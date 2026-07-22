# -*- coding: utf-8 -*-
"""
sinh_hieu.py — Pipeline 4: nhập nhanh sinh hiệu (§6.4 SPEC) + import Excel.

UX 6 ô/người (criterion P4.1): chiều cao, cân nặng, mạch, huyết áp, thị
lực, thính lực. Schema ho_so có RIÊNG cột mắt trái/phải và tai trái/phải
(nói thường/nói thầm) — §6.4 chỉ yêu cầu "6 ô/người" nên trường `thi_luc`
được ghi ĐỒNG THỜI vào khong_kinh_mat_phai + khong_kinh_mat_trai (2 mắt
cùng giá trị), và `thinh_luc` ghi vào tai_trai_noi_thuong + tai_phai_noi_thuong
(giữ nguyên 2 cột "nói thầm" — không thu thập trong màn hình nhanh này).
Đây là quyết định UX đơn giản hoá, ghi rõ trong PLAN.md/CRITERIA.md P4.

Cờ THIEU_SINH_HIEU (§4) được gỡ/tự thêm lại tự động (không cần xác nhận
thủ công như các cờ "suy luận" khác) khi đủ/thiếu 4 chỉ tiêu cốt lõi
(chieu_cao, can_nang, mach, huyet_ap) — ĐÚNG với luật lúc nạp
(import_data.py: cờ chỉ gắn khi CẢ 4 đều rỗng) và cùng cách xử lý đã áp
dụng cho THIEU_CCCD/so_cccd trong routers/ho_so.py:patch_ho_so().
"""
import io
import os
import re
import sys
from typing import Optional, Union

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402
import db  # noqa: E402
import auth  # noqa: E402
from services import fuzzy, qc, the_luc, sinh_hieu_valid  # noqa: E402
from routers.ho_so import build_where, TRANG_THAI_NHAN  # noqa: E402

config.ensure_build_on_path()
from build_import import clean_cccd, fmt_date  # noqa: E402

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

router = APIRouter(prefix='/api/sinh-hieu', tags=['sinh_hieu'])

# 6 trường nhập nhanh -> cột DB thật sự bị ảnh hưởng (criterion P4.1/P4.6)
FIELD_MAP = {
    'chieu_cao': ['chieu_cao'],
    'can_nang': ['can_nang'],
    'mach': ['mach'],
    'huyet_ap': ['huyet_ap'],
    'thi_luc': ['khong_kinh_mat_phai', 'khong_kinh_mat_trai'],
    'thinh_luc': ['tai_trai_noi_thuong', 'tai_phai_noi_thuong'],
}
CORE_VITAL_COLS = ('chieu_cao', 'can_nang', 'mach', 'huyet_ap')


def _du_sinh_hieu(row_like):
    """True nếu đủ 4 chỉ tiêu cốt lõi — nghịch đảo đúng luật rỗng dùng lúc
    nạp (import_data.py chỉ gắn cờ khi CẢ 4 đều None)."""
    def _get(c):
        return row_like[c] if c in row_like.keys() else None
    return all(_get(c) not in (None, '') for c in CORE_VITAL_COLS)


def _load_row(conn, ma_ho_so, user):
    row = conn.execute('SELECT * FROM ho_so WHERE ma_ho_so=?', (ma_ho_so,)).fetchone()
    if not row:
        return None, 'Không tìm thấy hồ sơ'
    # Đợt 2 criterion 4: hồ sơ CHƯA giao (nguoi_ra_soat_id IS NULL) mọi nhân viên
    # đều truy cập được; hồ sơ ĐÃ giao chỉ người được giao + admin.
    if (user['vai_tro'] != 'admin' and row['nguoi_ra_soat_id'] is not None
            and row['nguoi_ra_soat_id'] != user['id']):
        return None, 'Hồ sơ đã được giao cho nhân viên khác — không thuộc phạm vi rà soát của bạn'
    return row, None


def _apply_sinh_hieu(conn, row, changes_in, user_id):
    """Dùng CHUNG cho PATCH thủ công và import Excel — đảm bảo cả 2 đường
    đều ghi nhat_ky theo từng trường đổi + tự tính BMI + tự gỡ/thêm cờ
    THIEU_SINH_HIEU giống hệt nhau (criterion P4.3, P4.5, P4.6).

    Trả (updated: dict trường DB -> giá trị mới, error: str|None)."""
    ma_ho_so = row['ma_ho_so']
    db_changes = {}
    for ui_field, db_cols in FIELD_MAP.items():
        if ui_field in changes_in:
            for c in db_cols:
                db_changes[c] = changes_in[ui_field]

    if not db_changes:
        return {}, None

    merged = dict(row)
    merged.update(db_changes)
    if 'chieu_cao' in db_changes or 'can_nang' in db_changes:
        db_changes['chi_so_bmi'] = the_luc.bmi(merged.get('chieu_cao'), merged.get('can_nang'))

    updated = {}
    for field, new_val in db_changes.items():
        old_val = row[field] if field in row.keys() else None
        old_str = '' if old_val is None else str(old_val)
        new_str = '' if new_val is None else str(new_val)
        if old_str == new_str:
            continue
        conn.execute(f'UPDATE ho_so SET {field} = ? WHERE ma_ho_so = ?', (new_val, ma_ho_so))
        conn.execute(
            'INSERT INTO nhat_ky(ma_ho_so, nguoi_dung_id, ten_truong, '
            'gia_tri_cu, gia_tri_moi) VALUES (?,?,?,?,?)',
            (ma_ho_so, user_id, field, old_str, new_str))
        updated[field] = new_val

    if updated:
        conn.commit()

    # ---- tự gỡ/thêm cờ THIEU_SINH_HIEU (không cần xác nhận thủ công) ----
    new_row = conn.execute('SELECT * FROM ho_so WHERE ma_ho_so=?', (ma_ho_so,)).fetchone()
    before_co_qc = new_row['co_qc']
    if _du_sinh_hieu(new_row):
        after_co_qc = qc.remove_flags(conn, ma_ho_so, ['THIEU_SINH_HIEU'])
    else:
        after_co_qc = qc.add_flag(conn, ma_ho_so, 'THIEU_SINH_HIEU')
    if after_co_qc is not None and after_co_qc != before_co_qc:
        conn.execute(
            'INSERT INTO nhat_ky(ma_ho_so, nguoi_dung_id, ten_truong, '
            'gia_tri_cu, gia_tri_moi) VALUES (?,?,?,?,?)',
            (ma_ho_so, user_id, 'co_qc', before_co_qc or '', after_co_qc or ''))
        conn.commit()
        updated['co_qc'] = after_co_qc

    return updated, None


def _build_sinh_hieu_where(qp, user):
    """Dùng CHUNG cho /danh-sach và /mau-excel (criterion 3/5): xã multi +
    trạng thái + thiếu/đủ sinh hiệu + lọc ngày khám (Đợt 7 criterion 8, tái
    dùng _ymd/build_where của ho_so.py), cùng phạm vi vai trò của build_where()."""
    params = {
        'xa': qp.getlist('xa'), 'trang_thai': qp.getlist('trang_thai'),
        'ngay_tu': qp.get('ngay_tu'), 'ngay_den': qp.get('ngay_den'),
    }
    where_sql, args = build_where(params, user)
    sinh_hieu = qp.get('sinh_hieu')  # 'thieu' | 'du'
    if sinh_hieu == 'thieu':
        where_sql += " AND ((';'||co_qc||';') LIKE ?)"
        args.append('%;THIEU_SINH_HIEU;%')
    elif sinh_hieu == 'du':
        where_sql += " AND ((';'||co_qc||';') NOT LIKE ?)"
        args.append('%;THIEU_SINH_HIEU;%')
    return where_sql, args


@router.get('/danh-sach')
def danh_sach(request: Request, page: int = Query(1, ge=1),
              page_size: int = Query(50, ge=1, le=200),
              user=Depends(auth.get_current_user)):
    """Danh sách cho màn hình nhập nhanh — lọc: xã (multi) + trạng thái +
    thiếu/đủ sinh hiệu + họ tên (fuzzy, Đợt 3 criterion 5)."""
    qp = request.query_params
    conn = db.get_connection()
    try:
        where_sql, args = _build_sinh_hieu_where(qp, user)
        ho_ten_q = (qp.get('ho_ten') or '').strip()

        if ho_ten_q:
            # PLAN_PERF.md §2 — SQL-paginated bằng cột ho_ten_kd (đã tính
            # sẵn, bỏ dấu + lowercase) thay cho quét Python
            # (fuzzy.rank_by_name cũ, quét toàn bộ dòng đã lọc).
            q_kd = fuzzy.strip_diacritics(ho_ten_q)
            from routers.ho_so import _like_tokens
            like_sql, like_args = _like_tokens('ho_ten_kd', q_kd)
            where_q = f'{where_sql} AND {like_sql}'
            args_q = args + like_args
            total = conn.execute(
                f'SELECT COUNT(*) FROM ho_so WHERE {where_q}', args_q).fetchone()[0]
            offset = (page - 1) * page_size
            rows = conn.execute(
                f'SELECT * FROM ho_so WHERE {where_q} ORDER BY tt LIMIT ? OFFSET ?',
                args_q + [page_size, offset]).fetchall()
        else:
            total = conn.execute(f'SELECT COUNT(*) FROM ho_so WHERE {where_sql}', args).fetchone()[0]
            offset = (page - 1) * page_size
            rows = conn.execute(
                f'SELECT * FROM ho_so WHERE {where_sql} ORDER BY tt LIMIT ? OFFSET ?',
                args + [page_size, offset]).fetchall()

        items = []
        for r in rows:
            items.append({
                'ma_ho_so': r['ma_ho_so'],
                'ho_ten': r['ho_ten'],
                'nam_sinh': (r['ngay_sinh'] or '')[-4:] or None,
                'gioi_tinh': r['gioi_tinh'],
                'so_cccd': r['so_cccd'],
                'maxa_cu_tru': r['maxa_cu_tru'],
                'ngay_vao': r['ngay_vao'],
                'trang_thai': r['trang_thai'],
                'trang_thai_nhan': TRANG_THAI_NHAN.get(r['trang_thai'], r['trang_thai']),
                'chieu_cao': r['chieu_cao'],
                'can_nang': r['can_nang'],
                'chi_so_bmi': r['chi_so_bmi'],
                'mach': r['mach'],
                'huyet_ap': r['huyet_ap'],
                'kham_the_luc_pl': r['kham_the_luc_pl'],
                'thieu_sinh_hieu': not _du_sinh_hieu(r),
            })
    finally:
        conn.close()
    return {'total': total, 'page': page, 'page_size': page_size, 'items': items}


class PatchBody(BaseModel):
    model_config = ConfigDict(extra='forbid')
    # Đợt 6 criterion 1: Union[str, float] (KHÔNG chỉ float) — cho phép
    # chuỗi có dấu phẩy thập phân (vd '65,5') đi qua tới
    # sinh_hieu_valid.normalize_numeric_changes() bên dưới thay vì bị
    # pydantic tự chặn ngay ở tầng validate request (float('65,5') lỗi).
    chieu_cao: Optional[Union[str, float]] = None
    can_nang: Optional[Union[str, float]] = None
    mach: Optional[str] = None
    huyet_ap: Optional[str] = None


@router.patch('/{ma_ho_so}')
def patch_sinh_hieu(ma_ho_so: str, body: PatchBody, user=Depends(auth.get_current_user)):
    changes_in = body.model_dump(exclude_unset=True)
    conn = db.get_connection()
    try:
        row, err = _load_row(conn, ma_ho_so, user)
        if err:
            raise HTTPException(404 if 'Không tìm thấy' in err else 403, err)
        # Đợt 6 criterion 1: chuẩn hoá dấu thập phân (','->'.') TRƯỚC khi
        # validate ngưỡng — vd can_nang '65,5' -> '65.5'.
        changes_in, loi_chuan_hoa = sinh_hieu_valid.normalize_numeric_changes(changes_in)
        if loi_chuan_hoa:
            raise HTTPException(422, '; '.join(x['ly_do'] for x in loi_chuan_hoa))
        # Ngưỡng sinh hiệu (Đợt 3 criterion 2): ngoài ngưỡng -> 422; huyet_ap
        # được chuẩn hoá trước khi lưu (vd '12080' -> '120/80'), giá trị
        # chuẩn hoá sẽ xuất hiện trong `updated` của response bên dưới.
        nguong = sinh_hieu_valid.load_nguong(conn)
        changes_in, loi_nguong = sinh_hieu_valid.validate_changes(changes_in, nguong)
        if loi_nguong:
            raise HTTPException(422, '; '.join(x['ly_do'] for x in loi_nguong))

        # Đợt 5 criterion 2/3: chiều cao/cân nặng nằm trong bộ trường GỬI LÊN
        # (không phải trong FIELD_MAP đã dịch cột DB) -> tự tính lại PL thể
        # lực sau khi áp dụng; PATCH mach/huyet_ap solo không kích hoạt vì
        # the_luc chỉ dùng chiều cao/cân nặng.
        pl_can_tinh_lai = 'chieu_cao' in changes_in or 'can_nang' in changes_in
        pl_truoc = row['kham_the_luc_pl']

        updated, _ = _apply_sinh_hieu(conn, row, changes_in, user['id'])

        if pl_can_tinh_lai:
            # pl_moi = None nghĩa là PL bị XÓA (hết đủ dữ liệu) — vẫn báo về UI
            pl_moi = the_luc.tinh_va_ap_pl(conn, ma_ho_so, user['id'])
            if pl_moi != pl_truoc:
                updated['kham_the_luc_pl'] = pl_moi

        new_row = conn.execute('SELECT * FROM ho_so WHERE ma_ho_so=?', (ma_ho_so,)).fetchone()
    finally:
        conn.close()
    return {
        'ok': True, 'updated': updated, 'chi_so_bmi': new_row['chi_so_bmi'],
        'co_qc': qc.flags_of(new_row['co_qc']),
        'thieu_sinh_hieu': not _du_sinh_hieu(new_row),
        # Đợt 5 criterion 2: giá trị PL HIỆN TẠI (dù tự đổi lần này hay không)
        # để lưới sinh hiệu luôn đồng bộ đúng cột "Phân loại thể lực".
        'kham_the_luc_pl': new_row['kham_the_luc_pl'],
    }


# ============================= EXCEL =============================

# Đợt 3 criterion 3/4: cột định danh (khoá, chỉ đọc) + 4 cột sinh hiệu cần
# điền. KHÔNG còn cột thị lực/thính lực (bỏ theo phản hồi — "phải nhập chi
# tiết trong từng case").
TEMPLATE_HEADERS = ['MA_HO_SO', 'HO_TEN', 'SO_CCCD', 'NGAY_KHAM', 'XA_PHUONG',
                     'CHIEU_CAO', 'CAN_NANG', 'MACH', 'HUYET_AP']
IDENTITY_COL_COUNT = 5  # 5 cột đầu chỉ đọc/đối chiếu

# chuẩn hoá header (bỏ dấu, hoa, bỏ khoảng trắng/gạch) -> tên trường nội bộ,
# chấp nhận cả header tiếng Việt có dấu lẫn mã cột kiểu SO_CCCD (§4 P4.2/P4.7
# "flexible header matching"). THI_LUC/THINH_LUC vẫn nhận diện được (file
# mẫu cũ) nhưng bị BỎ QUA khi import (criterion 4).
_HEADER_SYNONYMS = {
    'MAHOSO': 'ma_ho_so',
    'SOCCCD': 'so_cccd', 'SOCANCUOC': 'so_cccd', 'CCCD': 'so_cccd',
    'SOCCCDCANCUOC': 'so_cccd',
    'HOTEN': 'ho_ten', 'HOVATEN': 'ho_ten',
    'NGAYKHAM': 'ngay_kham', 'NGAYVAO': 'ngay_kham',
    'XAPHUONG': 'xa_phuong',
    'CHIEUCAO': 'chieu_cao', 'CHIEUCAOCM': 'chieu_cao',
    'CANNANG': 'can_nang', 'CANNANGKG': 'can_nang',
    'MACH': 'mach',
    'HUYETAP': 'huyet_ap',
    'THILUC': 'thi_luc',      # deprecated — chỉ để báo "cột bỏ qua"
    'THINHLUC': 'thinh_luc',  # deprecated — chỉ để báo "cột bỏ qua"
}
_DEPRECATED_FIELDS = {'thi_luc', 'thinh_luc'}


def _norm_header(s):
    s = fuzzy.strip_diacritics(str(s or '')).upper()
    return re.sub(r'[^A-Z0-9]', '', s)


@router.get('/mau-excel')
def mau_excel(request: Request, user=Depends(auth.get_current_user)):
    """Sinh file .xlsx mẫu ĐÃ CÓ SẴN danh sách BN đang khớp bộ lọc hiện tại
    (Đợt 3 criterion 3) — 1 dòng/BN, 5 cột định danh (chỉ đọc, tô xám) + 4
    cột sinh hiệu để trống chờ điền."""
    qp = request.query_params
    conn = db.get_connection()
    try:
        where_sql, args = _build_sinh_hieu_where(qp, user)
        ho_ten_q = (qp.get('ho_ten') or '').strip()
        rows = conn.execute(
            f'SELECT * FROM ho_so WHERE {where_sql} ORDER BY tt', args).fetchall()
        if ho_ten_q:
            rows = fuzzy.rank_by_name(rows, ho_ten_q, limit=50)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'sinh_hieu'
    ws.append(TEMPLATE_HEADERS)

    bold = Font(bold=True)
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    for cell in ws[1]:
        cell.font = bold
    ws.cell(row=1, column=1).comment = Comment(
        'Chỉ điền 4 cột sinh hiệu: CHIEU_CAO, CAN_NANG, MACH, HUYET_AP.\n'
        '5 cột đầu (MA_HO_SO, HO_TEN, SO_CCCD, NGAY_KHAM, XA_PHUONG) dùng để '
        'đối chiếu — KHÔNG sửa.', 'KSK NCT')

    for r in rows:
        ws.append([r['ma_ho_so'], r['ho_ten'], r['so_cccd'], r['ngay_vao'],
                   r['maxa_cu_tru'], None, None, None, None])

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, IDENTITY_COL_COUNT + 1):
            ws.cell(row=row_idx, column=col_idx).fill = gray_fill

    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {'Content-Disposition': 'attachment; filename="mau_sinh_hieu.xlsx"'}
    return StreamingResponse(
        buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers)


def _match_ho_so(conn, ma_ho_so_raw, so_cccd_raw, ho_ten_raw, ngay_kham_raw):
    """Trả (row|None, cach_khop: 'ma_ho_so'|'cccd'|'ten_ngay'|None,
    ly_do_that_bai: str|None). Ưu tiên (criterion 4): MA_HO_SO chính xác ->
    SO_CCCD -> họ tên+ngày khám (waterfall — MA_HO_SO trống/không khớp thì
    thử tiếp CCCD, CCCD trống/không khớp thì thử tiếp tên+ngày)."""
    ma = str(ma_ho_so_raw).strip() if ma_ho_so_raw not in (None, '') else ''
    if ma:
        row = conn.execute('SELECT * FROM ho_so WHERE ma_ho_so=?', (ma,)).fetchone()
        if row:
            return row, 'ma_ho_so', None

    cccd_norm = None
    if so_cccd_raw not in (None, ''):
        cccd_norm, _err = clean_cccd(so_cccd_raw)
    if cccd_norm and len(cccd_norm) == 12:
        row = conn.execute('SELECT * FROM ho_so WHERE so_cccd=?', (cccd_norm,)).fetchone()
        if row:
            return row, 'cccd', None

    ho_ten_q = fuzzy.strip_diacritics(str(ho_ten_raw or '').strip())
    ngay_kham = fmt_date(ngay_kham_raw) if ngay_kham_raw not in (None, '') else ''
    if ho_ten_q and ngay_kham:
        candidates = conn.execute(
            'SELECT * FROM ho_so WHERE ngay_vao=?', (ngay_kham,)).fetchall()
        matches = [r for r in candidates
                   if fuzzy.strip_diacritics((r['ho_ten'] or '').strip()) == ho_ten_q]
        if len(matches) == 1:
            return matches[0], 'ten_ngay', None
        if len(matches) > 1:
            return None, None, (f"Khớp {len(matches)} hồ sơ trùng họ tên "
                                 f"'{ho_ten_raw}' + ngày khám '{ngay_kham}' — "
                                 'cần đối chiếu tay')

    ly_do_parts = []
    if ma:
        ly_do_parts.append(f"MA_HO_SO '{ma}' không tồn tại")
    if cccd_norm and len(cccd_norm) != 12:
        ly_do_parts.append(f"SO_CCCD '{so_cccd_raw}' không hợp lệ (không đủ 12 số)")
    elif cccd_norm:
        ly_do_parts.append(f"SO_CCCD '{so_cccd_raw}' không khớp hồ sơ nào")
    if not (ho_ten_q and ngay_kham):
        ly_do_parts.append('thiếu/không đủ dữ liệu họ tên + ngày khám để khớp dự phòng')
    if not ly_do_parts:
        ly_do_parts.append('thiếu MA_HO_SO/SO_CCCD/họ tên+ngày khám để khớp')
    return None, None, 'Không khớp được hồ sơ: ' + '; '.join(ly_do_parts)


@router.post('/import-excel')
async def import_excel(file: UploadFile = File(...), user=Depends(auth.get_current_user)):
    """Import Excel (Đợt 3 criterion 4): khớp ưu tiên MA_HO_SO -> SO_CCCD ->
    họ tên+ngày khám; chỉ đọc 4 trường sinh hiệu (thị lực/thính lực bị bỏ
    qua nếu còn trong file cũ); validate ngưỡng từng dòng — dòng có BẤT KỲ
    trường nào ngoài ngưỡng bị TỪ CHỐI TOÀN BỘ DÒNG (không áp dụng phần
    hợp lệ), lý do liệt kê rõ trong báo cáo."""
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f'Không đọc được file Excel: {e}')
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        raise HTTPException(400, 'File rỗng — không có dòng tiêu đề')
    col_field = {}
    cot_bo_qua = set()
    for cell in header_row:
        canon = _HEADER_SYNONYMS.get(_norm_header(cell.value))
        if not canon:
            continue
        if canon in _DEPRECATED_FIELDS:
            cot_bo_qua.add(str(cell.value))
            continue
        col_field[cell.column] = canon
    field_names = set(col_field.values())
    if not (('ma_ho_so' in field_names) or ('so_cccd' in field_names)
            or ('ho_ten' in field_names and 'ngay_kham' in field_names)):
        raise HTTPException(
            400, 'File thiếu cột bắt buộc: cần MA_HO_SO, hoặc SO_CCCD, hoặc '
                 f'HO_TEN + NGAY_KHAM (cột nhận diện được: {sorted(field_names)})')

    tong_dong = 0
    khop_ma_ho_so = 0
    khop_cccd = 0
    khop_ten_ngay = 0
    khong_khop = []
    loi_nguong = []
    chi_tiet = []

    conn = db.get_connection()
    try:
        nguong = sinh_hieu_valid.load_nguong(conn)

        for row_cells in ws.iter_rows(min_row=2):
            values = {}
            for cell in row_cells:
                field = col_field.get(cell.column)
                if field:
                    values[field] = cell.value
            if not any(v not in (None, '') for v in values.values()):
                continue  # dòng trắng thật sự — không tính vào tong_dong
            tong_dong += 1
            dong_so = row_cells[0].row

            matched_row, cach_khop, ly_do = _match_ho_so(
                conn, values.get('ma_ho_so'), values.get('so_cccd'),
                values.get('ho_ten'), values.get('ngay_kham'))

            if not matched_row:
                khong_khop.append({'dong': dong_so, 'ly_do': ly_do})
                continue

            if (user['vai_tro'] != 'admin' and matched_row['nguoi_ra_soat_id'] is not None
                    and matched_row['nguoi_ra_soat_id'] != user['id']):
                khong_khop.append({
                    'dong': dong_so,
                    'ly_do': (f"Khớp hồ sơ {matched_row['ma_ho_so']} nhưng đã được giao "
                              'cho nhân viên khác — ngoài phạm vi rà soát của bạn'),
                })
                continue

            changes_in = {}
            for ui_field in CORE_VITAL_COLS:  # chieu_cao, can_nang, mach, huyet_ap
                if ui_field in values and values[ui_field] not in (None, ''):
                    changes_in[ui_field] = values[ui_field]

            if changes_in:
                # Đợt 6 criterion 1: chuẩn hoá dấu thập phân từng ô số TRƯỚC
                # khi validate ngưỡng — openpyxl có thể trả cả số (float) lẫn
                # chuỗi (ô định dạng Text, vd '6,7') tuỳ cách người dùng nhập.
                changes_in, loi_chuan_hoa = sinh_hieu_valid.normalize_numeric_changes(changes_in)
                if loi_chuan_hoa:
                    ly_do_gop = '; '.join(f"{x['field']}: {x['ly_do']}" for x in loi_chuan_hoa)
                    loi_nguong.append({
                        'dong': dong_so, 'ma_ho_so': matched_row['ma_ho_so'],
                        'ly_do': ly_do_gop,
                    })
                    continue  # từ chối TOÀN BỘ dòng (criterion 4)
                normalized, loi = sinh_hieu_valid.validate_changes(changes_in, nguong)
                if loi:
                    ly_do_gop = '; '.join(f"{x['field']}: {x['ly_do']}" for x in loi)
                    loi_nguong.append({
                        'dong': dong_so, 'ma_ho_so': matched_row['ma_ho_so'],
                        'ly_do': ly_do_gop,
                    })
                    continue  # từ chối TOÀN BỘ dòng (criterion 4)
                changes_in = normalized

            updated, _ = _apply_sinh_hieu(conn, matched_row, changes_in, user['id'])

            if cach_khop == 'ma_ho_so':
                khop_ma_ho_so += 1
            elif cach_khop == 'cccd':
                khop_cccd += 1
            else:
                khop_ten_ngay += 1
            chi_tiet.append({
                'dong': dong_so, 'ma_ho_so': matched_row['ma_ho_so'],
                'cach_khop': cach_khop, 'cap_nhat': updated,
            })
    finally:
        conn.close()

    return {
        'tong_dong': tong_dong,
        'khop_ma_ho_so': khop_ma_ho_so,
        'khop_cccd': khop_cccd,
        'khop_ten_ngay': khop_ten_ngay,
        'so_khong_khop': len(khong_khop),
        'khong_khop': khong_khop,
        'so_loi_nguong': len(loi_nguong),
        'loi_nguong': loi_nguong,
        'cot_bo_qua': sorted(cot_bo_qua),
        'chi_tiet': chi_tiet,
    }
