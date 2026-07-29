# -*- coding: utf-8 -*-
"""
xuat_file.py — Pipeline 2: xuất file .xlsm nộp Bộ Y tế (§7 SPEC), admin-only.

Job chạy NỀN (thread) trong services/export_xlsm.py — mỗi xã lại là một
tiến trình con riêng (services/export_worker.py) để tránh tích luỹ bộ nhớ
openpyxl (bẫy §10). Endpoint khởi tạo job rồi trả job_id ngay; frontend
polling GET /api/xuat-file/jobs/{id} để theo dõi tiến độ.
"""
import os
import sys
from typing import List, Optional

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db  # noqa: E402
import auth  # noqa: E402
from services import export_xlsm, nhap_doi_soat, ghi_chu_ra_soat, auto_backup  # noqa: E402

from urllib.parse import quote

from fastapi import (APIRouter, Depends, File, Form, HTTPException, Query,
                     UploadFile)
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel

router = APIRouter(prefix='/api', tags=['xuat_file'])


class ExtendedOpt(BaseModel):
    enabled: bool = False
    columns: List[str] = []


class PreviewBody(BaseModel):
    pham_vi: str
    gia_tri: List[str] = []
    include_errors: bool = False
    # chỉ xuất hồ sơ đã tick ĐỦ 4 mục "đã rà soát xong" ở panel chi tiết
    chi_rs_xong: bool = False


class StartBody(PreviewBody):
    extended: Optional[ExtendedOpt] = None


@router.get('/xuat-file/cot-mo-rong')
def cot_mo_rong(admin=Depends(auth.require_admin)):
    """Danh mục cột mở rộng cho user tick chọn (§7.2)."""
    return [{'ma': c, 'ten': t} for c, t in export_xlsm.EXTENDED_COLUMNS]


@router.post('/xuat-file/preview')
def preview(body: PreviewBody, admin=Depends(auth.require_admin)):
    conn = db.get_connection()
    try:
        try:
            result = export_xlsm.preview(conn, body.pham_vi, body.gia_tri,
                                          body.include_errors, body.chi_rs_xong)
        except ValueError as e:
            raise HTTPException(400, str(e))
    finally:
        conn.close()
    return result


def _job_public(job):
    return {k: v for k, v in job.items() if k != 'job_dir'}


@router.post('/xuat-file/xlsx-chinh-sua')
def xuat_xlsx_chinh_sua(body: PreviewBody, admin=Depends(auth.require_admin)):
    """Xuất .xlsx như đơn thuần NHƯNG kèm cột MÃ ĐỊNH DANH (ma_ho_so) ở đầu —
    tải về, bổ sung/sửa dữ liệu, rồi nhập lại qua /nhap-doi-soat."""
    conn = db.get_connection()
    try:
        try:
            data, count = export_xlsm.build_plain_xlsx(
                conn, body.pham_vi, body.gia_tri, body.include_errors,
                body.chi_rs_xong, with_id=True)
        except ValueError as e:
            raise HTTPException(400, str(e))
    finally:
        conn.close()
    import datetime
    fn = f'KSK_ChinhSua_{count}ca_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx'
    return Response(
        content=data,
        media_type=('application/vnd.openxmlformats-officedocument'
                    '.spreadsheetml.sheet'),
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(fn)}"})


@router.post('/xuat-file/nhap-doi-soat')
async def nhap_doi_soat_ep(file: UploadFile = File(...),
                           ap_dung: bool = Form(False),
                           cho_ghi_de: bool = Form(False),
                           cot: str = Form(''),
                           sheet: str = Form(''),
                           ma_loai_tru: str = Form(''),
                           admin=Depends(auth.require_admin)):
    """Nhập lại file .xlsx đã chỉnh sửa, đối soát với DB. ap_dung=False -> chỉ
    XEM TRƯỚC (không ghi). ap_dung=True -> ghi (bổ sung luôn; ghi đè chỉ khi
    cho_ghi_de=True). `cot`: danh sách tên cột (field, chữ thường) cách nhau
    dấu phẩy — CHỈ đối soát/ghi các cột này; rỗng = tất cả cột phát hiện
    được. `sheet`: tên sheet cần đối soát — BẮT BUỘC nếu file có ≥2 sheet
    (phản hồi anh Khôi: trước đây tự đoán sheet, có thể đọc NHẦM sheet dẫn
    tới đối soát sai — vd ghi đè Ngày sinh thật thành 01/01 ước lượng).
    Chưa chỉ định mà file nhiều sheet -> trả 409 kèm danh sách sheet, để
    frontend hỏi user chọn rồi gọi lại. `ma_loai_tru`: danh sách mã hồ sơ
    (cách nhau dấu phẩy) bị bỏ tick ở checkbox từng DÒNG tại bước xem trước —
    CHỈ có tác dụng khi ap_dung=True (dòng đó không được ghi); các lời gọi
    xem trước không cần truyền tham số này. Trả tóm tắt + cot_phat_hien +
    mẫu chi tiết thay đổi."""
    content = await file.read()
    cot_chon = {c.strip() for c in cot.split(',') if c.strip()} if cot else None
    ma_loai_tru_set = ({m.strip() for m in ma_loai_tru.split(',') if m.strip()}
                       if ma_loai_tru else None)
    conn = db.get_connection()
    try:
        try:
            result = nhap_doi_soat.doi_soat(
                conn, content, apply=ap_dung, cho_ghi_de=cho_ghi_de,
                user_id=admin['id'], cot_chon=cot_chon, sheet=(sheet or None),
                ma_loai_tru=ma_loai_tru_set)
        except nhap_doi_soat.CanChonSheet as e:
            raise HTTPException(409, detail={
                'ly_do': 'can_chon_sheet',
                'sheet_list': e.sheet_names,
                'thong_bao': 'File có nhiều sheet — hãy chọn sheet cần đối soát.',
            })
        except ValueError as e:
            raise HTTPException(400, str(e))
    finally:
        conn.close()
    return result


@router.post('/xuat-file/ghi-chu-ra-soat')
async def ghi_chu_ra_soat_ep(file: UploadFile = File(...),
                             ap_dung: bool = Form(False),
                             admin=Depends(auth.require_admin)):
    """Nhập lại file "Đối soát" (xuất từ nút "Xuất Excel" của khối "Nhập lại
    file đã chỉnh sửa" — 9 cột: STT|Mã hồ sơ|Họ tên|Danh sách|Sẽ áp dụng|
    Loại thay đổi|Trường|Giá trị cũ|Giá trị mới, sheet 'Đối soát'), gộp các
    dòng thay đổi của TỪNG hồ sơ thành 1 đoạn ghi chú, APPEND vào cột
    ghi_chu_ra_soat. ap_dung=False -> chỉ XEM TRƯỚC (không ghi). ap_dung=True
    -> ghi thật (UPDATE + nhat_ky), commit 1 lần."""
    content = await file.read()
    conn = db.get_connection()
    try:
        try:
            result = ghi_chu_ra_soat.xu_ly(conn, content, apply=ap_dung,
                                           user_id=admin['id'])
        except ValueError as e:
            raise HTTPException(400, str(e))
    finally:
        conn.close()
    return result


@router.post('/xuat-file/doi-soat-xlsx')
async def doi_soat_xlsx_ep(file: UploadFile = File(...),
                           cot: str = Form(''),
                           sheet: str = Form(''),
                           ma_loai_tru: str = Form(''),
                           admin=Depends(auth.require_admin)):
    """Xuất TOÀN BỘ danh sách thay đổi phát hiện được ở bước đối soát (không
    giới hạn 200 hồ sơ như bảng xem trước trên màn hình) ra file .xlsx để
    kiểm tra kỹ hơn — CHỈ ĐỌC, không ghi gì vào DB (không có ap_dung/
    cho_ghi_de như /nhap-doi-soat). Mỗi DÒNG Excel = 1 THAY ĐỔI (1 hồ sơ có
    thể có nhiều thay đổi -> nhiều dòng). `ma_loai_tru` chỉ ảnh hưởng cột
    "Sẽ áp dụng" (tham chiếu checkbox đã bỏ tick ở bước xem trước), KHÔNG
    ảnh hưởng việc có liệt kê dòng đó hay không — file xuất luôn đầy đủ để
    người dùng kiểm tra hết, kể cả dòng đã định loại trừ."""
    content = await file.read()
    cot_chon = {c.strip() for c in cot.split(',') if c.strip()} if cot else None
    ma_loai_tru_set = ({m.strip() for m in ma_loai_tru.split(',') if m.strip()}
                       if ma_loai_tru else set())
    conn = db.get_connection()
    try:
        try:
            result = nhap_doi_soat.doi_soat(
                conn, content, apply=False, cho_ghi_de=False, user_id=None,
                cot_chon=cot_chon, sheet=(sheet or None),
                bo_gioi_han_chi_tiet=True)
        except nhap_doi_soat.CanChonSheet as e:
            raise HTTPException(409, detail={
                'ly_do': 'can_chon_sheet',
                'sheet_list': e.sheet_names,
                'thong_bao': 'File có nhiều sheet — hãy chọn sheet cần đối soát.',
            })
        except ValueError as e:
            raise HTTPException(400, str(e))
    finally:
        conn.close()

    import datetime
    import io
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Đối soát'
    headers = ['STT', 'Mã hồ sơ', 'Họ tên', 'Danh sách', 'Sẽ áp dụng',
               'Loại thay đổi', 'Trường', 'Giá trị cũ', 'Giá trị mới']
    bold = Font(bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = bold

    stt = 0
    for row in result['chi_tiet']:
        ma = row['ma_ho_so']
        ds = '; '.join(row.get('_danh_sach') or [])
        se_ap_dung = 'Không' if ma in ma_loai_tru_set else 'Có'
        for c in row['changes']:
            stt += 1
            r = stt + 1
            ws.cell(r, 1, stt)
            ws.cell(r, 2, ma)
            ws.cell(r, 3, row['ho_ten'])
            ws.cell(r, 4, ds)
            ws.cell(r, 5, se_ap_dung)
            ws.cell(r, 6, 'Bổ sung' if c['loai'] == 'bo_sung' else 'Ghi đè')
            ws.cell(r, 7, c['nhan'])
            ws.cell(r, 8, c['cu'])
            ws.cell(r, 9, c['moi'])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    data = buf.getvalue()

    fn = f'KSK_DoiSoat_{stt}thaydoi_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx'
    return Response(
        content=data,
        media_type=('application/vnd.openxmlformats-officedocument'
                    '.spreadsheetml.sheet'),
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(fn)}"})


@router.post('/xuat-file')
def start_export(body: StartBody, admin=Depends(auth.require_admin)):
    # Job xuất .xlsm chạy NỀN bằng subprocess + ghi job.json ra đĩa
    # (data/exports/) — không khả thi trên serverless (mỗi request là 1
    # container riêng, không có tiến trình nền/đĩa bền). Khóa theo biến
    # VERCEL (chỉ có trên bản đám mây), KHÔNG khóa theo TURSO_URL: máy local
    # vẫn có thể set TURSO_URL để NỐI DB online rồi xuất .xlsm chính thức.
    if os.getenv('VERCEL'):
        raise HTTPException(
            409, 'Xuất file .xlsm chưa hỗ trợ trên bản chạy đám mây — hãy '
                 'chạy app trên máy cá nhân (xem hướng dẫn câu lệnh ở trang '
                 'này), hoặc dùng nút "Tải .xlsx (kèm mã định danh)" ngay '
                 'tại đây.')
    extended = body.extended.model_dump() if body.extended else {'enabled': False, 'columns': []}
    try:
        job = export_xlsm.create_job(body.pham_vi, body.gia_tri,
                                      body.include_errors, extended, admin['id'],
                                      body.chi_rs_xong)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return _job_public(job)


@router.get('/xuat-file/jobs')
def list_jobs(admin=Depends(auth.require_admin)):
    return [_job_public(j) for j in export_xlsm.list_jobs()]


@router.get('/xuat-file/jobs/{job_id}')
def get_job(job_id: str, admin=Depends(auth.require_admin)):
    job = export_xlsm.get_job(job_id)
    if not job:
        raise HTTPException(404, 'Không tìm thấy job')
    return _job_public(job)


@router.get('/xuat-file/download')
def download(path: str = Query(...), admin=Depends(auth.require_admin)):
    """Chỉ cho tải file bên trong app/data/exports/ — chặn path traversal."""
    exports_dir = os.path.realpath(export_xlsm.EXPORTS_DIR)
    real = os.path.realpath(path)
    if not (real == exports_dir or real.startswith(exports_dir + os.sep)):
        raise HTTPException(403, 'Đường dẫn không hợp lệ')
    if not os.path.isfile(real):
        raise HTTPException(404, 'Không tìm thấy file')
    return FileResponse(real, filename=os.path.basename(real))


# ---- Đợt 16 (phản hồi anh Khôi): backup thủ công + liệt kê + khôi phục ----
@router.get('/backup/danh-sach')
def backup_danh_sach(admin=Depends(auth.require_admin)):
    """Liệt kê mọi bản sao lưu local (tự động + thủ công + hằng ngày...),
    mới nhất trước."""
    return auto_backup.list_backups()


@router.post('/backup/tao-thu-cong')
def backup_tao_thu_cong(admin=Depends(auth.require_admin)):
    """Sao lưu NGAY LẬP TỨC theo yêu cầu tay — lưu riêng thư mục 'manual'
    (không bị dọn bớt theo số lượng như bản tự động)."""
    duong_dan = auto_backup.backup_now(sub='manual')
    return {'duong_dan': os.path.basename(duong_dan)}


class KhoiPhucBody(BaseModel):
    duong_dan: str


@router.post('/backup/khoi-phuc')
def backup_khoi_phuc(body: KhoiPhucBody, admin=Depends(auth.require_admin)):
    """Khôi phục DB sống từ 1 bản backup đã liệt kê — GHI ĐÈ dữ liệu hiện
    tại của TOÀN BỘ hệ thống (ảnh hưởng mọi người dùng). Luôn tự tạo 1 bản
    an toàn của DB hiện tại trước khi ghi đè (thư mục before_restore/)."""
    try:
        safety = auto_backup.restore_backup(body.duong_dan)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {'ok': True, 'ban_an_toan_truoc_khoi_phuc': os.path.basename(safety)}
