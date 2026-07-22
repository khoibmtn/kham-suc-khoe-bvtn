# -*- coding: utf-8 -*-
"""
import_data.py — nạp dữ liệu KSK NCT vào SQLite (app/data/ksk.db).

Nguồn:
  - output/KSK_DuLieuQuanLy_TOANBO.xlsx  (sheet BENH_NHAN, BENH_CHI_TIET,
    PHAN_LOAI_CO_QUAN) — dữ liệu đã được pipeline build/ chuẩn hóa sẵn.
  - doc/Import_KSK_Tren 18.xlsm          (sheet dmicdme, dmtinh, dmdantoc,
    dmnghenghiep, dmkhac, 'Hướng dẫn')  — danh mục dropdown của Bộ Y tế.

KHÔNG viết lại pipeline chuẩn hóa: gọi thẳng build/classify.py (ORGAN_COLS)
và build/tien_su.py (suy_luan) để suy ra các trường tiền sử + khám cơ quan,
gọi build/build_import.py (CFG, ten_chinh_thuc) cho các trường hằng số và
tên bệnh chính thức.

IDEMPOTENT: chạy lại nhiều lần không đổi kết quả. Lựa chọn — bỏ qua
(INSERT OR IGNORE / skip) các `ma_ho_so` ĐÃ có trong bảng ho_so. Vì nhân viên
có thể đã sửa các trường rà soát (trang_thai, co_qc đã gỡ cờ, ghi_chu...),
KHÔNG ghi đè các hồ sơ đã tồn tại — chỉ nạp bản ghi MỚI (theo khóa
ma_ho_so). Danh mục (dm_icd, danh_muc) không phải dữ liệu người dùng sửa
nên được xóa & nạp lại mỗi lần chạy để luôn khớp với file nguồn mới nhất.
"""
import collections
import hashlib
import os
import re
import secrets
import sys
import time

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config  # noqa: E402
import db  # noqa: E402
from services import fuzzy  # noqa: E402

config.ensure_build_on_path()
from classify import ORGAN_COLS, ORGANS          # noqa: E402
import tien_su                                    # noqa: E402
from build_import import CFG, ten_chinh_thuc      # noqa: E402
from build_xlsm import TEN_CQ                     # noqa: E402

TEN_CQ_REV = {v: k for k, v in TEN_CQ.items()}    # tên đầy đủ -> mã cơ quan


# ============================= TIỆN ÍCH ĐỌC XLSX =============================
def _sheet_rows(path, sheet, **kw):
    """Đọc 1 sheet -> list[dict] theo header dòng 1. read_only để tiết kiệm RAM."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, **kw)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else '' for h in next(it)]
    rows = []
    for r in it:
        rows.append(dict(zip(header, r)))
    wb.close()
    return rows


def _cell_range(ws, col_letter, row_start, row_end):
    from openpyxl.utils import column_index_from_string
    ci = column_index_from_string(col_letter)
    out = []
    for r in range(row_start, row_end + 1):
        v = ws.cell(r, ci).value
        if v is not None and str(v).strip() != '':
            out.append(v)
    return out


# ============================= DANH MỤC (§1.2) =============================
def load_danh_muc(conn):
    """Nạp danh mục dropdown từ doc/Import_KSK_Tren 18.xlsm — xóa & nạp lại
    mỗi lần chạy (không phải dữ liệu người dùng sửa)."""
    wb = openpyxl.load_workbook(config.CATALOG_XLSM, read_only=True,
                                 data_only=True, keep_vba=False)
    rows = []  # (loai, ma, ten, thu_tu)

    # ---- dmtinh: header (TENTINH, TINH_ID, MATINH) ----
    ws = wb['dmtinh']
    it = ws.iter_rows(min_row=2, values_only=True)
    for i, r in enumerate(it):
        if r[0]:
            rows.append(('dmtinh', str(r[2]) if r[2] is not None else None,
                         str(r[0]).strip(), i))

    # ---- dmdantoc: header (Mã, Tên, Tên gọi khác) — CHỈ lấy các dòng có
    # cột Mã là số nguyên (1..56); bỏ 4 dòng cuối trang bị lệch cấu trúc
    # (cột A là tên, cột B là mã dạng chuỗi khác định dạng — dữ liệu rác
    # còn sót lại trong file mẫu gốc của Bộ Y tế).
    ws = wb['dmdantoc']
    it = ws.iter_rows(min_row=2, values_only=True)
    for r in it:
        if isinstance(r[0], int):
            rows.append(('dmdantoc', str(r[0]), str(r[1]).strip(), r[0]))

    # ---- dmnghenghiep: header (NGHE_NGHIEP, NGHENGHIEP_MA, NGHE_NGHIEP_ID) ----
    ws = wb['dmnghenghiep']
    it = ws.iter_rows(min_row=2, values_only=True)
    for i, r in enumerate(it):
        if r[0]:
            rows.append(('dmnghenghiep', str(r[1]) if r[1] is not None else None,
                         str(r[0]).strip(), i))

    # ---- dmkhac: các list nhỏ theo vùng ô cố định ----
    ws = wb['dmkhac']
    for i, v in enumerate(_cell_range(ws, 'AL', 2, 4)):
        rows.append(('gioi_tinh', None, str(v).strip(), i))
    for i, v in enumerate(_cell_range(ws, 'BF', 2, 12)):
        rows.append(('thi_luc', None, str(v).strip(), i))
    for i, v in enumerate(_cell_range(ws, 'BJ', 1, 16)):
        rows.append(('loai_kcb', None, str(v).strip(), i))

    # ---- Hướng dẫn: đối tượng khám + nguồn kinh phí ----
    ws = wb['Hướng dẫn']
    for i, v in enumerate(_cell_range(ws, 'B', 9, 24)):
        rows.append(('doi_tuong', None, str(v).strip(), i))
    for i, v in enumerate(_cell_range(ws, 'G', 8, 13)):
        rows.append(('nguon_kinh_phi', None, str(v).strip(), i))

    wb.close()

    conn.execute('DELETE FROM danh_muc')
    conn.executemany(
        'INSERT INTO danh_muc(loai, ma, ten, thu_tu) VALUES (?,?,?,?)', rows)
    conn.commit()
    return collections.Counter(r[0] for r in rows)


def load_dm_icd(conn):
    """Nạp danh mục ICD (dmicdme) -> dm_icd + đánh chỉ mục lại dm_icd_fts.

    dmicdme dùng quy ước dagger/asterisk (vd 'E11.4†'): dm_icd.ma giữ
    NGUYÊN VĂN, dm_icd.ma_tran bỏ hậu tố -> tránh bẫy tra ICD trượt (§10).
    File mẫu lặp lại khối ICD ~3 lần (phục vụ nhiều cột dropdown trong cùng
    workbook) -> dùng INSERT OR IGNORE để khử trùng theo khóa `ma`.
    """
    wb = openpyxl.load_workbook(config.CATALOG_XLSM, read_only=True,
                                 data_only=True, keep_vba=False)
    ws = wb['dmicdme']
    n_raw = 0
    conn.execute('DELETE FROM dm_icd')
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0] or not r[1]:
            continue
        n_raw += 1
        ma = str(r[0]).strip()
        ten = str(r[1]).strip()
        ma_tran = re.sub(r'[†*✝\s]+$', '', ma).strip()
        rows.append((ma, ma_tran, ten))
    wb.close()

    conn.executemany(
        'INSERT OR IGNORE INTO dm_icd(ma, ma_tran, ten) VALUES (?,?,?)', rows)
    conn.commit()
    # đánh lại chỉ mục FTS5 từ nội dung bảng gốc (external content table)
    conn.execute("INSERT INTO dm_icd_fts(dm_icd_fts) VALUES('rebuild')")
    conn.commit()
    n_loaded = conn.execute('SELECT COUNT(*) FROM dm_icd').fetchone()[0]
    return n_raw, n_loaded


# ============================= NGƯỜI DÙNG MẪU =============================
def _pbkdf2(password, salt=None, iterations=200_000):
    salt = salt or secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'),
                            bytes.fromhex(salt), iterations)
    return f'pbkdf2${iterations}${salt}${h.hex()}'


def seed_users(conn):
    users = [
        ('admin', 'Quản trị viên', 'admin', 'admin123'),
        ('raso1', 'Nhân viên rà soát 1', 'ra_soat', 'raso123'),
    ]
    for ten_dn, ho_ten, vai_tro, pw in users:
        conn.execute(
            'INSERT OR IGNORE INTO nguoi_dung'
            '(ten_dang_nhap, ho_ten, vai_tro, mat_khau_hash) VALUES (?,?,?,?)',
            (ten_dn, ho_ten, vai_tro, _pbkdf2(pw)))
    conn.commit()


# ============================= HỒ SƠ + BỆNH =============================
def _s(v):
    """Chuẩn hóa chuỗi rỗng -> None (giữ nguyên None đầu vào)."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s != '' else None


def build_ho_so_record(bn):
    """bn: dict 1 dòng sheet BENH_NHAN -> dict cột bảng ho_so (chưa gồm
    tiền sử suy luận và khám cơ quan — merge thêm sau)."""
    co_quan_bc_ten = _s(bn.get('CO_QUAN_BENH_CHINH'))
    rec = {
        'ma_ho_so': bn['MA_HO_SO'],
        'tt': bn.get('TT'),
        'ma_cskcb': CFG['MA_CSKCB'],
        'ma_gtin_cskcb': _s(CFG.get('MA_GTIN_CSKCB')),
        'ngay_vao': _s(bn.get('NGAY_KHAM')),
        'ho_ten': _s(bn.get('HO_TEN')),
        'gioi_tinh': _s(bn.get('GIOI_TINH')),
        'ngay_sinh': _s(bn.get('NGAY_SINH')),
        'ma_dan_toc': CFG['DAN_TOC'],
        'so_cccd': _s(bn.get('SO_CCCD')),
        'ngaycap_cccd': _s(bn.get('NGAYCAP_CCCD')),
        'noicap_cccd': _s(bn.get('NOICAP_CCCD')),
        'matinh_cu_tru': CFG['TINH'],
        'maxa_cu_tru': config.XA_DOI_TEN.get(_s(bn.get('XA_PHUONG')), _s(bn.get('XA_PHUONG'))),
        'dia_chi': _s(bn.get('THON_TDP')),
        'ma_nghe_nghiep': CFG['NGHE_NGHIEP'],
        'nhom_mau': None,
        'doi_tuong': _s(bn.get('DOI_TUONG')) or CFG['DOI_TUONG'],
        'nguon_chi_tra': CFG['NGUON_CHI_TRA'],
        'noi_lam_viec_hoc_tap': None,
        'dien_thoai': None,
        'ly_do_vv': CFG['LY_DO'],
        'ma_loai_kcb': CFG['LOAI_KCB'],
        # thể lực — §6.4: chưa có dữ liệu, để trống
        'chieu_cao': None, 'can_nang': None, 'chi_so_bmi': None,
        'mach': None, 'huyet_ap': None, 'kham_the_luc_pl': None,
        # thị lực (data), thính lực/hàm (trống — §6.4)
        'khong_kinh_mat_phai': _s(bn.get('THI_LUC_MAT_PHAI')),
        'khong_kinh_mat_trai': _s(bn.get('THI_LUC_MAT_TRAI')),
        'co_kinh_mat_phai': None, 'co_kinh_mat_trai': None,
        'tai_trai_noi_thuong': None, 'tai_trai_noi_tham': None,
        'tai_phai_noi_thuong': None, 'tai_phai_noi_tham': None,
        'ham_tren': None, 'ham_duoi': None,
        'phan_loai_sk': bn.get('PHAN_LOAI_SK'),
        'ket_luan_benh': _s(bn.get('TEN_BENH_CHINH')),
        'cac_benh_tat_neu_co': _s(bn.get('CHAN_DOAN_GOC')),
        # mở rộng
        'ma_benh_chinh': _s(bn.get('ICD_BENH_CHINH')),
        'co_quan_benh_chinh': TEN_CQ_REV.get(co_quan_bc_ten) if co_quan_bc_ten else None,
        'ma_benh_kem': _s(bn.get('ICD_BENH_KEM')),
        'ten_benh_kem': _s(bn.get('TEN_BENH_KEM')),
        'nam_sinh_nguon': bn.get('NAM_SINH'),
        'tuoi': bn.get('TUOI'),
        'glu_thoi_diem': _s(bn.get('GLU_THOI_DIEM')),
        'glu_gia_tri': bn.get('GLU_GIA_TRI'),
        'kq_dien_tim': _s(bn.get('KQ_DIEN_TIM')),
        'kq_sieu_am_o_bung': _s(bn.get('KQ_SIEU_AM_O_BUNG')),
        'chan_doan_goc': _s(bn.get('CHAN_DOAN_GOC')),
        # quản lý rà soát
        'nguoi_ra_soat_id': None,
        'trang_thai': 'chua_ra_soat',
        'co_qc': _s(bn.get('CO_QC')) or '',
        'so_loi': 0,
        'ghi_chu_ra_soat': _s(bn.get('GHI_CHU_RA_SOAT')),
        'ghi_chu_can_bo': None,
        'thoi_diem_hoan_thanh': None,
        'da_xuat_file': 0,
        'lan_xuat_cuoi': None,
    }
    # mặc định 14 cơ quan = None/None (sẽ điền từ PHAN_LOAI_CO_QUAN)
    for code, (col_txt, col_pl) in ORGAN_COLS.items():
        rec[col_txt.lower()] = None
        rec[col_pl.lower()] = None
    return rec


def apply_phan_loai(rec, plr_rows):
    """plr_rows: list dict sheet PHAN_LOAI_CO_QUAN của 1 người
    (MA_CO_QUAN, CO_QUAN, PHAN_LOAI, KET_QUA_KHAM)."""
    for r in plr_rows:
        code = r.get('MA_CO_QUAN')
        if code not in ORGAN_COLS:
            continue
        col_txt, col_pl = ORGAN_COLS[code]
        rec[col_txt.lower()] = _s(r.get('KET_QUA_KHAM'))
        pl = r.get('PHAN_LOAI')
        rec[col_pl.lower()] = int(pl) if pl not in (None, '') else None


def build_findings(benh_rows):
    """benh_rows: list dict sheet BENH_CHI_TIET của 1 người ->
    (findings cho tien_su.suy_luan, bc: finding của bệnh chính hoặc None)."""
    findings = []
    bc = None
    for r in benh_rows:
        f = {
            'icd': _s(r.get('MA_ICD')) or '',
            'ten_icd': _s(r.get('TEN_ICD')) or '',
            'co_quan': TEN_CQ_REV.get(_s(r.get('CO_QUAN')), _s(r.get('CO_QUAN'))),
            'concept': _s(r.get('KHAI_NIEM_CHUAN_HOA')) or '',
        }
        findings.append(f)
        if _s(r.get('LA_BENH_CHINH')) == 'CHÍNH':
            bc = f
    return findings, bc


def benh_row_tuple(ma_ho_so, r):
    return (
        ma_ho_so,
        r.get('STT_BENH'),
        1 if _s(r.get('LA_BENH_CHINH')) == 'CHÍNH' else 0,
        _s(r.get('MA_ICD')),
        _s(r.get('TEN_ICD')),
        TEN_CQ_REV.get(_s(r.get('CO_QUAN')), _s(r.get('CO_QUAN'))),
        r.get('MUC_DO_NANG') if r.get('MUC_DO_NANG') not in ('', None) else None,
        _s(r.get('CHUOI_GOC')),
        _s(r.get('KHAI_NIEM_CHUAN_HOA')),
        _s(r.get('NGUON_ANH_XA')),
        _s(r.get('DIEN_GIAI_CUA_BS')),
        1 if _s(r.get('CAN_RA_SOAT')) == 'CÓ' else 0,
    )


HO_SO_COLS = None  # điền lúc runtime từ build_ho_so_record đầu tiên


def import_ho_so_benh(conn):
    t0 = time.time()
    existed = {r[0] for r in conn.execute('SELECT ma_ho_so FROM ho_so')}
    print(f'  hồ sơ đã có sẵn trong DB: {len(existed)}')

    print('  đọc BENH_NHAN ...')
    bn_rows = _sheet_rows(config.SRC_QUANLY_XLSX, 'BENH_NHAN')
    print(f'  -> {len(bn_rows)} dòng')

    print('  đọc BENH_CHI_TIET ...')
    bct_rows = _sheet_rows(config.SRC_QUANLY_XLSX, 'BENH_CHI_TIET')
    print(f'  -> {len(bct_rows)} dòng')
    bct_by_hoso = collections.defaultdict(list)
    for r in bct_rows:
        bct_by_hoso[r['MA_HO_SO']].append(r)
    del bct_rows

    print('  đọc PHAN_LOAI_CO_QUAN ...')
    plr_rows = _sheet_rows(config.SRC_QUANLY_XLSX, 'PHAN_LOAI_CO_QUAN')
    print(f'  -> {len(plr_rows)} dòng')
    plr_by_hoso = collections.defaultdict(list)
    for r in plr_rows:
        plr_by_hoso[r['MA_HO_SO']].append(r)
    del plr_rows

    new_recs = []
    new_benh = []
    for bn in bn_rows:
        ma_ho_so = bn['MA_HO_SO']
        if ma_ho_so in existed:
            continue
        rec = build_ho_so_record(bn)
        apply_phan_loai(rec, plr_by_hoso.get(ma_ho_so, []))

        benh_rows = bct_by_hoso.get(ma_ho_so, [])
        findings, bc = build_findings(benh_rows)
        ts = tien_su.suy_luan(findings, bc, ten_chinh_thuc)
        for k, v in ts.items():
            rec[k.lower()] = v

        # PLAN_PERF.md §2 — cột hỗ trợ tìm kiếm SQL-paginated: tính ngay lúc
        # nạp để hồ sơ MỚI có sẵn dữ liệu (không phải đợi
        # scripts/build_search_cols.py chạy lại).
        rec['ho_ten_kd'], rec['search_blob_kd'] = fuzzy.build_search_cols(rec)

        new_recs.append(rec)
        for r in benh_rows:
            new_benh.append(benh_row_tuple(ma_ho_so, r))

    print(f'  hồ sơ MỚI cần nạp: {len(new_recs)}')
    if not new_recs:
        return {'ho_so_moi': 0, 'benh_moi': 0, 'cccd_trung_so': 0,
                'cccd_trung_ban_ghi': 0, 'thieu_sinh_hieu': 0}

    # ---- cờ CCCD_TRUNG (tự tính lúc nạp — §4) ----
    cccd_counter = collections.Counter(
        r['so_cccd'] for r in new_recs if r['so_cccd'])
    trung_so = {k for k, v in cccd_counter.items() if v > 1}
    n_cccd_trung_recs = 0
    for r in new_recs:
        if r['so_cccd'] and r['so_cccd'] in trung_so:
            flags = [f for f in r['co_qc'].split(';') if f]
            if 'CCCD_TRUNG' not in flags:
                flags.append('CCCD_TRUNG')
            r['co_qc'] = ';'.join(flags)
            n_cccd_trung_recs += 1

    # ---- cờ THIEU_SINH_HIEU (tự tính lúc nạp — §4) ----
    # Đợt này chiều cao/cân nặng/mạch/huyết áp CHƯA có dữ liệu (§6.4) nên
    # áp dụng cho toàn bộ hồ sơ mới nạp.
    n_thieu_sinh_hieu = 0
    for r in new_recs:
        if r['chieu_cao'] is None and r['can_nang'] is None and \
           r['mach'] is None and r['huyet_ap'] is None:
            flags = [f for f in r['co_qc'].split(';') if f]
            if 'THIEU_SINH_HIEU' not in flags:
                flags.append('THIEU_SINH_HIEU')
            r['co_qc'] = ';'.join(flags)
            n_thieu_sinh_hieu += 1

    for r in new_recs:
        r['so_loi'] = len([f for f in r['co_qc'].split(';') if f])

    cols = list(new_recs[0].keys())
    placeholders = ','.join('?' * len(cols))
    sql = (f'INSERT OR IGNORE INTO ho_so({",".join(cols)}) '
           f'VALUES ({placeholders})')
    conn.executemany(sql, [tuple(r[c] for c in cols) for r in new_recs])

    conn.executemany(
        'INSERT INTO benh(ma_ho_so, stt_benh, la_benh_chinh, ma_icd, ten_icd, '
        'co_quan, muc_do_nang, chuoi_goc, khai_niem, nguon_anh_xa, '
        'dien_giai_bs, can_ra_soat) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
        new_benh)
    conn.commit()
    print(f'  xong nạp hồ sơ+bệnh trong {time.time()-t0:.1f}s')

    return {
        'ho_so_moi': len(new_recs),
        'benh_moi': len(new_benh),
        'cccd_trung_so': len(trung_so),
        'cccd_trung_ban_ghi': n_cccd_trung_recs,
        'thieu_sinh_hieu': n_thieu_sinh_hieu,
    }


# ============================= BASELINE (lúc nạp) =============================
def compute_baseline(conn):
    """Chỉ tính 1 lần — nếu baseline_thongke đã có dữ liệu thì bỏ qua (baseline
    phải cố định kể từ lần nạp đầu tiên, không tính lại theo trạng thái rà soát
    hiện tại)."""
    n = conn.execute('SELECT COUNT(*) FROM baseline_thongke').fetchone()[0]
    if n > 0:
        print('  baseline_thongke đã có sẵn — bỏ qua (giữ nguyên số liệu nền)')
        return

    rows = []
    tong = conn.execute('SELECT COUNT(*) FROM ho_so').fetchone()[0]
    rows.append(('tong_quan', 'tong_ho_so', 'Tổng hồ sơ', tong))

    for ma_co, ten_co in [
        ('NGAY_SINH_UOC_LUONG', 'Ngày sinh ước lượng'),
        ('THIEU_CCCD', 'Thiếu CCCD'),
        ('CCCD_TRUNG', 'CCCD trùng'),
        ('CO_PHAN_LOAI_NHUNG_KHONG_CO_CHAN_DOAN', 'Có PL nhưng không chẩn đoán'),
        ('CON_CHAN_DOAN_CHUA_ANH_XA', 'Còn chẩn đoán chưa ánh xạ'),
        ('NGUON_DANH_DAU_NHIEU_PHAN_LOAI', 'Nguồn đánh dấu nhiều PL'),
        ('THI_LUC_CHUA_RO_BEN_MAT', 'Thị lực chưa rõ bên mắt'),
        ('ICD_MAY_TU_SUA_LOI_GO', 'ICD máy tự sửa lỗi gõ'),
        ('ICD_KHONG_DAC_HIEU', 'ICD không đặc hiệu'),
        ('NAM_SINH_SAI_NGUON', 'Năm sinh sai nguồn'),
        ('THIEU_SINH_HIEU', 'Thiếu sinh hiệu'),
    ]:
        c = conn.execute(
            "SELECT COUNT(*) FROM ho_so WHERE ';' || co_qc || ';' LIKE ?",
            (f'%;{ma_co};%',)).fetchone()[0]
        rows.append(('co_qc', ma_co, ten_co, c))

    for code, ten in TEN_CQ.items():
        c = conn.execute(
            'SELECT COUNT(*) FROM ho_so WHERE co_quan_benh_chinh = ?',
            (code,)).fetchone()[0]
        rows.append(('co_quan_benh_chinh', code, ten, c))

    conn.executemany(
        'INSERT OR IGNORE INTO baseline_thongke(nhom, ma, ten, gia_tri) '
        'VALUES (?,?,?,?)', rows)
    conn.commit()
    print(f'  đã lưu baseline_thongke: {len(rows)} dòng')


# ============================= MAIN =============================
def main():
    config.ensure_dirs()
    conn = db.get_connection()
    db.init_schema(conn)

    print('1) Nạp danh mục dropdown (doc/Import_KSK_Tren 18.xlsm) ...')
    dm_counts = load_danh_muc(conn)
    print('   ', dict(dm_counts))

    print('2) Nạp danh mục ICD (dmicdme) -> dm_icd + FTS5 ...')
    n_raw_icd, n_loaded_icd = load_dm_icd(conn)
    print(f'    dòng dữ liệu đọc được: {n_raw_icd} | dm_icd (đã khử trùng mã): {n_loaded_icd}')

    print('3) Seed tài khoản mẫu (admin/raso1) ...')
    seed_users(conn)

    print('4) Nạp hồ sơ (BENH_NHAN) + bệnh (BENH_CHI_TIET) + khám cơ quan '
          '(PHAN_LOAI_CO_QUAN) ...')
    stats = import_ho_so_benh(conn)
    print('   ', stats)

    print('5) Tính baseline (lúc nạp) ...')
    compute_baseline(conn)

    conn.close()
    print('\nHoàn tất nạp dữ liệu.')
    return stats


if __name__ == '__main__':
    main()
