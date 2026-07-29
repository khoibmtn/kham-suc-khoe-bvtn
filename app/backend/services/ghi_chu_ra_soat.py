# -*- coding: utf-8 -*-
"""
ghi_chu_ra_soat.py — Nhập lại file "Đối soát" (xuất từ nút "Xuất Excel" của
khối "Nhập lại file đã chỉnh sửa", xem xuat_file.py `POST
/api/xuat-file/doi-soat-xlsx` + nhap_doi_soat.py), rồi với MỖI hồ sơ ghép các
dòng thay đổi (Bổ sung/Ghi đè) thành 1 đoạn ghi chú, APPEND vào cột
`ghi_chu_ra_soat` (TEXT, bảng ho_so).

KHÁC HẲN nhap_doi_soat.py: file input ở đây là file "9 cột phẳng" do CHÍNH
app tự xuất ra (header: STT | Mã hồ sơ | Họ tên | Danh sách | Sẽ áp dụng |
Loại thay đổi | Trường | Giá trị cũ | Giá trị mới, sheet 'Đối soát'), KHÔNG
phải file header-103-cột chuẩn BYT — nên KHÔNG tái dùng doi_soat()/_apply_row()
của nhap_doi_soat.py, tự đọc/ghi độc lập ở đây.
"""
import io

REQUIRED_COLS = ['Mã hồ sơ', 'Họ tên', 'Loại thay đổi', 'Trường',
                 'Giá trị cũ', 'Giá trị mới']
OPTIONAL_COLS = ['Sẽ áp dụng']

SAMPLE = 200          # giới hạn chi_tiet trả về (preview trên màn hình)
SAMPLE_KHONG_KHOP = 50  # giới hạn danh sách mã không khớp trả về


def _norm(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _read_sheet(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True,
                                data_only=True)
    names = wb.sheetnames
    ws = wb['Đối soát'] if 'Đối soát' in names else wb[names[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _dong_ghi_chu(loai, truong, cu, moi):
    """Trả về 1 dòng ghi chú cho 1 thay đổi, hoặc None nếu 'Loại thay đổi'
    không nhận diện được (dữ liệu lạ/hỏng -> bỏ qua, không lỗi)."""
    if loai == 'Ghi đè':
        return f'Ghi đè: {truong}: {cu} -> {moi}'
    if loai == 'Bổ sung':
        return f'Bổ sung: {truong}: {moi}'
    return None


def xu_ly(conn, file_bytes, apply=False, user_id=None):
    """Đọc file "Đối soát" 9-cột, gộp ghi chú theo mã hồ sơ, APPEND vào
    ghi_chu_ra_soat của từng hồ sơ. apply=False -> chỉ tính toán (xem
    trước), KHÔNG ghi/commit gì. apply=True -> ghi UPDATE + INSERT nhat_ky,
    commit 1 LẦN ở cuối."""
    rows = _read_sheet(file_bytes)
    if not rows:
        raise ValueError('File rỗng — không đọc được dòng nào.')

    header = rows[0]
    col_idx = {}
    for ci, h in enumerate(header):
        h_s = _norm(h)
        if h_s:
            col_idx[h_s] = ci

    thieu = [c for c in REQUIRED_COLS if c not in col_idx]
    if thieu:
        raise ValueError('File thiếu cột bắt buộc: ' + ', '.join(thieu))

    i_ma = col_idx['Mã hồ sơ']
    i_ten = col_idx['Họ tên']
    i_loai = col_idx['Loại thay đổi']
    i_truong = col_idx['Trường']
    i_cu = col_idx['Giá trị cũ']
    i_moi = col_idx['Giá trị mới']
    i_apdung = col_idx.get('Sẽ áp dụng')

    def _get(r, i):
        return r[i] if i is not None and i < len(r) else None

    # Gom theo mã hồ sơ, GIỮ ĐÚNG thứ tự xuất hiện trong file gốc.
    theo_ma = {}       # ma -> {'ho_ten': str, 'dong': [str, ...]}
    thu_tu = []         # thứ tự mã hồ sơ xuất hiện lần đầu

    for r in rows[1:]:
        if not r:
            continue
        ma = _norm(_get(r, i_ma))
        if not ma:
            continue
        if i_apdung is not None:
            se_ap_dung = _norm(_get(r, i_apdung))
            if se_ap_dung == 'Không':
                continue

        loai = _norm(_get(r, i_loai))
        truong = _norm(_get(r, i_truong))
        cu = _norm(_get(r, i_cu))
        moi = _norm(_get(r, i_moi))
        dong = _dong_ghi_chu(loai, truong, cu, moi)
        if dong is None:
            continue

        if ma not in theo_ma:
            theo_ma[ma] = {'ho_ten': _norm(_get(r, i_ten)), 'dong': []}
            thu_tu.append(ma)
        theo_ma[ma]['dong'].append(dong)
        # Cập nhật họ tên nếu dòng trước để trống mà dòng sau có giá trị.
        if not theo_ma[ma]['ho_ten']:
            ho_ten_r = _norm(_get(r, i_ten))
            if ho_ten_r:
                theo_ma[ma]['ho_ten'] = ho_ten_r

    tong_ma_ho_so = len(thu_tu)
    khong_khop = []
    so_khong_khop = 0
    bo_qua_trung = 0
    so_luong_ghi = 0
    chi_tiet = []

    for ma in thu_tu:
        info = theo_ma[ma]
        dong_list = info['dong']
        if not dong_list:
            continue
        db_row = conn.execute('SELECT * FROM ho_so WHERE ma_ho_so=?', (ma,)).fetchone()
        if not db_row:
            so_khong_khop += 1
            if len(khong_khop) < SAMPLE_KHONG_KHOP:
                khong_khop.append(ma)
            continue

        block = '\n'.join(dong_list)
        hien_tai = db_row['ghi_chu_ra_soat'] if 'ghi_chu_ra_soat' in db_row.keys() else None
        hien_tai_co_noi_dung = bool(hien_tai and hien_tai.strip())

        # Chống ghi trùng: nếu block (đã strip) đã là substring của nội dung
        # hiện có -> bỏ qua hồ sơ này (không ghi lại lần 2).
        if hien_tai_co_noi_dung and block.strip() in hien_tai:
            bo_qua_trung += 1
            continue

        noi_dung_moi = (hien_tai.rstrip() + '\n\n' + block) if hien_tai_co_noi_dung else block
        so_luong_ghi += 1

        ho_ten = info['ho_ten'] or _norm(db_row['ho_ten'] if 'ho_ten' in db_row.keys() else None)
        if not apply and len(chi_tiet) < SAMPLE:
            chi_tiet.append({
                'ma_ho_so': ma, 'ho_ten': ho_ten,
                'so_dong_thay_doi': len(dong_list),
                'noi_dung_them': block,
            })
        elif apply:
            # apply=True: KHÔNG giới hạn — tính/ghi đủ toàn bộ, nhưng vẫn giữ
            # SAMPLE dòng đầu để trả về cho frontend hiển thị kết quả.
            if len(chi_tiet) < SAMPLE:
                chi_tiet.append({
                    'ma_ho_so': ma, 'ho_ten': ho_ten,
                    'so_dong_thay_doi': len(dong_list),
                    'noi_dung_them': block,
                })
            conn.execute('UPDATE ho_so SET ghi_chu_ra_soat=? WHERE ma_ho_so=?',
                        (noi_dung_moi, ma))
            conn.execute(
                'INSERT INTO nhat_ky(ma_ho_so, nguoi_dung_id, ten_truong, '
                'gia_tri_cu, gia_tri_moi) VALUES (?,?,?,?,?)',
                (ma, user_id, 'ghi_chu_ra_soat',
                 hien_tai if hien_tai else '', noi_dung_moi))

    if apply:
        conn.commit()

    return {
        'tong_ma_ho_so': tong_ma_ho_so,
        'so_luong_ghi': so_luong_ghi,
        'bo_qua_trung': bo_qua_trung,
        'khong_khop': khong_khop,
        'so_khong_khop': so_khong_khop,
        'chi_tiet': chi_tiet,
    }
