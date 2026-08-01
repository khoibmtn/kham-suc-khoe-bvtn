# -*- coding: utf-8 -*-
"""
nhap_doi_soat.py — Nhập lại file .xlsx đã chỉnh sửa (xuất từ nút "Xuất để chỉnh
sửa & nhập lại"), đối soát với DB:

- Match từng dòng theo cột MÃ ĐỊNH DANH (ma_ho_so) — KHÔNG theo họ tên/CCCD
  (các trường này có thể bị sửa nên không tin cậy để định danh).
- So sánh từng trường sửa được (patchable) của bảng ho_so:
    * BỔ SUNG: DB đang trống -> file có giá trị.
    * GHI ĐÈ : DB có giá trị -> file khác giá trị.
    * Ô trong file để TRỐNG -> BỎ QUA (không bao giờ xóa dữ liệu cũ).
- 2 chế độ: xem trước (apply=False, không ghi gì) / áp dụng (apply=True; bổ sung
  luôn được ghi, ghi đè chỉ ghi khi cho_ghi_de=True).
- Khi ghi: chuẩn hoá số + kiểm ngưỡng sinh hiệu, ghi nhat_ky, tính lại BMI /
  phân loại thể lực / cờ CCCD — hệt PATCH thủ công.
"""
import datetime
import io
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from services import export_xlsm, fuzzy, ngay_thang_valid, sinh_hieu_valid, qc, the_luc  # noqa: E402

# Trường KHÔNG cho nhập lại (khóa/định danh/quản lý rà soát/tự tính).
MANAGED = {
    'ma_ho_so', 'tt', 'nguoi_ra_soat_id', 'trang_thai', 'co_qc', 'so_loi',
    'thoi_diem_hoan_thanh', 'da_xuat_file', 'lan_xuat_cuoi', 'chan_doan_goc',
    'ho_ten_kd', 'search_blob_kd', 'chi_so_bmi',
    'rs_hanh_chinh', 'rs_sinh_ton', 'rs_the_luc', 'rs_canh_bao_khac',
}


def _patchable(conn):
    cols = [r['name'] for r in conn.execute('PRAGMA table_info(ho_so)')]
    return set(cols) - MANAGED


def _norm(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


# Nhận diện chuỗi ngày kiểu ISO (có thể kèm giờ) — khi user chỉnh sửa file
# .xlsx (xuất dạng TEXT '@') trong Excel rồi lưu lại, Excel có thể tự động
# chuyển ô ngày thành kiểu datetime; openpyxl(data_only=True) khi đó trả về
# object datetime.datetime, hoặc (tuỳ cách lưu) một chuỗi ISO như
# "1960-01-01 00:00:00" — cả 2 trường hợp đều PHẢI được ép về đúng
# dd/mm/yyyy mà toàn bộ ứng dụng dùng, KHÔNG được ghi thẳng vào DB/hiển thị
# nguyên dạng ISO (phản hồi anh Khôi — lỗi ảnh hưởng dữ liệu thật).
_ISO_DATE_RE = re.compile(r'^(\d{4})-(\d{1,2})-(\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?$')


def _chuan_hoa_ngay(v):
    """Ép 1 giá trị thô của ô ngày (object datetime/date, chuỗi ISO, hoặc
    chuỗi dd/mm/yyyy sẵn có) về chuỗi dd/mm/yyyy. Nếu không nhận diện được
    (không phải datetime/date, không khớp ISO, hoặc ISO nhưng ngày không có
    thật trong lịch) -> trả nguyên chuỗi đã _norm() để
    ngay_thang_valid.validate_date_changes() ở bước sau tự phát hiện & báo
    lỗi rõ ràng (không tự chế thêm thông báo lỗi ở đây)."""
    # datetime.datetime LÀ subclass của datetime.date -> check date là đủ.
    if isinstance(v, datetime.date):
        return v.strftime('%d/%m/%Y')
    s = _norm(v)
    m = _ISO_DATE_RE.match(s)
    if m:
        nam, thang, ngay = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(nam, thang, ngay).strftime('%d/%m/%Y')
        except ValueError:
            return s
    return s


def _chuan_hoa_gia_tri(field, v):
    """Chuẩn hoá field-aware: 3 trường ngày (ngay_sinh/ngay_vao/ngaycap_cccd)
    ép về dd/mm/yyyy qua _chuan_hoa_ngay(); mọi trường khác dùng _norm() y
    hệt hành vi cũ (không đổi)."""
    if field in ngay_thang_valid.DATE_FIELDS:
        return _chuan_hoa_ngay(v)
    return _norm(v)


def _recompute_bmi(cao, can):
    try:
        cao_f, can_f = float(cao), float(can)
        if cao_f > 0:
            return round(can_f / ((cao_f / 100) ** 2), 2)
    except (TypeError, ValueError):
        pass
    return None


class CanChonSheet(Exception):
    """Ném khi file có NHIỀU sheet mà chưa chỉ định `sheet` — trước đây
    code tự đoán ('Trên 18' nếu có, không thì lấy sheet ĐẦU TIÊN) — SAI LẦM
    thật đã xảy ra (phản hồi anh Khôi): sheet đầu tiên có thể là 1 sheet
    khác hẳn (bản nháp/tham chiếu cũ...) khiến đối soát đọc nhầm dữ liệu,
    ví dụ đè "Ngày sinh" thật thành 01/01 ước lượng — người dùng KHÔNG hề
    biết mình đang đối soát nhầm sheet. Giờ BẮT BUỘC hỏi khi có ≥2 sheet."""
    def __init__(self, sheet_names):
        self.sheet_names = sheet_names
        super().__init__('File có nhiều sheet — cần chọn sheet để đối soát.')


def _sheet_names(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _read_sheet(file_bytes, sheet=None):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True,
                                data_only=True)
    names = wb.sheetnames
    if sheet:
        if sheet not in names:
            wb.close()
            raise ValueError(f'Không tìm thấy sheet "{sheet}" trong file.')
        ws = wb[sheet]
    elif len(names) == 1:
        ws = wb[names[0]]
    else:
        wb.close()
        raise CanChonSheet(names)
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def doi_soat(conn, file_bytes, apply=False, cho_ghi_de=False, user_id=None,
             cot_chon=None, sheet=None, ma_loai_tru=None,
             bo_gioi_han_chi_tiet=False):
    """sheet: tên sheet cần đối soát — BẮT BUỘC chỉ định nếu file có ≥2
    sheet (ném CanChonSheet kèm danh sách tên nếu chưa chỉ định, để caller
    hỏi user rồi gọi lại). File chỉ 1 sheet thì tự dùng, không cần hỏi.

    cot_chon: None/rỗng -> dùng TẤT CẢ cột phát hiện được (hành vi cũ,
    tương thích ngược); tập hợp tên field (chữ thường, khớp schema.sql) ->
    CHỈ đối soát/ghi những cột đó — các cột khác coi như KHÔNG có trong file
    (bỏ qua hoàn toàn, kể cả nếu có giá trị). Trả kèm 'cot_phat_hien' (TOÀN
    BỘ cột hợp lệ tìm thấy trong file, không phụ thuộc cot_chon) để frontend
    liệt kê + hỏi user chọn trước khi đối soát/ghi (phản hồi anh Khôi).

    ma_loai_tru: tập hợp mã hồ sơ bị bỏ tick ở bước xem trước (checkbox từng
    dòng) — CHỈ có tác dụng khi apply=True (dòng đó KHÔNG được ghi/không có
    nhat_ky mới); khi apply=False (xem trước) tham số này KHÔNG ảnh hưởng gì,
    chi_tiet vẫn liệt kê ĐẦY ĐỦ mọi dòng có thay đổi. None/rỗng -> tương
    thích ngược (mọi dòng đều được ghi như trước).

    bo_gioi_han_chi_tiet: mặc định False (hành vi cũ — chi_tiet chỉ giữ tối
    đa SAMPLE=200 hồ sơ đầu có thay đổi, dùng cho preview trên màn hình).
    True -> chi_tiet giữ TẤT CẢ hồ sơ có thay đổi, không giới hạn — dùng
    riêng cho mục đích xuất Excel đầy đủ (endpoint doi-soat-xlsx), KHÔNG
    dùng cho preview/apply thường vì có thể rất lớn."""
    ma_loai_tru = ma_loai_tru or set()
    rows = _read_sheet(file_bytes, sheet=sheet)
    if len(rows) < 4:
        raise ValueError('File không đúng định dạng (cần 3 dòng tiêu đề + dữ liệu).')

    label_row, code_row = rows[0], rows[1]
    code_by_col = {ci: str(c).strip() for ci, c in enumerate(code_row) if c not in (None, '')}
    label_by_col = {ci: (str(label_row[ci]).strip() if ci < len(label_row) and label_row[ci] else code_by_col[ci])
                    for ci in code_by_col}
    id_col = next((ci for ci, c in code_by_col.items() if c == export_xlsm.ID_COL_CODE), None)
    if id_col is None:
        raise ValueError('File thiếu cột MÃ ĐỊNH DANH — hãy xuất bằng nút '
                         '"Xuất để chỉnh sửa & nhập lại".')

    patch = _patchable(conn)
    col_field_headers = {ci: c.lower() for ci, c in code_by_col.items()
                         if ci != id_col and c.lower() in patch}
    if not col_field_headers:
        raise ValueError('Không có cột dữ liệu nào hợp lệ để nhập.')

    # Phản hồi anh Khôi: file xuất luôn giữ ĐỦ header 103 cột chuẩn BYT dù
    # người dùng chỉ điền vài cột — liệt kê theo HEADER (như trước) khiến
    # danh sách chọn cột dài cả trăm dòng dù file thực tế chỉ có ~12 cột có
    # dữ liệu. Chỉ liệt kê cột có ÍT NHẤT 1 ô khác trống trong toàn bộ dòng
    # dữ liệu (quét 1 lượt, dừng sớm khi đã đủ mọi cột).
    con_thieu = set(col_field_headers.values())
    for r in rows[3:]:
        if not r or not con_thieu:
            break
        for ci, f in col_field_headers.items():
            if f in con_thieu and ci < len(r) and _norm(r[ci]) != '':
                con_thieu.discard(f)
    co_du_lieu = set(col_field_headers.values()) - con_thieu

    col_field_all = {ci: f for ci, f in col_field_headers.items() if f in co_du_lieu}
    if not col_field_all:
        raise ValueError('File không có cột dữ liệu nào có giá trị để nhập.')
    # Thứ tự xuất hiện trong file — giữ nguyên để frontend liệt kê tự nhiên.
    cot_phat_hien = [{'field': f, 'nhan': label_by_col.get(ci, f)}
                      for ci, f in col_field_all.items()]

    if cot_chon:
        col_field = {ci: f for ci, f in col_field_all.items() if f in cot_chon}
        if not col_field:
            raise ValueError('Chưa chọn cột nào hợp lệ để cập nhật.')
    else:
        col_field = col_field_all

    nguong = sinh_hieu_valid.load_nguong(conn)

    tong = so_khop = so_khong_khop = so_bo_sung = so_ghi_de = 0
    da_ghi_bo_sung = da_ghi_ghi_de = 0
    khong_khop = []          # danh sách ma_ho_so không tìm thấy
    chi_tiet = []            # mẫu preview (giới hạn)
    loi_validate = []        # lỗi ngưỡng khi áp dụng
    SAMPLE = 200

    for r in rows[3:]:
        if not r or id_col >= len(r):
            continue
        ma = _norm(r[id_col])
        if not ma:
            continue
        tong += 1
        db_row = conn.execute('SELECT * FROM ho_so WHERE ma_ho_so=?', (ma,)).fetchone()
        if not db_row:
            so_khong_khop += 1
            if len(khong_khop) < 50:
                khong_khop.append(ma)
            continue
        so_khop += 1

        changes = []         # {field, nhan, cu, moi, loai}
        for ci, field in col_field.items():
            if ci >= len(r):
                continue
            up = r[ci]
            up_s = _chuan_hoa_gia_tri(field, up)
            if up_s == '':
                continue                       # ô trống -> bỏ qua (không xóa)
            db_s = _chuan_hoa_gia_tri(field, db_row[field] if field in db_row.keys() else None)
            if up_s == db_s:
                continue
            loai = 'bo_sung' if db_s == '' else 'ghi_de'
            changes.append({'field': field, 'nhan': label_by_col.get(ci, field),
                            'cu': db_s, 'moi': up_s, 'loai': loai})
        if not changes:
            continue
        so_bo_sung += sum(1 for c in changes if c['loai'] == 'bo_sung')
        so_ghi_de += sum(1 for c in changes if c['loai'] == 'ghi_de')
        if bo_gioi_han_chi_tiet or len(chi_tiet) < SAMPLE:
            chi_tiet.append({'ma_ho_so': ma, 'ho_ten': _norm(db_row['ho_ten']),
                             'changes': changes})

        if apply and ma not in ma_loai_tru:
            da_bs, da_gd = _apply_row(conn, ma, db_row, changes, cho_ghi_de,
                                      nguong, user_id, loi_validate)
            da_ghi_bo_sung += da_bs
            da_ghi_ghi_de += da_gd

    # Cột "Danh sách" ở bước xem trước đối soát (criterion 4.3): mỗi dòng
    # trong chi_tiet thuộc những danh sách tùy chỉnh nào — 1 TRUY VẤN PHỤ
    # DUY NHẤT theo TOÀN BỘ mã hồ sơ đã có trong chi_tiet (không N+1), khuôn
    # mẫu ho_so.py dòng ~544-553 / export_xlsm.create_job().
    if chi_tiet:
        ma_ct = [c['ma_ho_so'] for c in chi_tiet]
        ph_ct = ','.join('?' * len(ma_ct))
        ds_map = {}
        for r_ds in conn.execute(
                f'SELECT dsh.ma_ho_so AS ma_ho_so, ds.ten AS ten '
                f'FROM danh_sach_ho_so dsh JOIN danh_sach ds ON ds.id = dsh.danh_sach_id '
                f'WHERE dsh.ma_ho_so IN ({ph_ct})', ma_ct):
            ds_map.setdefault(r_ds['ma_ho_so'], []).append(r_ds['ten'])
        for c in chi_tiet:
            c['_danh_sach'] = ds_map.get(c['ma_ho_so'], [])

    if apply:
        conn.commit()

    return {
        'apply': apply,
        'cot_phat_hien': cot_phat_hien,
        'tong_dong': tong,
        'so_khop': so_khop,
        'so_khong_khop': so_khong_khop,
        'khong_khop': khong_khop,
        'so_bo_sung': so_bo_sung,
        'so_ghi_de': so_ghi_de,
        'da_ghi_bo_sung': da_ghi_bo_sung,
        'da_ghi_ghi_de': da_ghi_ghi_de,
        'loi_validate': loi_validate[:50],
        'chi_tiet': chi_tiet,
    }


def _apply_row(conn, ma, db_row, changes, cho_ghi_de, nguong, user_id, loi_out):
    """Ghi các thay đổi được phép của 1 hồ sơ. Trả (số bổ sung, số ghi đè)."""
    apply_changes = [c for c in changes
                     if c['loai'] == 'bo_sung' or (c['loai'] == 'ghi_de' and cho_ghi_de)]
    if not apply_changes:
        return 0, 0
    field_moi = {c['field']: c['moi'] for c in apply_changes}

    # Chuẩn hoá số + kiểm ngưỡng sinh hiệu (giống PATCH). Lỗi -> loại field đó.
    field_moi, loi1 = sinh_hieu_valid.normalize_numeric_changes(field_moi)
    field_moi, loi2 = sinh_hieu_valid.validate_changes(field_moi, nguong)
    for e in (loi1 or []) + (loi2 or []):
        loi_out.append({'ma_ho_so': ma, 'ly_do': e.get('ly_do', str(e))})

    # Trường ngày (ngay_sinh/ngay_vao/ngaycap_cccd): field_moi ở đây đã được
    # ép chuẩn dd/mm/yyyy từ bước preview (_chuan_hoa_gia_tri) NHƯNG có thể
    # vẫn là giá trị "rác" không parse được thành ngày thật (vd chuỗi ISO
    # không tồn tại trong lịch, hoặc chuỗi bất kỳ không khớp dd/mm/yyyy) —
    # CHỦ Ý loại HẲN field đó khỏi field_moi trước khi ghi (khác cách xử lý
    # lỗi ngưỡng sinh hiệu ở trên — ở đây KHÔNG được ghi giá trị ngày sai/rác
    # vào DB, theo đúng yêu cầu; KHÔNG áp dụng .pop() tương tự cho vitals).
    field_moi, loi3 = ngay_thang_valid.validate_date_changes(field_moi)
    for e in (loi3 or []):
        loi_out.append({'ma_ho_so': ma, 'ly_do': e.get('ly_do', str(e))})
        field_moi.pop(e['field'], None)

    if not field_moi:
        return 0, 0

    old_cccd = db_row['so_cccd'] if 'so_cccd' in db_row.keys() else None
    # BMI tự tính khi chiều cao/cân nặng đổi
    if 'chieu_cao' in field_moi or 'can_nang' in field_moi:
        cao = field_moi.get('chieu_cao', db_row['chieu_cao'])
        can = field_moi.get('can_nang', db_row['can_nang'])
        field_moi['chi_so_bmi'] = _recompute_bmi(cao, can)

    set_clauses, args, applied = [], [], []
    for field, moi in field_moi.items():
        old = db_row[field] if field in db_row.keys() else None
        set_clauses.append(f'{field} = ?')
        args.append(moi)
        conn.execute(
            'INSERT INTO nhat_ky(ma_ho_so, nguoi_dung_id, ten_truong, '
            'gia_tri_cu, gia_tri_moi) VALUES (?,?,?,?,?)',
            (ma, user_id, field, '' if old is None else str(old),
             '' if moi is None else str(moi)))
        applied.append(field)
    if set_clauses:
        args.append(ma)
        conn.execute(f'UPDATE ho_so SET {", ".join(set_clauses)} WHERE ma_ho_so=?', args)
        # Phản hồi anh Khôi: file đối soát có thể sửa CCCD/họ tên/ngày sinh...
        # (1 trong 19 trường blob tìm kiếm) -> đồng bộ lại ho_ten_kd/
        # search_blob_kd ngay, giống hệt lý do PATCH thủ công qua panel chi
        # tiết (routers/ho_so.py patch_ho_so) — xem services/fuzzy.
        # dong_bo_search_cols(). CÙNG conn đang ghi, TRƯỚC commit của caller
        # (doi_soat() chỉ commit 1 lần ở cuối khi apply=True).
        fuzzy.dong_bo_search_cols(conn, ma)

    if 'so_cccd' in field_moi:
        qc.recompute_cccd_flags(conn, ma, old_cccd, field_moi['so_cccd'], user_id)
    if 'chieu_cao' in field_moi or 'can_nang' in field_moi:
        the_luc.tinh_va_ap_pl(conn, ma, user_id)

    # đếm theo phân loại gốc (áp dụng)
    fields_applied = set(applied)
    bs = sum(1 for c in apply_changes if c['loai'] == 'bo_sung' and c['field'] in fields_applied)
    gd = sum(1 for c in apply_changes if c['loai'] == 'ghi_de' and c['field'] in fields_applied)
    return bs, gd
