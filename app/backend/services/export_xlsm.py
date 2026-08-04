# -*- coding: utf-8 -*-
"""
export_xlsm.py — Pipeline 2: xuất file .xlsm nộp Bộ Y tế (§7 SPEC).

KHÔNG viết lại pipeline xuất: bảng `ho_so` được thiết kế để tên cột (viết
hoa) khớp THẲNG với mã trường mẫu BYT (đã kiểm chứng: 99/99 mã trường ở
dòng 2 của `doc/Import_KSK_Tren 18.xlsm` có cột `ho_so` cùng tên viết
thường) — nên "ánh xạ ngược" của import_data.py chỉ là `col.upper()`,
không cần bảng tra riêng. Việc ghi file thật sự vẫn gọi lại
`build/build_xlsm.write_xlsm` (qua tiến trình con `export_worker.py`),
đúng bẫy §10: KHÔNG tạo workbook mới, tách theo xã, mỗi xã một tiến trình.

Job chạy NỀN trên 1 thread; mỗi xã lại là 1 SUBPROCESS riêng (export_worker.py)
để hệ điều hành thu hồi bộ nhớ openpyxl sau từng file (bẫy §10 SPEC).
"""
import datetime
import io
import json
import os
import re
import subprocess
import sys
import threading
import time
import uuid

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402
import db  # noqa: E402
from services import qc  # noqa: E402

WORKER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'export_worker.py')
EXPORTS_DIR = os.path.join(config.DATA_DIR, 'exports')

SCOPE_TYPES = ('toan_bo', 'xa', 'can_bo', 'trang_thai', 'chon_tay', 'danh_sach')

# Cờ "đã rà soát xong" (checkbox panel chi tiết). "chi_rs_xong" ở tuỳ chọn
# xuất => chỉ lấy hồ sơ đã tick ĐỦ 4 mục.
_RS_XONG_COLS = ('rs_hanh_chinh', 'rs_sinh_ton', 'rs_the_luc', 'rs_canh_bao_khac')
_RS_XONG_SQL = ' AND '.join(f'{c}=1' for c in _RS_XONG_COLS)


def _apply_rs_xong(where_sql, chi_rs_xong):
    """Bọc điều kiện phạm vi + (tuỳ chọn) AND đã-rà-soát-xong-đủ-4-mục."""
    return f'({where_sql}) AND {_RS_XONG_SQL}' if chi_rs_xong else where_sql

TRANG_THAI_NHAN = {
    'chua_ra_soat': 'Chưa rà soát', 'dang_ra_soat': 'Đang rà soát',
    'hoan_thanh': 'Hoàn thành', 'can_doi_chieu_giay': 'Cần đối chiếu giấy',
}

# Danh sách cột mở rộng CHÍNH XÁC theo §7.2 SPEC — cột từ 104 trở đi, mặc
# định KHÔNG ghi, chỉ ghi khi user tick chọn & bật option (mặc định TẮT).
EXTENDED_COLUMNS = [
    ('MA_BENH_CHINH', 'Mã bệnh chính'),
    ('MA_BENH_KEM', 'Mã bệnh kèm'),
    ('TEN_BENH_KEM', 'Tên bệnh kèm'),
    ('CO_QUAN_BENH_CHINH', 'Cơ quan bệnh chính'),
    ('CHAN_DOAN_GOC', 'Chẩn đoán gốc'),
    ('CO_QC', 'Cờ QC'),
    ('SO_LOI', 'Số lỗi'),
    ('GHI_CHU_RA_SOAT', 'Ghi chú rà soát'),
    ('GHI_CHU_CAN_BO', 'Ghi chú nhân viên'),
    ('NGUOI_RA_SOAT', 'Người rà soát'),
    ('TRANG_THAI', 'Trạng thái'),
    ('THOI_DIEM_HOAN_THANH', 'Thời điểm hoàn thành'),
    ('GLU_THOI_DIEM', 'Thời điểm đo glucose'),
    ('GLU_GIA_TRI', 'Giá trị glucose'),
    ('KQ_DIEN_TIM', 'Kết quả điện tim'),
    ('KQ_SIEU_AM_O_BUNG', 'Kết quả siêu âm ổ bụng'),
    ('MA_HO_SO', 'Mã hồ sơ'),
    ('DANH_SACH', 'Danh sách'),
]
EXTENDED_LABELS = dict(EXTENDED_COLUMNS)
EXTENDED_CODES = {c for c, _ in EXTENDED_COLUMNS}

_JOBS = {}
_JOBS_LOCK = threading.Lock()


# ============================= PHẠM VI =============================
def resolve_scope_where(pham_vi, gia_tri):
    gia_tri = gia_tri or []
    if pham_vi == 'toan_bo':
        return '1=1', []
    if pham_vi == 'xa':
        vals = [v for v in gia_tri if v]
        if not vals:
            raise ValueError('Thiếu danh sách xã/phường')
        return f"maxa_cu_tru IN ({','.join('?' * len(vals))})", vals
    if pham_vi == 'can_bo':
        try:
            vals = [int(v) for v in gia_tri]
        except (TypeError, ValueError):
            raise ValueError('Mã nhân viên không hợp lệ')
        if not vals:
            raise ValueError('Thiếu danh sách nhân viên')
        return f"nguoi_ra_soat_id IN ({','.join('?' * len(vals))})", vals
    if pham_vi == 'trang_thai':
        vals = [v for v in gia_tri if v]
        if not vals:
            raise ValueError('Thiếu danh sách trạng thái')
        return f"trang_thai IN ({','.join('?' * len(vals))})", vals
    if pham_vi == 'chon_tay':
        vals = [v for v in gia_tri if v]
        if not vals:
            raise ValueError('Danh sách mã hồ sơ rỗng')
        return f"ma_ho_so IN ({','.join('?' * len(vals))})", vals
    if pham_vi == 'danh_sach':
        try:
            vals = [int(v) for v in gia_tri]
        except (TypeError, ValueError):
            raise ValueError('Mã danh sách không hợp lệ')
        if not vals:
            raise ValueError('Thiếu danh sách cần xuất')
        return (f"ma_ho_so IN (SELECT ma_ho_so FROM danh_sach_ho_so "
                f"WHERE danh_sach_id IN ({','.join('?' * len(vals))}))", vals)
    raise ValueError(f"pham_vi không hợp lệ: {pham_vi!r} (phải là {SCOPE_TYPES})")


def _red_flag_where():
    parts, args = [], []
    for f in sorted(qc.RED_FLAGS):
        parts.append("(';'||co_qc||';') LIKE ?")
        args.append(f'%;{f};%')
    return '(' + ' OR '.join(parts) + ')', args


def preview(conn, pham_vi, gia_tri, include_errors, chi_rs_xong=False):
    """Đếm trước khi xuất (§7.2): tổng trong phạm vi, số còn cờ 🔴, sẽ xuất,
    sẽ loại trừ (nếu include_errors=False, các ca cờ đỏ bị loại)."""
    where_sql, args = resolve_scope_where(pham_vi, gia_tri)
    where_sql = _apply_rs_xong(where_sql, chi_rs_xong)
    tong = conn.execute(f'SELECT COUNT(*) FROM ho_so WHERE {where_sql}', args).fetchone()[0]
    red_sql, red_args = _red_flag_where()
    do_flag_count = conn.execute(
        f'SELECT COUNT(*) FROM ho_so WHERE {where_sql} AND {red_sql}',
        args + red_args).fetchone()[0]
    se_loai_tru = 0 if include_errors else do_flag_count
    return {'tong': tong, 'do_flag_count': do_flag_count,
            'se_xuat': tong - se_loai_tru, 'se_loai_tru': se_loai_tru}


# ================= XUẤT .XLSX ĐƠN THUẦN (không macro) =================
# Khác pipeline .xlsm nộp Bộ: KHÔNG mở template nặng (dmicdme 35.735 dòng),
# KHÔNG job nền/đĩa, KHÔNG subprocess — chỉ dựng 1 workbook write_only trong
# bộ nhớ rồi trả bytes. Nhờ vậy chạy được CẢ trên bản đám mây (serverless
# không có tiến trình nền/đĩa bền) lẫn máy local, tải về ngay.
FIRST_DATA_ROW = 4              # mẫu BYT: dòng 1 nhãn, 2 mã trường, 3 hướng dẫn
_NCOL = 103                     # đúng 103 cột như mẫu 'Trên 18'
_PLAIN_TEXT_CODES = {'SO_CCCD', 'NGAY_VAO', 'NGAY_SINH', 'NGAYCAP_CCCD'}
_TPL_HEADER_CACHE = None

# Phản hồi anh Khôi: file mẫu BYT dùng ',' làm dấu phân cách thập phân cho
# 4 cột thính lực (m) — nhân viên nhập qua ô number trên trình duyệt (luôn
# dùng '.') nên phải đổi lại lúc XUẤT, không đổi ở nơi lưu trữ (ho_so vẫn
# giữ '.' để không phá vỡ validate/parse số ở nơi khác của app).
_TRUONG_THAP_PHAN_PHAY = ('TAI_TRAI_NOI_THUONG', 'TAI_TRAI_NOI_THAM',
                          'TAI_PHAI_NOI_THUONG', 'TAI_PHAI_NOI_THAM')


def _doi_dau_thap_phan(rec):
    for code in _TRUONG_THAP_PHAN_PHAY:
        v = rec.get(code)
        if v not in (None, '') and '.' in str(v):
            rec[code] = str(v).replace('.', ',')
    return rec


def _template_header():
    """Đọc 3 dòng header (nhãn/mã/hướng dẫn) của sheet 'Trên 18' từ template
    — CHỈ 3 dòng đầu (read_only, không đụng sheet danh mục nặng). Trả
    (header_rows: list[3][103], code2col: {MÃ_TRƯỜNG: chỉ_số_cột 1-based}).
    Cache ở module vì header cố định, dùng lại cho mọi lần xuất."""
    global _TPL_HEADER_CACHE
    if _TPL_HEADER_CACHE is None:
        import openpyxl
        wb = openpyxl.load_workbook(config.CATALOG_XLSM, read_only=True,
                                    keep_vba=False)
        ws = wb['Trên 18']
        header_rows = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            vals = list(row[:_NCOL])
            vals += [None] * (_NCOL - len(vals))
            header_rows.append(vals)
        wb.close()
        code2col = {}
        for c, code in enumerate(header_rows[1], 1):   # dòng 2 = mã trường
            if code:
                code2col[str(code).strip()] = c
        _TPL_HEADER_CACHE = (header_rows, code2col)
    return _TPL_HEADER_CACHE


# Cột định danh (khóa nhập lại) đặt TRƯỚC 103 cột mẫu khi with_id=True.
ID_COL_LABEL = 'MÃ ĐỊNH DANH (khóa — KHÔNG sửa/xóa)'
ID_COL_CODE = 'MA_HO_SO'
ID_COL_INSTR = 'Cột để đối soát khi nhập lại. Giữ nguyên, không sửa/xóa.'


def _plain_xlsx_bytes(rows, with_id=False, ext_columns=None, ext_data=None):
    """Dựng .xlsx write_only: 1 sheet 'Trên 18' (3 dòng header giống mẫu +
    dữ liệu từ dòng 4). with_id=True chèn cột MÃ ĐỊNH DANH (ma_ho_so) ở ĐẦU để
    nhập lại đối soát. `ext_columns`: list mã cột mở rộng (subset của
    EXTENDED_CODES) cần NỐI THÊM vào CUỐI mỗi dòng — None/rỗng giữ nguyên
    hành vi cũ (không cột mở rộng). `ext_data`: {ma_ho_so: {MÃ_CỘT: giá_trị}}
    đã tính SẴN ở nơi gọi (build_plain_xlsx) — hàm này chỉ đọc, không tự tính.
    Trả (bytes, số_dòng)."""
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill

    ext_columns = ext_columns or []
    ext_data = ext_data or {}

    header_rows, code2col = _template_header()
    text_cols = {code2col[c] for c in _PLAIN_TEXT_CODES if c in code2col}

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Trên 18')
    ws.freeze_panes = 'C4' if with_id else 'B4'
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical='top')
    id_fill = PatternFill('solid', fgColor='FFF2CC')  # tô vàng cột khóa

    id_head = [ID_COL_LABEL, ID_COL_CODE, ID_COL_INSTR]
    # Dòng header cột mở rộng: dòng 1 = nhãn (in đậm), dòng 2 = mã (khớp cơ
    # chế đối soát theo tên cột của nhap_doi_soat.py), dòng 3 = trống.
    ext_head_rows = [
        [EXTENDED_LABELS[c] for c in ext_columns],
        list(ext_columns),
        [None] * len(ext_columns),
    ]
    for ri, hrow in enumerate(header_rows, 1):
        cells = []
        if with_id:
            c = WriteOnlyCell(ws, value=id_head[ri - 1])
            if ri == 1:
                c.font = bold
            c.alignment = wrap
            c.fill = id_fill
            cells.append(c)
        for v in hrow:
            c = WriteOnlyCell(ws, value=(v if v not in (None, '') else None))
            if ri == 1:
                c.font = bold
            c.alignment = wrap
            cells.append(c)
        for v in ext_head_rows[ri - 1]:
            c = WriteOnlyCell(ws, value=v)
            if ri == 1:
                c.font = bold
            c.alignment = wrap
            cells.append(c)
        ws.append(cells)

    for i, row in enumerate(rows):
        rec = _doi_dau_thap_phan({k.upper(): row[k] for k in row.keys()})
        line = [None] * _NCOL
        line[0] = i + 1                                # cột TT
        for code, col in code2col.items():
            if code in rec and rec[code] not in (None, ''):
                line[col - 1] = rec[code]
        cells = []
        if with_id:
            c = WriteOnlyCell(ws, value=row['ma_ho_so'])
            c.number_format = '@'
            c.fill = id_fill
            cells.append(c)
        for ci, v in enumerate(line, 1):
            c = WriteOnlyCell(ws, value=v)
            if ci in text_cols and v not in (None, ''):
                c.number_format = '@'                  # giữ CCCD/ngày dạng text
            cells.append(c)
        row_ext = ext_data.get(row['ma_ho_so'], {}) if ext_columns else {}
        for code in ext_columns:
            cells.append(WriteOnlyCell(ws, value=row_ext.get(code)))
        ws.append(cells)

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue(), len(rows)


def build_plain_xlsx(conn, pham_vi, gia_tri, include_errors, chi_rs_xong=False,
                     with_id=False, extended=None):
    """Xuất .xlsx đơn thuần cho phạm vi đã chọn (gộp mọi hồ sơ vào 1 sheet).
    Loại hồ sơ còn cờ đỏ khi include_errors=False, hệt logic .xlsm. with_id=True
    thêm cột MÃ ĐỊNH DANH để chỉnh sửa rồi nhập lại. `extended`:
    {'enabled': bool, 'columns': [...]} hoặc None (mặc định, tương thích
    ngược) — bật thì NỐI THÊM cột mở rộng (§7.2) vào cuối, hệt danh mục nút
    .xlsm. Ném ValueError nếu phạm vi rỗng — router chuyển thành 400."""
    where_sql, args = resolve_scope_where(pham_vi, gia_tri)
    where_sql = _apply_rs_xong(where_sql, chi_rs_xong)
    rows = conn.execute(
        f'SELECT * FROM ho_so WHERE {where_sql} ORDER BY maxa_cu_tru, tt',
        args).fetchall()
    if not rows:
        raise ValueError('Không có hồ sơ nào khớp phạm vi đã chọn')
    if not include_errors:
        red_sql, red_args = _red_flag_where()
        red_set = {r['ma_ho_so'] for r in conn.execute(
            f'SELECT ma_ho_so FROM ho_so WHERE {where_sql} AND {red_sql}',
            args + red_args)}
        rows = [r for r in rows if r['ma_ho_so'] not in red_set]
        if not rows:
            raise ValueError('Toàn bộ hồ sơ trong phạm vi đều còn cờ đỏ — '
                             'bật "Xuất kèm cả hồ sơ lỗi" nếu vẫn muốn xuất')

    ext_enabled = bool(extended and extended.get('enabled'))
    ext_columns = []
    if ext_enabled:
        ext_columns = [c for c in (extended.get('columns') or []) if c in EXTENDED_CODES]

    ext_data = {}
    if ext_columns:
        user_map = {r['id']: r['ho_ten'] for r in
                    conn.execute('SELECT id, ho_ten FROM nguoi_dung')}
        ma_list = [r['ma_ho_so'] for r in rows]
        danh_sach_map = {}
        if ma_list:
            ph_ds = ','.join('?' * len(ma_list))
            for r_ds in conn.execute(
                    f'SELECT dsh.ma_ho_so AS ma_ho_so, ds.ten AS ten '
                    f'FROM danh_sach_ho_so dsh JOIN danh_sach ds ON ds.id = dsh.danh_sach_id '
                    f'WHERE dsh.ma_ho_so IN ({ph_ds})', ma_list):
                danh_sach_map.setdefault(r_ds['ma_ho_so'], []).append(r_ds['ten'])
        for r in rows:
            full = _row_ext(r, user_map.get(r['nguoi_ra_soat_id'], ''), danh_sach_map)
            ext_data[r['ma_ho_so']] = {c: full[c] for c in ext_columns}

    return _plain_xlsx_bytes(rows, with_id=with_id, ext_columns=ext_columns,
                             ext_data=ext_data)


# ============================= XUẤT HIS (.xlsx) =============================
# Nút "Xuất HIS" — KHÁC HẲN 2 pipeline trên: không dựng theo mẫu BYT (103/
# 119 cột khác nhau, mã trường khác nhau) mà theo ĐÚNG định dạng import của
# phần mềm HIS đơn vị đang dùng — 119 cột, đối chiếu field-by-field với file
# mẫu thật `doc/minh-lo/Minhlo_KSK trên 18T TT25.xls` (dòng 1 = mã cột, dữ
# liệu từ dòng 2, ô trống ghi chuỗi "NULL"). Tái dùng CÙNG resolve_scope_where
# + _apply_rs_xong + _red_flag_where như build_plain_xlsx — mọi phạm vi/tuỳ
# chọn hoạt động giống hệt nút "Tải .xlsx (kèm mã định danh)".
HIS_COLUMNS = [
    'makcb', 'soksk', 'ngay_lap_phieu', 'nguon_kinh_phi', 'nhom_mau', 'ly_do_ksk',
    'tsgd', 'tsgd_ma_benh',
] + [f'tien_su_ban_than_{i}' for i in range(1, 23)] + [
    'benh_tat_khac', 'dang_dieu_tri', 'dang_dieu_tri_benh', 'thai_san', 'thai_san_benh',
    'chieu_cao', 'can_nang', 'bmi', 'mach', 'huyet_ap_tam_thu', 'huyet_ap_tam_truong',
    'kham_the_luc_pl',
    'tuan_hoan_manv', 'tuan_hoan', 'tuan_hoan_pl',
    'ho_hap_manv', 'ho_hap', 'ho_hap_pl',
    'tieu_hoa_manv', 'tieu_hoa', 'tieu_hoa_pl',
    'than_tiet_nieu_manv', 'than_tiet_nieu', 'than_tiet_nieu_pl',
    'noi_tiet_manv', 'noi_tiet', 'noi_tiet_pl',
    'co_xuong_khop_manv', 'co_xuong_khop', 'co_xuong_khop_pl',
    'than_kinh_manv', 'than_kinh', 'than_kinh_pl',
    'tam_than_manv', 'tam_than', 'tam_than_pl',
    'ngoai_khoa_manv', 'ngoai_khoa', 'ngoai_khoa_pl',
    'da_lieu_manv', 'da_lieu', 'da_lieu_pl',
    'san_manv', 'san', 'san_pl',
    'mat_manv', 'mat_kk_mat_phai', 'mat_kk_mat_trai', 'mat_ck_mat_phai',
    'mat_ck_mat_trai', 'mat_benh', 'mat_pl',
    'tmh_manv', 'tmh_tai_trai_noi_thuong', 'tmh_tai_trai_noi_tham',
    'tmh_tai_phai_noi_thuong', 'tmh_tai_phai_noi_tham', 'tmh_benh', 'tmh_pl',
    'rhm_manv', 'rhm_ham_tren', 'rhm_ham_duoi', 'rhm_benh', 'rhm_pl',
    'mau_manv', 'huyet_hoc', 'mau_duong_mau', 'mau_ure', 'mau_creatinin',
    'mau_sgot', 'mau_alat',
    'nuocTieu_manv', 'nuoc_tieu', 'nuoc_tieu_khac',
    'cdha_manv', 'cdha',
    'cls_khac_manv', 'cls_khac',
    'cls_manv', 'cls_ket_qua',
    'kl_manv', 'kl_phan_loai', 'kl_ten_benh', 'kl_ngay', 'isDangLam', 'ailam',
    'daky', 'kl_benh', 'doi_tuong',
]
assert len(HIS_COLUMNS) == 119, f'HIS_COLUMNS phải đúng 119 cột, đang có {len(HIS_COLUMNS)}'

# 4 cột phụ thêm SAU 119 cột chính — để user VLOOKUP makcb rồi xóa đi.
HIS_EXTRA_COLUMNS = ['phu_ma_ho_so', 'phu_ho_ten', 'phu_so_cccd', 'phu_ngay_sinh']

# Thứ tự ĐÚNG 22 cờ tiền sử bản thân — khớp tien_su_ban_than_1..22.
_TSBT_ORDER = (
    'tsbt_benh_trong_5_nam_qua', 'tsbt_benh_than_kinh', 'tsbt_benh_mat',
    'tsbt_benh_tai', 'tsbt_benh_tim', 'tsbt_phau_thuat_tim',
    'tsbt_tang_huyet_ap', 'tsbt_kho_tho', 'tsbt_benh_phoi', 'tsbt_benh_than',
    'tsbt_nghien_ruou', 'tsbt_dai_thao_duong', 'tsbt_benh_tam_than',
    'tsbt_mat_y_thuc', 'tsbt_ngat', 'tsbt_benh_tieu_hoa',
    'tsbt_roi_loan_giac_ngu', 'tsbt_tai_bien', 'tsbt_benh_cot_song',
    'tsbt_ruou_thuong_xuyen', 'tsbt_ma_tuy', 'tsbt_benh_khac',
)
assert len(_TSBT_ORDER) == 22

# 11 khối cơ quan (mã cột HIS -> field text/pl trong ho_so) dựng theo mẫu
# tuan_hoan_manv/tuan_hoan/tuan_hoan_pl — dùng chung cho vòng lặp cột VÀ để
# dựng kl_benh (cộng thêm mắt/TMH/RHM ở _KL_BLOCKS bên dưới).
_ORGAN_TEXT_BLOCKS = (
    ('tuan_hoan', 'noi_khoa_tuan_hoan', 'noi_khoa_tuan_hoan_pl'),
    ('ho_hap', 'noi_khoa_ho_hap', 'noi_khoa_ho_hap_pl'),
    ('tieu_hoa', 'noi_khoa_tieu_hoa', 'noi_khoa_tieu_hoa_pl'),
    ('than_tiet_nieu', 'noi_khoa_than_tn_sd', 'noi_khoa_than_tietnieu_pl'),
    ('noi_tiet', 'noi_khoa_noi_tiet', 'noi_khoa_noi_tiet_pl'),
    ('co_xuong_khop', 'noi_khoa_co_xuong_khop', 'noi_khoa_co_xuong_khop_pl'),
    ('than_kinh', 'noi_khoa_than_kinh', 'noi_khoa_than_kinh_pl'),
    ('tam_than', 'noi_khoa_tam_than', 'noi_khoa_tam_than_pl'),
    ('ngoai_khoa', 'ket_qua_kham_ngoai_khoa', 'kham_ngoai_khoa_pl'),
    ('da_lieu', 'ket_qua_kham_da_lieu', 'kham_da_lieu_pl'),
    ('san', 'ket_qua_kham_san_phu_khoa', 'kham_san_phu_khoa_pl'),
)

# 14 khối cơ quan (nhãn hiển thị -> field text/pl) để dựng cột 118 kl_benh —
# 11 khối nội khoa ở trên + mắt/TMH/RHM (khác cấu trúc cột nên liệt kê riêng).
_KL_BLOCKS = (
    ('Tuần hoàn', 'noi_khoa_tuan_hoan', 'noi_khoa_tuan_hoan_pl'),
    ('Hô hấp', 'noi_khoa_ho_hap', 'noi_khoa_ho_hap_pl'),
    ('Tiêu hóa', 'noi_khoa_tieu_hoa', 'noi_khoa_tieu_hoa_pl'),
    ('Thận-Tiết niệu', 'noi_khoa_than_tn_sd', 'noi_khoa_than_tietnieu_pl'),
    ('Nội tiết', 'noi_khoa_noi_tiet', 'noi_khoa_noi_tiet_pl'),
    ('Cơ xương khớp', 'noi_khoa_co_xuong_khop', 'noi_khoa_co_xuong_khop_pl'),
    ('Thần kinh', 'noi_khoa_than_kinh', 'noi_khoa_than_kinh_pl'),
    ('Tâm thần', 'noi_khoa_tam_than', 'noi_khoa_tam_than_pl'),
    ('Ngoại khoa', 'ket_qua_kham_ngoai_khoa', 'kham_ngoai_khoa_pl'),
    ('Da liễu', 'ket_qua_kham_da_lieu', 'kham_da_lieu_pl'),
    ('Sản', 'ket_qua_kham_san_phu_khoa', 'kham_san_phu_khoa_pl'),
    ('Mắt', 'benh_khac_mat', 'kham_mat_pl'),
    ('Tai mũi họng', 'benh_khac_tai_mui_hong', 'kham_tai_mui_hong_pl'),
    ('Răng hàm mặt', 'benh_khac_rang_ham_mat', 'kham_rang_ham_mat_pl'),
)

# 4 cột thính lực HIS (m) — cùng cơ chế đổi dấu thập phân '.'->',' như
# _doi_dau_thap_phan() ở trên, áp dụng cho TỪNG giá trị đơn lẻ.
_TRUONG_THINH_LUC_HIS = ('tai_trai_noi_thuong', 'tai_trai_noi_tham',
                         'tai_phai_noi_thuong', 'tai_phai_noi_tham')


def _his_null(v):
    """Ô trống -> chuỗi "NULL" (khớp file mẫu HIS thật), ngược lại giữ nguyên."""
    return 'NULL' if v in (None, '') else v


def _his_pl(v):
    """Quy tắc cột *_pl (phân loại 1..5): 0/None/'' -> "NULL", ngược lại số."""
    return 'NULL' if v in (None, 0, '') else v


def _his_flag(v):
    """22 cờ tiền sử + đang điều trị/thai sản: ho_so lưu chuỗi 'Có'/'Không'
    (KHÔNG phải 1/'1'/True như mô tả ban đầu — đã kiểm tra thực tế DB, mọi
    giá trị chỉ là 'Có'/'Không'/rỗng) -> "1" nếu 'Có', ngược lại "NULL"."""
    return '1' if v == 'Có' else 'NULL'


def _his_thinh_luc(v):
    if v in (None, ''):
        return 'NULL'
    s = str(v)
    return s.replace('.', ',') if '.' in s else s


def _his_split_huyet_ap(v):
    if v in (None, ''):
        return 'NULL', 'NULL'
    m = re.search(r'(\d+)\s*/\s*(\d+)', str(v))
    if not m:
        return 'NULL', 'NULL'
    return m.group(1), m.group(2)


def _his_cls_khac(kq_sieu_am, kq_dien_tim):
    parts = []
    if kq_sieu_am not in (None, ''):
        parts.append(f'Siêu âm: {kq_sieu_am}')
    if kq_dien_tim not in (None, ''):
        parts.append(f'Điện tim: {kq_dien_tim}')
    return '\n'.join(parts) if parts else 'NULL'


def _his_kl_benh(row):
    lines = []
    for nhan, text_field, pl_field in _KL_BLOCKS:
        text = row[text_field]
        pl = row[pl_field]
        bat_thuong_pl = pl not in (None, 0, '') and int(pl) >= 2
        bat_thuong_text = (text not in (None, '')
                           and str(text).strip().lower() != 'bình thường')
        if bat_thuong_pl or bat_thuong_text:
            lines.append(f'{nhan}: {text if text not in (None, "") else ""}')
    return '\n'.join(lines) if lines else 'NULL'


def _his_row(row):
    """Dựng 1 dòng 119 giá trị theo ĐÚNG thứ tự HIS_COLUMNS từ 1 dòng ho_so."""
    rec = {
        'makcb': '',                                     # để trống thật, user tự dán
        'soksk': 'NULL',
        'ngay_lap_phieu': _his_null(row['ngay_vao']),
        'nguon_kinh_phi': 'NULL',
        'nhom_mau': _his_null(row['nhom_mau']),
        'ly_do_ksk': _his_null(row['ly_do_vv']),
        'tsgd': _his_null(row['tsgd_mac_benh']),
        'tsgd_ma_benh': _his_null(row['tsgd_ma_benh']),
        'benh_tat_khac': _his_null(row['tsbt_ma_benh_khac']),
        'dang_dieu_tri': _his_flag(row['tsbt_dang_dieu_tri_benh']),
        'dang_dieu_tri_benh': _his_null(row['tsbt_ma_benh']),
        'thai_san': _his_flag(row['tsbt_thai_san']),
        'thai_san_benh': _his_null(row['tsbt_ma_benh_thai_san']),
        'chieu_cao': _his_null(row['chieu_cao']),
        'can_nang': _his_null(row['can_nang']),
        'bmi': _his_null(row['chi_so_bmi']),
        'mach': _his_null(row['mach']),
        'kham_the_luc_pl': _his_pl(row['kham_the_luc_pl']),
        'mat_manv': 'NULL',
        'mat_kk_mat_phai': _his_null(row['khong_kinh_mat_phai']),
        'mat_kk_mat_trai': _his_null(row['khong_kinh_mat_trai']),
        'mat_ck_mat_phai': _his_null(row['co_kinh_mat_phai']),
        'mat_ck_mat_trai': _his_null(row['co_kinh_mat_trai']),
        'mat_benh': _his_null(row['benh_khac_mat']),
        'mat_pl': _his_pl(row['kham_mat_pl']),
        'tmh_manv': 'NULL',
        'tmh_tai_trai_noi_thuong': _his_thinh_luc(row['tai_trai_noi_thuong']),
        'tmh_tai_trai_noi_tham': _his_thinh_luc(row['tai_trai_noi_tham']),
        'tmh_tai_phai_noi_thuong': _his_thinh_luc(row['tai_phai_noi_thuong']),
        'tmh_tai_phai_noi_tham': _his_thinh_luc(row['tai_phai_noi_tham']),
        'tmh_benh': _his_null(row['benh_khac_tai_mui_hong']),
        'tmh_pl': _his_pl(row['kham_tai_mui_hong_pl']),
        'rhm_manv': 'NULL',
        'rhm_ham_tren': _his_null(row['ham_tren']),
        'rhm_ham_duoi': _his_null(row['ham_duoi']),
        'rhm_benh': _his_null(row['benh_khac_rang_ham_mat']),
        'rhm_pl': _his_pl(row['kham_rang_ham_mat_pl']),
        'mau_manv': 'NULL', 'huyet_hoc': 'NULL',
        'mau_duong_mau': _his_null(row['glu_gia_tri']),
        'mau_ure': 'NULL', 'mau_creatinin': 'NULL',
        'mau_sgot': 'NULL', 'mau_alat': 'NULL',
        'nuocTieu_manv': 'NULL', 'nuoc_tieu': 'NULL', 'nuoc_tieu_khac': 'NULL',
        'cdha_manv': 'NULL', 'cdha': 'NULL',
        'cls_khac_manv': 'NULL',
        'cls_khac': _his_cls_khac(row['kq_sieu_am_o_bung'], row['kq_dien_tim']),
        'cls_manv': 'NULL', 'cls_ket_qua': 'NULL',
        'kl_manv': 'NULL',
        'kl_phan_loai': _his_pl(row['phan_loai_sk']),
        'kl_ten_benh': _his_null(row['ket_luan_benh']),
        'kl_ngay': _his_null(row['ngay_vao']),
        'isDangLam': 'NULL', 'ailam': 'NULL', 'daky': 'NULL',
        'kl_benh': _his_kl_benh(row),
        'doi_tuong': _his_null(row['doi_tuong']),
    }
    for i, field in enumerate(_TSBT_ORDER, 1):
        rec[f'tien_su_ban_than_{i}'] = _his_flag(row[field])
    sys_v, dia_v = _his_split_huyet_ap(row['huyet_ap'])
    rec['huyet_ap_tam_thu'] = sys_v
    rec['huyet_ap_tam_truong'] = dia_v
    for code_prefix, text_field, pl_field in _ORGAN_TEXT_BLOCKS:
        rec[f'{code_prefix}_manv'] = 'NULL'
        rec[code_prefix] = _his_null(row[text_field])
        rec[f'{code_prefix}_pl'] = _his_pl(row[pl_field])
    return [rec[c] for c in HIS_COLUMNS]


def _his_xlsx_bytes(rows):
    """Dựng .xlsx write_only: 1 sheet 'Table1' — dòng 1 = header (119 mã cột
    HIS + 4 mã cột phụ), dữ liệu từ dòng 2. Cột phụ tô vàng nhạt để phân biệt
    với 119 cột chính (khớp mã HIS, KHÔNG được sửa thứ tự/tên)."""
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Table1')
    ws.freeze_panes = 'A2'
    bold = Font(bold=True)
    phu_fill = PatternFill('solid', fgColor='FFF2CC')

    header_cells = []
    for code in HIS_COLUMNS:
        c = WriteOnlyCell(ws, value=code)
        c.font = bold
        header_cells.append(c)
    for code in HIS_EXTRA_COLUMNS:
        c = WriteOnlyCell(ws, value=code)
        c.font = bold
        c.fill = phu_fill
        header_cells.append(c)
    ws.append(header_cells)

    for row in rows:
        cells = [WriteOnlyCell(ws, value=v) for v in _his_row(row)]
        for v in (row['ma_ho_so'], row['ho_ten'], row['so_cccd'] or '',
                 row['ngay_sinh'] or ''):
            c = WriteOnlyCell(ws, value=v)
            c.fill = phu_fill
            cells.append(c)
        ws.append(cells)

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue(), len(rows)


def build_his_xlsx(conn, pham_vi, gia_tri, include_errors, chi_rs_xong=False):
    """Xuất .xlsx đúng định dạng import phần mềm HIS (khác hẳn build_plain_xlsx/
    build_xlsm — mẫu BYT). Tái dùng CÙNG resolve_scope_where + _apply_rs_xong +
    _red_flag_where như build_plain_xlsx nên bộ lọc phạm vi/tuỳ chọn hoạt
    động giống hệt nút "Tải .xlsx (kèm mã định danh)". Ném ValueError nếu
    phạm vi rỗng — router chuyển thành 400. Trả (bytes, count)."""
    where_sql, args = resolve_scope_where(pham_vi, gia_tri)
    where_sql = _apply_rs_xong(where_sql, chi_rs_xong)
    rows = conn.execute(
        f'SELECT * FROM ho_so WHERE {where_sql} ORDER BY maxa_cu_tru, tt',
        args).fetchall()
    if not rows:
        raise ValueError('Không có hồ sơ nào khớp phạm vi đã chọn')
    if not include_errors:
        red_sql, red_args = _red_flag_where()
        red_set = {r['ma_ho_so'] for r in conn.execute(
            f'SELECT ma_ho_so FROM ho_so WHERE {where_sql} AND {red_sql}',
            args + red_args)}
        rows = [r for r in rows if r['ma_ho_so'] not in red_set]
        if not rows:
            raise ValueError('Toàn bộ hồ sơ trong phạm vi đều còn cờ đỏ — '
                             'bật "Xuất kèm cả hồ sơ lỗi" nếu vẫn muốn xuất')
    return _his_xlsx_bytes(rows)


# ============================= XÂY BẢN GHI =============================
def _merged_filename(so_ca):
    """Đợt 18 (phản hồi anh Khôi): GỘP 1 file .xlsm duy nhất theo đúng phạm
    vi đã chọn — không tách theo xã nữa (xem ghi chú tại _run_job về việc
    kiểm chứng lại giới hạn bộ nhớ ở §7.1 mục 7 SPEC)."""
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    return f'KSK_Import_{so_ca}ca_{ts}.xlsm'


def _row_to_rec(row, tt):
    """Đảo ngược import_data.py: tên cột `ho_so` viết hoa == mã trường BYT
    (đã kiểm chứng field-by-field với dòng 2 của template — xem docstring
    module). write_xlsm() tự bỏ qua các khoá không khớp mã trường mẫu."""
    rec = _doi_dau_thap_phan({k.upper(): row[k] for k in row.keys()})
    rec['TT'] = tt
    return rec


def _row_ext(row, ho_ten_ra_soat, danh_sach_map):
    """Giá trị cho các cột mở rộng §7.2 — chỉ dùng khi user bật option.
    danh_sach_map: {ma_ho_so: [tên danh sách,...]} (đợt xuất — xem create_job),
    dùng để điền cột DANH_SACH — rỗng nếu hồ sơ không thuộc danh sách nào."""
    return {
        'MA_BENH_CHINH': row['ma_benh_chinh'],
        'MA_BENH_KEM': row['ma_benh_kem'],
        'TEN_BENH_KEM': row['ten_benh_kem'],
        'CO_QUAN_BENH_CHINH': qc.TEN_CQ.get(row['co_quan_benh_chinh'], row['co_quan_benh_chinh']),
        'CHAN_DOAN_GOC': row['chan_doan_goc'],
        'CO_QC': row['co_qc'],
        'SO_LOI': row['so_loi'],
        'GHI_CHU_RA_SOAT': row['ghi_chu_ra_soat'],
        'GHI_CHU_CAN_BO': row['ghi_chu_can_bo'],
        'NGUOI_RA_SOAT': ho_ten_ra_soat,
        'TRANG_THAI': TRANG_THAI_NHAN.get(row['trang_thai'], row['trang_thai']),
        'THOI_DIEM_HOAN_THANH': row['thoi_diem_hoan_thanh'],
        'GLU_THOI_DIEM': row['glu_thoi_diem'],
        'GLU_GIA_TRI': row['glu_gia_tri'],
        'KQ_DIEN_TIM': row['kq_dien_tim'],
        'KQ_SIEU_AM_O_BUNG': row['kq_sieu_am_o_bung'],
        'MA_HO_SO': row['ma_ho_so'],
        'DANH_SACH': ', '.join(danh_sach_map.get(row['ma_ho_so'], [])),
    }


# ============================= JOB =============================
def _save_job(job):
    path = os.path.join(job['job_dir'], 'job.json')
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(job, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _log(job, msg):
    line = f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}"
    job['log'].append(line)
    _save_job(job)


def create_job(pham_vi, gia_tri, include_errors, extended, admin_id,
               chi_rs_xong=False):
    """Chuẩn bị job (đọc DB, gom theo xã, đếm cờ đỏ) rồi khởi động thread nền.
    Ném ValueError nếu phạm vi rỗng/không hợp lệ — router chuyển thành 400."""
    conn = db.get_connection()
    try:
        where_sql, args = resolve_scope_where(pham_vi, gia_tri)
        where_sql = _apply_rs_xong(where_sql, chi_rs_xong)
        rows = conn.execute(
            f'SELECT * FROM ho_so WHERE {where_sql} ORDER BY maxa_cu_tru, tt',
            args).fetchall()
        red_sql, red_args = _red_flag_where()
        red_set = {r['ma_ho_so'] for r in conn.execute(
            f'SELECT ma_ho_so FROM ho_so WHERE {where_sql} AND {red_sql}',
            args + red_args)}
        user_map = {r['id']: r['ho_ten'] for r in
                    conn.execute('SELECT id, ho_ten FROM nguoi_dung')}
        # Cột mở rộng DANH_SACH (criterion 1.4): mỗi hồ sơ trong `rows` thuộc
        # những danh sách tùy chỉnh nào — 1 TRUY VẤN PHỤ DUY NHẤT theo mã hồ
        # sơ của TOÀN BỘ phạm vi (không N+1), khuôn mẫu ho_so.py dòng ~544-553.
        ma_list = [r['ma_ho_so'] for r in rows]
        danh_sach_map = {}
        if ma_list:
            ph_ds = ','.join('?' * len(ma_list))
            for r_ds in conn.execute(
                    f'SELECT dsh.ma_ho_so AS ma_ho_so, ds.ten AS ten '
                    f'FROM danh_sach_ho_so dsh JOIN danh_sach ds ON ds.id = dsh.danh_sach_id '
                    f'WHERE dsh.ma_ho_so IN ({ph_ds})', ma_list):
                danh_sach_map.setdefault(r_ds['ma_ho_so'], []).append(r_ds['ten'])
    finally:
        conn.close()

    if not rows:
        raise ValueError('Không có hồ sơ nào khớp phạm vi đã chọn')

    ext_enabled = bool(extended and extended.get('enabled'))
    ext_columns = []
    if ext_enabled:
        ext_columns = [c for c in (extended.get('columns') or []) if c in EXTENDED_CODES]

    do_flag_count = sum(1 for r in rows if r['ma_ho_so'] in red_set)
    se_loai_tru = 0 if include_errors else do_flag_count

    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    job_id = f'{ts}_{uuid.uuid4().hex[:6]}'
    job_dir = os.path.join(EXPORTS_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)

    job = {
        'id': job_id, 'status': 'queued',
        'created_at': datetime.datetime.now().isoformat(),
        'admin_id': admin_id,
        'params': {'pham_vi': pham_vi, 'gia_tri': gia_tri,
                    'include_errors': include_errors, 'chi_rs_xong': chi_rs_xong,
                    'extended': {'enabled': ext_enabled, 'columns': ext_columns}},
        'tong_pham_vi': len(rows), 'do_flag_count': do_flag_count,
        'se_xuat': len(rows) - se_loai_tru, 'se_loai_tru': se_loai_tru,
        'progress_percent': 0,
        'log': [], 'files': [], 'error': None, 'job_dir': job_dir,
    }
    with _JOBS_LOCK:
        _JOBS[job_id] = job
    _save_job(job)

    t = threading.Thread(target=_run_job,
                          args=(job_id, rows, red_set, user_map, include_errors,
                                ext_enabled, ext_columns, danh_sach_map),
                          daemon=True)
    t.start()
    return job


def _run_job(job_id, rows, red_set, user_map, include_errors, ext_enabled, ext_columns,
             danh_sach_map):
    """Đợt 18 (phản hồi anh Khôi): GỘP 1 file .xlsm duy nhất theo đúng phạm vi
    đã chọn — không tách theo xã/phường nữa như trước. §7.1 mục 7 SPEC từng
    ghi "gộp 13.326 ca vào 1 file làm openpyxl vượt bộ nhớ" — đã ĐO LẠI THẬT
    trên máy hiện tại (write_xlsm với TOÀN BỘ 13.326 ca thật trong DB): peak
    ~840MB RAM, ~15s, chạy trót lọt — với openpyxl/phần cứng hiện tại, giới
    hạn cũ không còn đúng. Vẫn chạy trong 1 SUBPROCESS riêng (không phải
    ngay tiến trình server) để nếu có bất ngờ thật sự xảy ra, chỉ tiến trình
    con bị ảnh hưởng, server chính không bị kéo theo."""
    job = _JOBS[job_id]
    job['status'] = 'running'
    _log(job, f"Bắt đầu xuất {len(rows)} hồ sơ trong phạm vi đã chọn "
              f"(cờ đỏ: {job['do_flag_count']}, "
              f"{'gồm cả hồ sơ lỗi' if include_errors else 'loại trừ hồ sơ lỗi'})")

    included_rows, skip_rows = [], []
    for r in rows:
        is_red = r['ma_ho_so'] in red_set
        if is_red and not include_errors:
            skip_rows.append(r)
        else:
            included_rows.append(r)
    ke_records = [(r, False, None) for r in skip_rows]

    exported_ma = []
    had_error = False

    if not included_rows:
        job['status'] = 'error'
        job['error'] = 'Không có hồ sơ nào để xuất — toàn bộ đã bị loại do cờ 🔴.'
        _log(job, job['error'])
        _save_job(job)
        return

    _log(job, f"Đang ghi {len(included_rows)} ca vào 1 file .xlsm gộp ...")

    recs = []
    for i, r in enumerate(included_rows, 1):
        rec = _row_to_rec(r, i)
        if ext_enabled:
            rec['_EXT'] = _row_ext(r, user_map.get(r['nguoi_ra_soat_id'], ''), danh_sach_map)
        recs.append(rec)

    filename = _merged_filename(len(included_rows))
    output_path = os.path.join(job['job_dir'], filename)
    handoff_path = os.path.join(job['job_dir'], '.handoff.json')
    with open(handoff_path, 'w', encoding='utf-8') as f:
        json.dump({'records': recs,
                   'extended': {'enabled': ext_enabled, 'columns': ext_columns,
                                'labels': EXTENDED_LABELS}},
                  f, ensure_ascii=False)

    # % tiến trình chỉ là ƯỚC TÍNH theo thời gian đã trôi qua so với thời
    # gian dự kiến (hiệu chỉnh thực đo: ~3.5s cố định mở template+danh mục +
    # ~0.85ms/dòng ghi 103 cột; cột mở rộng nếu bật mở lại toàn file 1 lần
    # nữa, ~3s + ~1.8ms/dòng) — write_xlsm() ở build/ không có hook báo tiến
    # độ thật và KHÔNG được sửa file đó (§9 SPEC: không viết lại pipeline).
    est_total_s = 3.5 + 0.00085 * len(included_rows)
    if ext_enabled:
        est_total_s += 3.0 + 0.0018 * len(included_rows)

    proc = subprocess.Popen(
        [sys.executable, WORKER_PATH, handoff_path, output_path],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        text=True, encoding='utf-8', errors='replace')
    t0 = time.time()
    while proc.poll() is None:
        job['progress_percent'] = min(95, int((time.time() - t0) / est_total_s * 100))
        _save_job(job)
        time.sleep(0.7)
    _, stderr = proc.communicate()

    try:
        os.remove(handoff_path)
    except OSError:
        pass

    if proc.returncode == 0:
        job['progress_percent'] = 100
        job['files'].append({'ten': filename, 'duong_dan': output_path, 'loai': 'xlsm'})
        for r in included_rows:
            exported_ma.append(r['ma_ho_so'])
            ke_records.append((r, True, filename))
        _log(job, f"Xong — {len(included_rows)} ca -> {filename}")
    else:
        had_error = True
        err_tail = (stderr or '').strip()[-800:]
        job['error'] = err_tail
        _log(job, f"LỖI — {err_tail}")
        for r in included_rows:
            ke_records.append((r, False, None))
    _save_job(job)

    _log(job, 'Đang tạo file kê ...')
    file_ke_path = os.path.join(job['job_dir'], 'file_ke.xlsx')
    _write_file_ke(ke_records, file_ke_path)
    job['files'].append({'ten': 'file_ke.xlsx', 'duong_dan': file_ke_path,
                          'loai': 'file_ke'})
    _save_job(job)

    if exported_ma:
        conn = db.get_connection()
        try:
            now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for ma in exported_ma:
                old = conn.execute('SELECT da_xuat_file FROM ho_so WHERE ma_ho_so=?',
                                    (ma,)).fetchone()
                conn.execute(
                    'UPDATE ho_so SET da_xuat_file=1, lan_xuat_cuoi=? WHERE ma_ho_so=?',
                    (now, ma))
                conn.execute(
                    'INSERT INTO nhat_ky(ma_ho_so, nguoi_dung_id, ten_truong, '
                    'gia_tri_cu, gia_tri_moi) VALUES (?,?,?,?,?)',
                    (ma, job['admin_id'], 'da_xuat_file',
                     str(old['da_xuat_file']) if old else '0', '1'))
            conn.commit()
        finally:
            conn.close()
        _log(job, f'Đã cập nhật da_xuat_file cho {len(exported_ma)} hồ sơ')

    job['status'] = 'error' if had_error else 'done'
    job['finished_at'] = datetime.datetime.now().isoformat()
    _log(job, 'Hoàn tất — có lỗi, xem chi tiết ở trên.' if had_error else 'Hoàn tất.')
    _save_job(job)


def _write_file_ke(records, path):
    """File kê .xlsx (§7.2/§7 mục 10) — workbook THƯỜNG (không phải file
    nộp Bộ), liệt kê MỌI hồ sơ trong lô + cờ còn lại + đã đưa vào file hay
    bị loại + tên file .xlsm chứa nó."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kê danh sách xuất'
    headers = ['Mã hồ sơ', 'Họ tên', 'Xã/phường', 'Trạng thái', 'Cờ còn lại',
               'Kết quả', 'Tên file .xlsm']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDEBF7')
    r = 2
    for row, included, filename in records:
        ws.cell(r, 1, row['ma_ho_so'])
        ws.cell(r, 2, row['ho_ten'])
        ws.cell(r, 3, row['maxa_cu_tru'])
        ws.cell(r, 4, TRANG_THAI_NHAN.get(row['trang_thai'], row['trang_thai']))
        ws.cell(r, 5, row['co_qc'] or '')
        ws.cell(r, 6, 'Đã đưa vào file' if included else 'Bị loại (còn cờ 🔴)')
        ws.cell(r, 7, filename or '')
        r += 1
    for col, w in zip('ABCDEFG', (20, 24, 20, 18, 40, 20, 30)):
        ws.column_dimensions[col].width = w
    wb.save(path)
    wb.close()


def get_job(job_id):
    with _JOBS_LOCK:
        if job_id in _JOBS:
            return _JOBS[job_id]
    path = os.path.join(EXPORTS_DIR, job_id, 'job.json')
    if os.path.isfile(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return None


def list_jobs():
    out = []
    if os.path.isdir(EXPORTS_DIR):
        for name in sorted(os.listdir(EXPORTS_DIR), reverse=True):
            if not os.path.isdir(os.path.join(EXPORTS_DIR, name)):
                continue
            job = get_job(name)
            if job:
                out.append(job)
    return out
