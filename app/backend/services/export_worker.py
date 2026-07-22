# -*- coding: utf-8 -*-
"""
export_worker.py — tiến trình CON, xuất 1 FILE .xlsm CHO 1 XÃ.

Chạy:  python3 export_worker.py <handoff_json> <output_xlsm_path>

Đọc danh sách bản ghi (đã ở dạng dict theo mã trường BYT, do
services/export_xlsm.py chuẩn bị) từ file JSON "handoff", gọi lại
build/build_xlsm.write_xlsm() — KHÔNG viết lại pipeline xuất (§9 SPEC) —
để ghi vào bản sao template (copy + keep_vba=True), giữ nguyên 100% data
validation/định dạng/VBA (bẫy §10: không tạo workbook mới).

Mỗi xã MỘT TIẾN TRÌNH RIÊNG (§7.1 mục 7): tiến trình cha
(services/export_xlsm.py) spawn subprocess này cho từng xã rồi chờ xong mới
sang xã kế — nhờ vậy hệ điều hành thu hồi bộ nhớ sau khi tiến trình con
thoát, tránh tích luỹ RAM qua nhiều xã (build_xlsm.write_xlsm tự
wb.close()+gc.collect(), nhưng openpyxl vẫn không trả hết bộ nhớ cho tiến
trình đang chạy — phải kết thúc hẳn tiến trình).
"""
import gc
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

config.ensure_build_on_path()
from build_xlsm import write_xlsm  # noqa: E402

FIRST_ROW = 4
EXT_START_COL = 104  # §7.2: cột mở rộng bắt đầu từ 104, KHÔNG động vào 1..103


def write_extended_columns(path, records, columns, labels):
    """Mở lại file VỪA GHI (keep_vba=True) và thêm cột mở rộng — chỉ gọi khi
    user bật option (mặc định TẮT). Header dòng 1 = nhãn tiếng Việt, dòng 2 =
    mã cột (giữ đúng quy ước của mẫu: dòng 1 nhãn, dòng 2 mã trường)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.load_workbook(path, keep_vba=True)
    ws = wb['Trên 18']
    warn_fill = PatternFill('solid', fgColor='FFF3CD')

    for j, code in enumerate(columns):
        col = EXT_START_COL + j
        label = labels.get(code, code)
        if j == 0:
            label = '⚠ CỘT MỞ RỘNG — KHÔNG NỘP BỘ — ' + label
        c1 = ws.cell(1, col, label)
        c1.font = Font(bold=True)
        c1.fill = warn_fill
        c2 = ws.cell(2, col, code)
        c2.font = Font(bold=True)
        c2.fill = warn_fill

    for i, rec in enumerate(records):
        r = FIRST_ROW + i
        ext = rec.get('_EXT') or {}
        for j, code in enumerate(columns):
            col = EXT_START_COL + j
            v = ext.get(code)
            ws.cell(r, col, v if v not in ('', None) else None)

    wb.save(path)
    wb.close()
    del wb, ws
    gc.collect()


def main():
    if len(sys.argv) != 3:
        print('Dùng: export_worker.py <handoff_json> <output_xlsm_path>',
              file=sys.stderr)
        sys.exit(2)
    handoff_path, output_path = sys.argv[1], sys.argv[2]

    with open(handoff_path, encoding='utf-8') as f:
        data = json.load(f)

    records = data['records']
    ext = data.get('extended') or {'enabled': False, 'columns': []}

    # write_xlsm() chỉ đọc các khoá khớp mã trường mẫu (1..103) + TT — khoá
    # nội bộ _EXT bị bỏ qua tự nhiên, nhưng loại bỏ tường minh cho rõ ràng.
    recs_for_template = [{k: v for k, v in r.items() if k != '_EXT'} for r in records]
    write_xlsm(recs_for_template, output_path)

    if ext.get('enabled') and ext.get('columns'):
        write_extended_columns(output_path, records, ext['columns'],
                                ext.get('labels') or {})

    print(f'OK {len(records)} ca -> {output_path}')


if __name__ == '__main__':
    main()
