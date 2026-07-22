# -*- coding: utf-8 -*-
"""
export_unmapped.py — Xuất danh sách chẩn đoán CHƯA ÁNH XẠ ĐƯỢC để anh Khôi bổ sung.

Hai mức độ:
  [1] CHƯA HIỂU        — không nhận diện được bệnh, đã BỎ QUA (mất thông tin)
  [2] CHỈ ĐOÁN CƠ QUAN — nhận ra thuộc cơ quan nào nhưng không rõ bệnh cụ thể,
                          đang gán mã "không đặc hiệu" (thông tin bị làm nghèo)

Chạy:
    python3 export_unmapped.py --sample 100
    python3 export_unmapped.py --all
"""
import argparse, os, sys, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as L

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from normalize import process_cell, concept_key, basic_clean
from icd_map import map_concept, ORGAN_FALLBACK
from build_import import SRC, OUT, parse_pl
from build_xlsm import diem_day_du

TEN_CQ = {
    'TH': 'Tuần hoàn', 'HH': 'Hô hấp', 'TIEUHOA': 'Tiêu hóa',
    'THAN': 'Thận-Tiết niệu-Sinh dục', 'NOITIET': 'Nội tiết',
    'CXK': 'Cơ-Xương-Khớp', 'TK': 'Thần kinh', 'TT': 'Tâm thần',
    'NGOAI': 'Ngoại khoa', 'DALIEU': 'Da liễu', 'SAN': 'Sản phụ khoa',
    'MAT': 'Mắt', 'TMH': 'Tai-Mũi-Họng', 'RHM': 'Răng-Hàm-Mặt',
}

# Phỏng đoán của máy cho các mẩu còn mơ hồ — anh Khôi chỉ cần XÁC NHẬN hoặc sửa.
# key: khóa sau chuẩn hóa (concept_key)
SUGGEST = {
    'sỏi': ('Nhiều khả năng là "Sỏi thận" — cùng ca có "nang thận 2 bên", '
            'người nhập gõ "Sỏi.nang thận 2 bên" nên bị tách rời',
            'N20.0', 'Thận-Tiết niệu-Sinh dục'),
    'mắt trái': ('Ghi thiếu — chỉ có tên bên mắt, không có bệnh. '
                 'Cần đối chiếu sổ khám gốc', '', 'Mắt'),
    'bàn tay p': ('Ghi thiếu — chỉ có vị trí, không có bệnh. '
                  'Cần đối chiếu sổ khám gốc', '', 'Cơ-Xương-Khớp'),
}

HDR = PatternFill('solid', fgColor='DDEBF7')
FILL_IN = PatternFill('solid', fgColor='FFF2CC')     # cột anh Khôi điền
MUC1 = PatternFill('solid', fgColor='FFC7CE')        # chưa hiểu
MUC2 = PatternFill('solid', fgColor='FFEB9C')        # chỉ đoán cơ quan
MUC3 = PatternFill('solid', fgColor='DDEBF7')        # máy tự suy đoán


MUC_TEN = {1: 'CHƯA HIỂU', 2: 'CHỈ ĐOÁN CƠ QUAN', 3: 'MÁY TỰ SUY ĐOÁN'}
MUC_NGUON = {'unmapped': 1, 'organ_fallback': 2, 'fuzzy': 3}


def quet(rows):
    """-> (chi_tiet, tan_suat_toan_bo)"""
    chi_tiet = []
    for idx, r in enumerate(rows, 1):
        raw = basic_clean(r[37] or '')
        if not raw:
            continue
        atoms = process_cell(raw)
        ok = []
        for a in atoms:
            ck = concept_key(a)
            if ck:
                ok.append((a, ck, map_concept(ck)))
        for a, ck, m in ok:
            if m['nguon'] == 'rule':
                continue
            # ngữ cảnh: các chẩn đoán KHÁC trong cùng ca đã hiểu được
            ctx = '; '.join(x[2]['ten_icd'] for x in ok
                            if x[0] != a and x[2]['nguon'] == 'rule')
            chi_tiet.append({
                'tt': idx,
                'ho_ten': basic_clean(r[3]).upper(),
                'xa': basic_clean(r[0]),
                'raw': raw,
                'atom': a,
                'ck': ck,
                'muc': MUC_NGUON[m['nguon']],
                'co_quan': TEN_CQ.get(m['co_quan'], ''),
                'icd_tam': m['icd'],
                'ten_tam': m['ten_icd'],
                'ctx': ctx,
                'neo': m.get('neo', ''),
                'do_giong': m.get('do_giong', ''),
            })
    return chi_tiet


def tan_suat_toan_bo():
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    rows = [r for r in wb['TONG HOP FULL'].iter_rows(min_row=3, values_only=True)
            if r[3]]
    wb.close()
    c = collections.Counter()
    for r in rows:
        for a in process_cell(r[37] or ''):
            ck = concept_key(a)
            if ck:
                c[ck] += 1
    return c, len(rows)


def xuat(chi_tiet, freq, path, tong_ca):
    wb = openpyxl.Workbook()

    # ================= SHEET 1: TỪ ĐIỂN CẦN BỔ SUNG =================
    ws = wb.active
    ws.title = '1. Từ điển cần bổ sung'
    cols = [
        ('Mức', 6), ('Mẩu chẩn đoán CHƯA ÁNH XẠ', 30),
        ('Khóa sau chuẩn hóa', 26), ('Số ca trong mẫu này', 10),
        ('Số lần trong TOÀN BỘ 13.326 ca', 12),
        ('Cơ quan đoán được', 20), ('Mã ICD đang gán tạm', 14),
        ('Tên ICD đang gán tạm', 30),
        ('Máy suy đoán theo khái niệm', 26), ('Độ giống', 8),
        ('Ví dụ chuỗi chẩn đoán gốc chứa mẩu này', 45),
        ('GỢI Ý CỦA MÁY (chỉ cần xác nhận)', 42),
        ('ICD gợi ý', 12),
        ('➜ NGHĨA ĐẦY ĐỦ (anh Khôi điền)', 30),
        ('➜ MÃ ICD-10 đúng', 14),
        ('➜ Cơ quan khám', 18),
        ('➜ Ghi chú / bỏ qua nếu không phải bệnh', 26),
    ]
    for c, (h, w) in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.fill = FILL_IN if h.startswith('➜') else HDR
        ws.column_dimensions[L(c)].width = w
    ws.freeze_panes = 'C2'
    ws.row_dimensions[1].height = 42

    gom = collections.defaultdict(lambda: {'n': 0, 'vd': set(), 'atom': ''})
    for d in chi_tiet:
        g = gom[(d['muc'], d['ck'])]
        g['n'] += 1
        g['atom'] = d['atom']
        g['co_quan'] = d['co_quan']
        g['icd_tam'] = d['icd_tam']
        g['ten_tam'] = d['ten_tam']
        g['neo'] = d.get('neo', '')
        g['do_giong'] = d.get('do_giong', '')
        if len(g['vd']) < 3:
            g['vd'].add(d['raw'])

    items = sorted(gom.items(),
                   key=lambda kv: (kv[0][0], -freq.get(kv[0][1], 0), -kv[1]['n']))
    for i, ((muc, ck), g) in enumerate(items):
        r = 2 + i
        sg = SUGGEST.get(ck, ('', '', ''))
        vals = [muc, g['atom'], ck, g['n'], freq.get(ck, 0),
                g['co_quan'], g['icd_tam'], g['ten_tam'],
                g.get('neo', ''), g.get('do_giong', ''),
                ' ⏐ '.join(sorted(g['vd'])), sg[0], sg[1], '', '', '', '']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c >= 14:
                cell.fill = FILL_IN
            elif c in (12, 13):
                cell.fill = PatternFill('solid', fgColor='E2EFDA')
        ws.cell(r, 1).fill = {1: MUC1, 2: MUC2, 3: MUC3}[muc]

    # ================= SHEET 2: CHI TIẾT THEO CA =================
    w2 = wb.create_sheet('2. Chi tiết theo ca')
    cols2 = [('TT ca', 6), ('Họ tên', 22), ('Xã/Phường', 18),
             ('CHẨN ĐOÁN ĐẦY ĐỦ (GỐC)', 55),
             ('Mẩu CHƯA ánh xạ được', 28), ('Mức', 6),
             ('Cơ quan đoán được', 18), ('Mã ICD gán tạm', 13),
             ('Các chẩn đoán khác trong ca (đã hiểu)', 55)]
    for c, (h, w) in enumerate(cols2, 1):
        cell = w2.cell(1, c, h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.fill = HDR
        w2.column_dimensions[L(c)].width = w
    w2.freeze_panes = 'E2'
    for i, d in enumerate(chi_tiet):
        r = 2 + i
        vals = [d['tt'], d['ho_ten'], d['xa'], d['raw'], d['atom'], d['muc'],
                d['co_quan'], d['icd_tam'], d['ctx']]
        for c, v in enumerate(vals, 1):
            cell = w2.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        w2.cell(r, 6).fill = {1: MUC1, 2: MUC2, 3: MUC3}[d['muc']]

    # ================= SHEET 3: HƯỚNG DẪN =================
    w3 = wb.create_sheet('3. Hướng dẫn')
    guide = [
        ('CÁCH DÙNG FILE NÀY', True),
        ('', False),
        ('Mục tiêu: bổ sung các chẩn đoán mà máy chưa đọc hiểu, để lần chạy sau '
         'ánh xạ được đầy đủ.', False),
        ('', False),
        ('MỨC 1 (nền đỏ) — CHƯA HIỂU:', True),
        ('   Máy không nhận ra đây là bệnh gì. Mẩu này đã bị BỎ QUA, '
         'thông tin bị mất khỏi file import.', False),
        ('   ➜ Ưu tiên xử lý trước.', False),
        ('', False),
        ('MỨC 2 (nền vàng) — CHỈ ĐOÁN ĐƯỢC CƠ QUAN:', True),
        ('   Máy biết thuộc cơ quan nào nhưng không rõ bệnh cụ thể, nên đang gán '
         'mã "không đặc hiệu".', False),
        ('   Dữ liệu KHÔNG mất nhưng bị làm nghèo (vd: "Bệnh của hệ tiêu hóa, '
         'không đặc hiệu" thay vì tên bệnh thật).', False),
        ('', False),
        ('MỨC 3 (nền xanh) — MÁY TỰ SUY ĐOÁN:', True),
        ('   Máy đoán đây là lỗi gõ của một chẩn đoán đã biết và tự sửa '
         '(cột "Máy suy đoán theo khái niệm" + "Độ giống").', False),
        ('   Dữ liệu ĐÃ được ánh xạ, anh chỉ cần LƯỚT XEM để bắt suy đoán sai.', False),
        ('', False),
        ('CÁCH ĐIỀN (sheet 1, các cột nền vàng có dấu ➜):', True),
        ('   • NGHĨA ĐẦY ĐỦ: viết đầy đủ chẩn đoán. VD "TMCT" -> "Thiếu máu cơ tim"', False),
        ('   • MÃ ICD-10: nếu anh biết mã thì điền, không thì để trống, '
         'em tra theo nghĩa đầy đủ.', False),
        ('   • CƠ QUAN KHÁM: một trong 14 nhóm — Tuần hoàn, Hô hấp, Tiêu hóa, '
         'Thận-Tiết niệu-Sinh dục, Nội tiết, Cơ-Xương-Khớp, Thần kinh, Tâm thần, '
         'Ngoại khoa, Da liễu, Sản phụ khoa, Mắt, Tai-Mũi-Họng, Răng-Hàm-Mặt.', False),
        ('   • GHI CHÚ: ghi "BỎ QUA" nếu mẩu đó không phải chẩn đoán '
         '(vd: trị số, ghi chú hành chính, lỗi gõ phím).', False),
        ('', False),
        ('LƯU Ý: cột "Số lần trong TOÀN BỘ" cho biết mẩu đó xuất hiện bao nhiêu lần '
         'trên cả 13.326 ca — hãy ưu tiên các dòng có số lớn.', False),
    ]
    w3.column_dimensions['A'].width = 110
    for i, (t, bold) in enumerate(guide, 1):
        cell = w3.cell(i, 1, t)
        cell.font = Font(bold=bold, size=11 if bold else 10)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(path)
    return len(items)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sample', type=int, default=100)
    ap.add_argument('--all', action='store_true')
    a = ap.parse_args()
    os.makedirs(OUT, exist_ok=True)

    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    rows = [r for r in wb['TONG HOP FULL'].iter_rows(min_row=3, values_only=True)
            if r[3]]
    wb.close()

    if a.all:
        sel, mode = rows, 'TOANBO'
    else:
        sel = sorted(rows, key=lambda r: -diem_day_du(r))[:a.sample]
        mode = f'MAU{a.sample}'

    print('Đang quét tần suất trên toàn bộ dữ liệu...')
    freq, tong = tan_suat_toan_bo()
    from icd_map import bat_fuzzy
    print('Dựng mỏ neo cho tầng khớp mờ:', bat_fuzzy(freq), 'khái niệm')
    chi_tiet = quet(sel)
    path = os.path.join(OUT, f'TU_DIEN_CAN_BO_SUNG_{mode}.xlsx')
    n_unique = xuat(chi_tiet, freq, path, tong)

    print(f'\n{len(sel)} ca đã quét:')
    for k in (1, 2, 3):
        print(f'  Mức {k} {MUC_TEN[k]:<18}: '
              f'{sum(1 for d in chi_tiet if d["muc"] == k):>5} lượt')
    print(f'  Gộp lại: {n_unique} mẩu cần bổ sung')
    print(f'  Số ca bị ảnh hưởng: {len(set(d["tt"] for d in chi_tiet))}/{len(sel)}')
    print('\nFile:', path)


if __name__ == '__main__':
    main()
