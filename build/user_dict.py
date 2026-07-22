# -*- coding: utf-8 -*-
"""
user_dict.py — Nạp TỪ ĐIỂN DO ANH KHÔI DIỄN GIẢI.

Nguồn: output/TU_DIEN_CAN_BO_SUNG_TOANBO_da_dien_giai.xlsx, sheet 1.
  cột C  Khóa sau chuẩn hóa   (khớp với normalize.concept_key)
  cột F  Cơ quan đoán được    -> CƠ QUAN CHÍNH THỨC (ưu tiên tuyệt đối)
  cột N  Nghĩa đầy đủ         -> đưa lại qua bộ rule ICD để lấy mã

QUY ƯỚC: cột N TRỐNG = BỎ QUA khái niệm đó (thiếu dữ kiện, hoặc đã được
diễn giải ở mục khác). Không ánh xạ, không cảnh báo.
"""
import os
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PATH = os.path.join(BASE, 'output',
                    'TU_DIEN_CAN_BO_SUNG_TOANBO_da_dien_giai.xlsx')

# tên cơ quan (cột F) -> mã cơ quan nội bộ
CQ_MA = {
    'tuần hoàn': 'TH', 'hô hấp': 'HH', 'tiêu hóa': 'TIEUHOA',
    'thận-tiết niệu-sinh dục': 'THAN', 'thận - tiết niệu - sinh dục': 'THAN',
    'nội tiết': 'NOITIET', 'cơ-xương-khớp': 'CXK', 'cơ - xương - khớp': 'CXK',
    'thần kinh': 'TK', 'tâm thần': 'TT', 'ngoại khoa': 'NGOAI',
    'da liễu': 'DALIEU', 'sản phụ khoa': 'SAN', 'mắt': 'MAT',
    'tai-mũi-họng': 'TMH', 'tai - mũi - họng': 'TMH',
    'răng-hàm-mặt': 'RHM', 'răng - hàm - mặt': 'RHM',
}

_DICT = None      # {khóa -> {'nghia', 'co_quan', 'bo_qua'}}


def load(path=None):
    global _DICT
    if _DICT is not None:
        return _DICT
    p = path or PATH
    _DICT = {}
    if not os.path.exists(p):
        return _DICT
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb['1. Từ điển cần bổ sung']
    for r in range(2, ws.max_row + 1):
        khoa = ws.cell(r, 3).value
        if not khoa:
            continue
        khoa = str(khoa).strip().lower()
        nghia = ws.cell(r, 14).value
        cq_ten = str(ws.cell(r, 6).value or '').strip().lower()
        nghia = str(nghia).strip() if nghia not in (None, '') else ''
        _DICT[khoa] = {
            'nghia': nghia,
            'co_quan': CQ_MA.get(cq_ten, ''),
            'bo_qua': nghia == '',      # cột N trống -> bỏ qua
        }
    wb.close()
    return _DICT


def tra(concept):
    """-> dict hoặc None."""
    return load().get(str(concept).strip().lower())


def thong_ke():
    d = load()
    return {
        'tong': len(d),
        'co_nghia': sum(1 for v in d.values() if not v['bo_qua']),
        'bo_qua': sum(1 for v in d.values() if v['bo_qua']),
        'co_co_quan': sum(1 for v in d.values() if v['co_quan']),
    }
