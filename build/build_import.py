# -*- coding: utf-8 -*-
"""
build_import.py — Sinh file import chuẩn Bộ Y tế từ tong-hop.xlsx

Chạy:
    python3 build_import.py --sample 100     # sinh mẫu kiểm tra
    python3 build_import.py --all            # chạy toàn bộ

Đầu ra (thư mục ../output/):
    KSK_Import_<mode>.xlsx      sheet 'Trên 18' + 'DM_CanLamSang'
    KSK_DoiChieu_<mode>.xlsx    bảng đối chiếu chuỗi gốc <-> kết quả chuẩn hóa
    CAN_RA_SOAT.xlsx            các ca lỗi CCCD / cảnh báo, KHÔNG đưa vào file nộp
"""
import argparse, os, re, sys, random, datetime, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from normalize import process_cell, concept_key, basic_clean
from icd_map import map_concept
from classify import classify_person, ORGAN_COLS, ORGANS, BINH_THUONG

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE, 'tong-hop.xlsx')
TPL = os.path.join(BASE, 'doc', 'Import_KSK_Tren 18.xlsm')
OUT = os.path.join(BASE, 'output')

# ======================= THAM SỐ CẤU HÌNH =======================
# Anh Khôi chỉnh trực tiếp tại đây nếu cần.
CFG = {
    'MA_CSKCB':      '31006',                    # TTYT Thủy Nguyên
    'MA_GTIN_CSKCB': '',                         # mã GLN 13 ký tự — chưa có
    'TINH':          'Thành phố Hải Phòng',
    'DAN_TOC':       'Kinh',
    'NGHE_NGHIEP':   'Không xác định',
    'DOI_TUONG':     'Người cao tuổi',
    'NGUON_CHI_TRA': 'Ngân sách Địa phương',
    'LOAI_KCB':      'Khám sức khoẻ định kỳ',
    'LY_DO':         'Khám sức khỏe định kỳ người cao tuổi',
    'NOI_CAP_CCCD':  'Cục Cảnh sát QLHC về TTXH',
    'NAM_KHAM':      2026,
}

# Ánh xạ tên xã trong nguồn -> tên đơn vị hành chính đầy đủ
XA_MAP = {
    '1. Thủy Nguyên': 'Phường Thủy Nguyên',
    '2. Hòa Bình':    'Phường Hòa Bình',
    '3. Thiên Hương': 'Phường Thiên Hương',
    '4. Lưu Kiếm':    'Xã Lưu Kiếm',
    '5. Bạch Đằng':   'Phường Bạch Đằng',
    '6. Nam Triệu':   'Xã Nam Triệu',
    '7. Lê Ích Mộc':  'Xã Lê Ích Mộc',
    '8. Việt Khê':    'Xã Việt Khê',
}

# ======================= CHUẨN HÓA TRƯỜNG =======================
def fmt_date(v):
    """-> dd/mm/yyyy dạng TEXT."""
    if v is None or v == '':
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$', s)
    if m:
        return f'{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}'
    m = re.match(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$', s)
    if m:
        return f'{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}'
    return ''


def clean_cccd(v):
    """
    -> (cccd_text, ghi_chu). CCCD phải là TEXT 12 chữ số.
    Sửa được lỗi định dạng rõ ràng; không sửa được thì trả cờ lỗi.
    """
    if v is None or str(v).strip() == '':
        return '', 'THIẾU CCCD'
    s = str(v).strip()
    # số bị Excel đọc thành float: 3.1042001109e+10
    if re.match(r'^\d+\.\d+e\+\d+$', s, re.I):
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    if s.endswith('.0'):
        s = s[:-2]
    digits = re.sub(r'\D', '', s)
    if not digits:
        return '', 'CCCD KHÔNG HỢP LỆ: ' + str(v)
    if len(digits) == 12:
        return digits, ''
    if len(digits) == 11:                      # mất số 0 đầu do định dạng số
        return '0' + digits, 'Đã thêm số 0 đầu'
    if len(digits) == 9:                       # CMND cũ 9 số
        return digits, 'CMND 9 số (chưa đổi CCCD)'
    if len(digits) > 12:                       # dính ngày cấp vào sau
        return digits[:12], 'Đã cắt lấy 12 số đầu, cần kiểm tra'
    return digits, f'CCCD {len(digits)} số — cần rà soát'


def clean_gioi(v):
    s = basic_clean(v).lower().replace('ũ', 'ữ').replace('ư', 'ữ')
    if s.startswith('na'):
        return 'Nam'
    if s.startswith('n'):
        return 'Nữ'
    return 'Chưa xác định'


def clean_nam_sinh(nam, tuoi):
    """-> (ngay_sinh 'dd/mm/yyyy', ghi_chu). Quy ước 01/01/yyyy."""
    y = None
    try:
        y = int(float(nam))
    except (TypeError, ValueError):
        pass
    note = ''
    if y is None or not (1900 <= y <= 2010):
        try:
            t = int(float(tuoi))
            if 18 <= t <= 120:
                y = CFG['NAM_KHAM'] - t
                note = f'Năm sinh nguồn không hợp lệ ({nam}) — suy từ tuổi {t}'
        except (TypeError, ValueError):
            pass
    if y is None or not (1900 <= y <= 2010):
        return '', f'KHÔNG XÁC ĐỊNH ĐƯỢC NĂM SINH (nguồn: {nam})'
    return f'01/01/{y}', note or 'Ngày/tháng ước lượng (nguồn chỉ có năm sinh)'


def clean_noi_cap(v):
    if not v or not str(v).strip():
        return ''
    s = str(v).strip().lower()
    if 'cục' in s or s.startswith('cs') or s.startswith('ccs'):
        return CFG['NOI_CAP_CCCD']
    if 'bộ ca' in s or 'bộ công an' in s:
        return 'Bộ Công an'
    return basic_clean(v)


def parse_pl(row):
    """Phân loại sức khỏe I..V từ 5 cột đánh dấu 'x' (idx 11..15)."""
    marks = [i for i in range(5) if row[11 + i] not in (None, '')]
    if len(marks) == 1:
        return marks[0] + 1, ''
    if not marks:
        return None, 'Nguồn TRỐNG phân loại sức khỏe'
    return max(marks) + 1, ('Nguồn đánh dấu nhiều phân loại: '
                            + ','.join('I II III IV V'.split()[i] for i in marks)
                            + ' — đã lấy mức nặng nhất')


CLS_NOT_ABNORMAL = re.compile(r'^\s*(bt|bình thường|nhịp xoang|bthg)\s*$', re.I)

def is_bat_thuong(v):
    return bool(v) and not CLS_NOT_ABNORMAL.match(str(v).strip())


# ======================= XỬ LÝ 1 BẢN GHI =======================
def build_record(row, stt):
    """row: tuple 44 cột từ tong-hop.xlsx (bắt đầu từ dòng 3)."""
    notes = []

    cccd, n = clean_cccd(row[6]);            notes += [n] if n else []
    ngay_sinh, n = clean_nam_sinh(row[4], row[43]); notes += [n] if n else []
    pl_chung, n = parse_pl(row);             notes += [n] if n else []

    # --- chẩn đoán -> findings ---
    raw_dx = row[37] or ''
    findings = []
    for atom in process_cell(raw_dx):
        ck = concept_key(atom)
        if not ck:
            continue
        m = map_concept(ck)
        if m['nguon'] == 'unmapped':
            notes.append(f'Chưa ánh xạ được ICD: "{atom}"')
            continue
        findings.append({'atom': atom, 'concept': ck, **m})

    res = classify_person(findings, pl_chung)
    notes += res['canh_bao']

    # --- kết quả cận lâm sàng ghép vào cơ quan tương ứng ---
    organ_text = dict(res['organ_text'])
    dtim, sieuam = row[41], row[42]
    if is_bat_thuong(dtim):
        cur = organ_text['TH']
        add = 'Điện tim: ' + basic_clean(dtim)
        organ_text['TH'] = add if cur == BINH_THUONG else cur + '; ' + add
    if is_bat_thuong(sieuam):
        cur = organ_text['TIEUHOA']
        add = 'Siêu âm ổ bụng: ' + basic_clean(sieuam)
        organ_text['TIEUHOA'] = add if cur == BINH_THUONG else cur + '; ' + add

    bc = res['benh_chinh']
    bk = res['benh_kem']
    # khử trùng lặp mã bệnh kèm, bỏ mã trùng bệnh chính
    seen = {bc['icd']} if bc else set()
    bk_u = []
    for f in bk:
        if f['icd'] not in seen:
            seen.add(f['icd'])
            bk_u.append(f)

    rec = {
        'TT': stt,
        'MA_CSKCB': CFG['MA_CSKCB'],
        'MA_GTIN_CSKCB': CFG['MA_GTIN_CSKCB'],
        'NGAY_VAO': fmt_date(row[2]),
        'HO_TEN': basic_clean(row[3]).upper(),
        'GIOI_TINH': clean_gioi(row[5]),
        'NGAY_SINH': ngay_sinh,
        'MA_DAN_TOC': CFG['DAN_TOC'],
        'SO_CCCD': cccd,
        'NGAYCAP_CCCD': fmt_date(row[7]),
        'NOICAP_CCCD': clean_noi_cap(row[8]),
        'MATINH_CU_TRU': CFG['TINH'],
        'MAXA_CU_TRU': XA_MAP.get(basic_clean(row[0]), basic_clean(row[0])),
        'DIA_CHI': basic_clean(row[10]),
        'MA_NGHE_NGHIEP': CFG['NGHE_NGHIEP'],
        'DOI_TUONG': CFG['DOI_TUONG'],
        'NGUON_CHI_TRA': CFG['NGUON_CHI_TRA'],
        'LY_DO_VV': CFG['LY_DO'],
        'MA_LOAI_KCB': CFG['LOAI_KCB'],
        'PHAN_LOAI_SK': res['pl_chung'],
        # tên bệnh phải lấy NGUYÊN VĂN từ danh mục BYT để qua được dropdown
        'KET_LUAN_BENH': ten_chinh_thuc(bc['icd'], bc['ten_icd']) if bc else '',
        'CAC_BENH_TAT_NEU_CO': basic_clean(raw_dx),
        # --- cột bổ sung (ngoài mẫu BYT) ---
        '_MA_BENH_CHINH': bc['icd'] if bc else '',
        '_MA_BENH_KEM': ';'.join(f['icd'] for f in bk_u),
        '_TEN_BENH_KEM': ';'.join(ten_chinh_thuc(f['icd'], f['ten_icd'])
                                  for f in bk_u),
        '_CO_QUAN_BENH_CHINH': bc['co_quan'] if bc else '',
        '_GHI_CHU': ' | '.join(dict.fromkeys(notes)),
    }
    for o in ORGANS:
        col_txt, col_pl = ORGAN_COLS[o]
        rec[col_txt] = organ_text[o]
        rec[col_pl] = res['organ_class'][o]

    # cận lâm sàng (glucose) -> sheet DM_CanLamSang
    cls = []
    if row[40] not in (None, ''):
        try:
            gv = float(str(row[40]).replace(',', '.'))
            thoi_diem = basic_clean(row[39])
            nguong = 7.0 if thoi_diem.lower().startswith('đói') else 11.1
            cls.append({
                'HO_TEN': rec['HO_TEN'], 'SO_CCCD': cccd,
                'MA_CHI_SO': '- Glucose', 'MA_DICH_VU': 'S01',
                'DON_VI_DO': 'mmol/ L', 'GIA_TRI': gv,
                'MO_TA': f'Đường máu mao mạch ({thoi_diem})',
                'KET_LUAN': 'Cao hơn bình thường' if gv >= nguong else 'Bình thường',
            })
        except ValueError:
            notes.append(f'Giá trị glucose không hợp lệ: {row[40]}')
            rec['_GHI_CHU'] = ' | '.join(dict.fromkeys(notes))

    return rec, cls


# ======================= GHI FILE =======================
HDR_FILL = PatternFill('solid', fgColor='DDEBF7')
EXTRA_FILL = PatternFill('solid', fgColor='FFF2CC')

_ICD_OFFICIAL = None

def icd_official():
    """
    {MÃ ICD -> TÊN CHÍNH THỨC} lấy từ sheet dmicdme của file mẫu BYT.

    BẮT BUỘC dùng tên này cho các cột có ràng buộc dropdown (KET_LUAN_BENH...),
    vì cổng BYT đối chiếu nguyên văn. Tên tự đặt dù đúng về y học vẫn bị từ chối.
    """
    global _ICD_OFFICIAL
    if _ICD_OFFICIAL is None:
        wb = openpyxl.load_workbook(TPL, data_only=True, read_only=True)
        _ICD_OFFICIAL = {}
        for r in wb['dmicdme'].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]:
                ma = str(r[0]).strip().upper()
                ten = str(r[1]).strip()
                _ICD_OFFICIAL[ma] = ten
                # Danh mục BYT dùng quy ước dagger/asterisk của ICD-10: một số
                # mã có hậu tố † hoặc * (vd 'E11.4†' = ĐTĐ có biến chứng thần
                # kinh). Tra theo mã trần 'E11.4' sẽ TRƯỢT nếu không đánh chỉ
                # mục cả dạng đã bỏ ký hiệu.
                tran = ma.rstrip('†*✝ ').strip()
                if tran and tran not in _ICD_OFFICIAL:
                    _ICD_OFFICIAL[tran] = ten
        wb.close()
    return _ICD_OFFICIAL


def ten_chinh_thuc(code, fallback=''):
    return icd_official().get(str(code).strip().upper(), fallback)


def read_template_headers():
    wb = openpyxl.load_workbook(TPL, data_only=True, read_only=True)
    ws = wb['Trên 18']
    r1 = [ws.cell(1, c).value for c in range(1, 104)]
    r2 = [ws.cell(2, c).value for c in range(1, 104)]
    wb.close()
    return r1, r2


def write_import(recs, cls_rows, path):
    r1, r2 = read_template_headers()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Trên 18'

    EXTRA = [('_MA_BENH_CHINH', 'Mã bệnh chính (ICD-10)'),
             ('_MA_BENH_KEM', 'Mã bệnh kèm theo (cách nhau bởi ";")'),
             ('_TEN_BENH_KEM', 'Tên bệnh kèm theo (cách nhau bởi ";")'),
             ('_CO_QUAN_BENH_CHINH', 'Cơ quan có bệnh chính'),
             ('_GHI_CHU', 'Ghi chú rà soát')]

    for c in range(1, 104):
        ws.cell(1, c, r1[c - 1]); ws.cell(2, c, r2[c - 1])
    for i, (k, lbl) in enumerate(EXTRA):
        c = 104 + i
        ws.cell(1, c, lbl); ws.cell(2, c, k)
        ws.cell(1, c).fill = EXTRA_FILL; ws.cell(2, c).fill = EXTRA_FILL
    for c in range(1, 104):
        ws.cell(1, c).fill = HDR_FILL; ws.cell(2, c).fill = HDR_FILL
    for c in range(1, 104 + len(EXTRA)):
        ws.cell(1, c).font = Font(bold=True, size=9)
        ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(2, c).font = Font(bold=True, size=8, color='7F7F7F')
    ws.freeze_panes = 'C3'
    ws.row_dimensions[1].height = 46

    # dòng 3 = dòng dữ liệu đầu tiên (mẫu BYT có dòng 3 là hướng dẫn; ta bỏ
    # dòng hướng dẫn để file sạch, cổng đọc theo mã trường ở dòng 2)
    TEXT_COLS = {'SO_CCCD', 'NGAY_VAO', 'NGAY_SINH', 'NGAYCAP_CCCD'}
    for i, rec in enumerate(recs):
        r = 3 + i
        for c in range(1, 104):
            key = r2[c - 1]
            if key and key in rec:
                v = rec[key]
                cell = ws.cell(r, c, v)
                if key in TEXT_COLS:
                    cell.number_format = '@'
            elif c == 1:
                ws.cell(r, c, rec['TT'])
        for j, (k, _) in enumerate(EXTRA):
            ws.cell(r, 104 + j, rec.get(k, ''))

    for c in range(1, 104 + len(EXTRA)):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ---- sheet cận lâm sàng ----
    w2 = wb.create_sheet('DM_CanLamSang')
    h1 = ['TT', 'Họ tên', 'Số CMND/CCCD/Mã định danh', 'Tên chỉ số cận lâm sàng',
          'DV Cận lâm sàng', 'Đơn vị đo chỉ số cận lâm sàng',
          'Ghi giá trị chỉ số cận lâm sàng', 'Mô tả kết quả chỉ số cận lâm sàng',
          'Kết luận chỉ số cận lâm sàng']
    h2 = [None, 'HO_TEN', 'SO_CCCD', 'MA_CHI_SO', 'MA_DICH_VU', 'DON_VI_DO',
          'GIA_TRI', 'MO_TA', 'KET_LUAN']
    for c, (a, b) in enumerate(zip(h1, h2), 1):
        w2.cell(1, c, a).font = Font(bold=True, size=9)
        w2.cell(1, c).fill = HDR_FILL
        w2.cell(2, c, b).font = Font(bold=True, size=8, color='7F7F7F')
        w2.column_dimensions[get_column_letter(c)].width = 20
    for i, cr in enumerate(cls_rows):
        r = 3 + i
        w2.cell(r, 1, i + 1)
        for c, k in enumerate(h2[1:], 2):
            cell = w2.cell(r, c, cr[k])
            if k == 'SO_CCCD':
                cell.number_format = '@'
    w2.freeze_panes = 'A3'

    wb.save(path)


def write_doichieu(pairs, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Đối chiếu'
    hdr = ['TT', 'Họ tên', 'CCCD', 'PL SK', 'CHUỖI CHẨN ĐOÁN GỐC',
           'Bệnh chính (ICD)', 'Mã BC', 'Cơ quan', 'Bệnh kèm theo (ICD)', 'Mã BK',
           'Tuần hoàn', 'PL', 'Hô hấp', 'PL', 'Tiêu hóa', 'PL',
           'Thận-TN-SD', 'PL', 'Nội tiết', 'PL', 'Cơ-xương-khớp', 'PL',
           'Thần kinh', 'PL', 'Tâm thần', 'PL', 'Ngoại khoa', 'PL',
           'Da liễu', 'PL', 'Sản phụ khoa', 'PL', 'Mắt', 'PL',
           'Tai-Mũi-Họng', 'PL', 'Răng-Hàm-Mặt', 'PL', 'Ghi chú rà soát']
    for c, h in enumerate(hdr, 1):
        ws.cell(1, c, h).font = Font(bold=True, size=9)
        ws.cell(1, c).fill = HDR_FILL
        ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['E'].width = 55
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['F'].width = 34
    ws.column_dimensions['I'].width = 50
    for c in [11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31, 33, 35, 37]:
        ws.column_dimensions[get_column_letter(c)].width = 30
    ws.column_dimensions[get_column_letter(39)].width = 45
    ws.freeze_panes = 'F2'
    for i, rec in enumerate(pairs):
        r = 2 + i
        vals = [rec['TT'], rec['HO_TEN'], rec['SO_CCCD'], rec['PHAN_LOAI_SK'],
                rec['CAC_BENH_TAT_NEU_CO'], rec['KET_LUAN_BENH'],
                rec['_MA_BENH_CHINH'], rec['_CO_QUAN_BENH_CHINH'],
                rec['_TEN_BENH_KEM'], rec['_MA_BENH_KEM']]
        for o in ORGANS:
            t, p = ORGAN_COLS[o]
            vals += [rec[t], rec[p]]
        vals.append(rec['_GHI_CHU'])
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c == 3:
                cell.number_format = '@'
    wb.save(path)


def write_rasoat(bad, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Cần rà soát'
    hdr = ['TT', 'Họ tên', 'CCCD', 'Ngày khám', 'Xã', 'Thôn',
           'Chuỗi chẩn đoán gốc', 'Lý do cần rà soát']
    for c, h in enumerate(hdr, 1):
        ws.cell(1, c, h).font = Font(bold=True); ws.cell(1, c).fill = HDR_FILL
    for w, c in zip([6, 22, 16, 12, 16, 22, 50, 55], range(1, 9)):
        ws.column_dimensions[get_column_letter(c)].width = w
    for i, b in enumerate(bad):
        r = 2 + i
        for c, v in enumerate(b, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c == 3:
                cell.number_format = '@'
    ws.freeze_panes = 'A2'
    wb.save(path)


# ======================= MAIN =======================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sample', type=int, default=0)
    ap.add_argument('--all', action='store_true')
    ap.add_argument('--seed', type=int, default=20260719)
    a = ap.parse_args()
    if not a.all and not a.sample:
        a.sample = 100
    os.makedirs(OUT, exist_ok=True)

    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    rows = list(wb['TONG HOP FULL'].iter_rows(min_row=3, values_only=True))
    wb.close()
    rows = [r for r in rows if r[3]]           # phải có họ tên
    print(f'Đọc {len(rows)} bản ghi từ tong-hop.xlsx')

    if a.all:
        sel, mode = rows, 'TOANBO'
    else:
        # lấy mẫu PHÂN TẦNG: theo xã × phân loại sức khỏe, ưu tiên ca nhiều bệnh
        random.seed(a.seed)
        strata = collections.defaultdict(list)
        for r in rows:
            pl, _ = parse_pl(r)
            strata[(basic_clean(r[0]), pl)].append(r)
        sel, keys = [], sorted(strata.keys(), key=lambda k: str(k))
        i = 0
        while len(sel) < a.sample:
            k = keys[i % len(keys)]
            if strata[k]:
                sel.append(strata[k].pop(random.randrange(len(strata[k]))))
            i += 1
            if i > 10000:
                break
        # đảm bảo có ca khó: nhiều bệnh nhất + ca ghi chú lạ
        hard = sorted(rows, key=lambda r: -(len(str(r[37] or ''))))[:8]
        for h in hard:
            if h not in sel:
                sel[random.randrange(len(sel))] = h
        mode = f'MAU{a.sample}'

    recs, cls_rows, bad = [], [], []
    for i, r in enumerate(sel, 1):
        rec, cls = build_record(r, i)
        recs.append(rec); cls_rows += cls
        gc = rec['_GHI_CHU']
        if 'THIẾU CCCD' in gc or 'KHÔNG HỢP LỆ' in gc or 'KHÔNG XÁC ĐỊNH' in gc:
            bad.append([rec['TT'], rec['HO_TEN'], rec['SO_CCCD'],
                        rec['NGAY_VAO'], rec['MAXA_CU_TRU'], rec['DIA_CHI'],
                        rec['CAC_BENH_TAT_NEU_CO'], gc])

    # phát hiện CCCD trùng
    dup = [k for k, v in collections.Counter(
        r['SO_CCCD'] for r in recs if r['SO_CCCD']).items() if v > 1]
    for rec in recs:
        if rec['SO_CCCD'] in dup:
            rec['_GHI_CHU'] = (rec['_GHI_CHU'] + ' | ' if rec['_GHI_CHU'] else '') \
                              + 'CCCD TRÙNG với bản ghi khác'
            bad.append([rec['TT'], rec['HO_TEN'], rec['SO_CCCD'], rec['NGAY_VAO'],
                        rec['MAXA_CU_TRU'], rec['DIA_CHI'],
                        rec['CAC_BENH_TAT_NEU_CO'], 'CCCD TRÙNG'])

    p1 = os.path.join(OUT, f'KSK_Import_{mode}.xlsx')
    p2 = os.path.join(OUT, f'KSK_DoiChieu_{mode}.xlsx')
    p3 = os.path.join(OUT, f'CAN_RA_SOAT_{mode}.xlsx')
    write_import(recs, cls_rows, p1)
    write_doichieu(recs, p2)
    write_rasoat(bad, p3)

    # ---- thống kê ----
    print(f'\n✔ {len(recs)} bản ghi | {len(cls_rows)} dòng cận lâm sàng '
          f'| {len(bad)} ca cần rà soát')
    pl = collections.Counter(r['PHAN_LOAI_SK'] for r in recs)
    print('Phân loại SK:', dict(sorted(pl.items(), key=lambda x: str(x[0]))))
    cq = collections.Counter(r['_CO_QUAN_BENH_CHINH'] for r in recs)
    print('Cơ quan bệnh chính:', cq.most_common())
    nobc = sum(1 for r in recs if not r['_MA_BENH_CHINH'])
    print(f'Không xác định được bệnh chính: {nobc}')
    print('\nFile:\n ', p1, '\n ', p2, '\n ', p3)


if __name__ == '__main__':
    main()
