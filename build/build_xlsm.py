# -*- coding: utf-8 -*-
"""
build_xlsm.py — Xuất THẲNG ra file .xlsm đúng chuẩn mẫu Bộ Y tế.

Cách làm: mở BẢN SAO của file mẫu (keep_vba=True) rồi ghi dữ liệu vào từ dòng 4.
Nhờ vậy giữ nguyên 100%: data validation (dropdown), định dạng cột, cột ẩn M/O,
các sheet danh mục (dmicdme, dmtinh...) và macro VBA.

Chạy:
    python3 build_xlsm.py --sample 100     # 100 ca ĐẦY ĐỦ THÔNG TIN NHẤT
    python3 build_xlsm.py --all
"""
import argparse, os, re, sys, collections, datetime, shutil, gc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as L

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from normalize import process_cell, concept_key, basic_clean
from icd_map import map_concept
from classify import classify_person, ORGAN_COLS, ORGANS, BINH_THUONG, severity_of
import tien_su
import mapper
from build_import import (CFG, XA_MAP, fmt_date, clean_cccd, clean_gioi,
                          clean_nam_sinh, clean_noi_cap, parse_pl, is_bat_thuong,
                          icd_official, ten_chinh_thuc, SRC, TPL, OUT, BASE,
                          write_doichieu, write_rasoat)

FIRST_ROW = 4          # mẫu BYT: dòng 1 tiêu đề, 2 mã trường, 3 hướng dẫn


def build_record(row, stt, gidx=None):
    """
    Sinh 1 bản ghi (dict theo mã trường) từ 1 dòng tong-hop.xlsx.

    stt  : số thứ tự TRONG FILE đang xuất (cột TT của mẫu BYT)
    gidx : vị trí TOÀN CỤC trong tong-hop.xlsx — dùng cho MA_HO_SO để mã hồ sơ
           luôn giống nhau dù xuất toàn bộ hay xuất tách theo xã. Nếu lấy theo
           stt thì mỗi file xã sẽ đánh lại từ 1 -> trùng mã và lệch với file
           dữ liệu quản lý.
    """
    notes = []
    cccd, n = clean_cccd(row[6]);                    notes += [n] if n else []
    ngay_sinh, n = clean_nam_sinh(row[4], row[43]);  notes += [n] if n else []
    pl_chung, n = parse_pl(row);                     notes += [n] if n else []

    raw_dx = row[37] or ''
    # mapper: từ điển anh Khôi -> rule -> khớp mờ -> theo cơ quan
    findings, thi_luc, dx_notes = mapper.phan_tich(raw_dx)
    notes += dx_notes
    for f in findings:
        f['_sev'] = severity_of(f['icd'])

    res = classify_person(findings, pl_chung)
    notes += res['canh_bao']

    organ_text = dict(res['organ_text'])
    if is_bat_thuong(row[41]):
        cur = organ_text['TH']; add = 'Điện tim: ' + basic_clean(row[41])
        organ_text['TH'] = add if cur == BINH_THUONG else cur + '; ' + add
    if is_bat_thuong(row[42]):
        cur = organ_text['TIEUHOA']; add = 'Siêu âm ổ bụng: ' + basic_clean(row[42])
        organ_text['TIEUHOA'] = add if cur == BINH_THUONG else cur + '; ' + add

    bc, bk = res['benh_chinh'], res['benh_kem']
    seen = {bc['icd']} if bc else set()
    bk_u = []
    for f in bk:
        if f['icd'] not in seen:
            seen.add(f['icd']); bk_u.append(f)

    rec = {
        'MA_CSKCB': CFG['MA_CSKCB'],
        'MA_GTIN_CSKCB': CFG['MA_GTIN_CSKCB'],
        'NGAY_VAO': fmt_date(row[2]),
        'HO_TEN': basic_clean(row[3]).upper(),
        'GIOI_TINH': clean_gioi(row[5]),
        'NGAY_SINH': ngay_sinh,
        'MA_DAN_TOC': CFG['DAN_TOC'],
        'SO_CCCD': cccd,
        'NGAYCAP_CCCD': fmt_date(row[7]),
        'NOICAP_CCCD': clean_noi_cap(row[8]),
        'MATINH_CU_TRU': CFG['TINH'],
        'MAXA_CU_TRU': XA_MAP.get(basic_clean(row[0]), basic_clean(row[0])),
        'DIA_CHI': basic_clean(row[10]),
        'MA_NGHE_NGHIEP': CFG['NGHE_NGHIEP'],
        'DOI_TUONG': CFG['DOI_TUONG'],
        'NGUON_CHI_TRA': CFG['NGUON_CHI_TRA'],
        'NOI_LAM_VIEC_HOC_TAP': '',
        'LY_DO_VV': CFG['LY_DO'],
        'MA_LOAI_KCB': CFG['LOAI_KCB'],
        # thị lực trích từ chuỗi chẩn đoán (dạng x/10 đúng danh mục DM_KQmat)
        'KHONG_KINH_MAT_PHAI': thi_luc['mp'],
        'KHONG_KINH_MAT_TRAI': thi_luc['mt'],
        'PHAN_LOAI_SK': res['pl_chung'],
        'KET_LUAN_BENH': ten_chinh_thuc(bc['icd'], bc['ten_icd']) if bc else '',
        'CAC_BENH_TAT_NEU_CO': basic_clean(raw_dx),
        '_MA_BENH_CHINH': bc['icd'] if bc else '',
        '_MA_BENH_KEM': ';'.join(f['icd'] for f in bk_u),
        '_TEN_BENH_KEM': ';'.join(ten_chinh_thuc(f['icd'], f['ten_icd'])
                                  for f in bk_u),
        '_CO_QUAN_BENH_CHINH': bc['co_quan'] if bc else '',
    }
    # --- TIỀN SỬ (cột Y, AA..AY) ---
    rec.update(tien_su.suy_luan(findings, bc, ten_chinh_thuc))

    for o in ORGANS:
        ct, cp = ORGAN_COLS[o]
        rec[ct] = organ_text[o]
        rec[cp] = res['organ_class'][o]

    rec['_GHI_CHU'] = ' | '.join(dict.fromkeys(notes))
    rec['TT'] = stt
    # --- dữ liệu phục vụ app quản lý (không đưa vào file .xlsm nộp Bộ) ---
    rec['_MA_HO_SO'] = (f"{CFG['MA_CSKCB']}-{CFG['NAM_KHAM']}-"
                        f"{(gidx if gidx is not None else stt):05d}")
    rec['_XA_NGUON'] = basic_clean(row[0])
    rec['_THON'] = basic_clean(row[10])
    rec['_NAM_SINH_NGUON'] = row[4]
    rec['_TUOI'] = row[43]
    rec['_GLU_THOI_DIEM'] = basic_clean(row[39])
    rec['_GLU_GIA_TRI'] = row[40]
    rec['_DIEN_TIM'] = basic_clean(row[41])
    rec['_SIEU_AM'] = basic_clean(row[42])
    rec['_TONG_SO_BENH'] = row[38]
    rec['_FINDINGS'] = findings
    rec['_ICD_CHINH_OBJ'] = bc
    rec['_THI_LUC_MP'] = thi_luc['mp']
    rec['_THI_LUC_MT'] = thi_luc['mt']

    # ---- CỜ KIỂM SOÁT CHẤT LƯỢNG cho ứng dụng quản lý ----
    co = []
    if not cccd:
        co.append('THIEU_CCCD')
    if 'Ngày/tháng ước lượng' in rec['_GHI_CHU']:
        co.append('NGAY_SINH_UOC_LUONG')
    if 'Năm sinh nguồn không hợp lệ' in rec['_GHI_CHU']:
        co.append('NAM_SINH_SAI_NGUON')
    if any(f['nguon'] == 'fuzzy' for f in findings):
        co.append('ICD_MAY_TU_SUA_LOI_GO')
    if any(f['nguon'] == 'organ_fallback' for f in findings):
        co.append('ICD_KHONG_DAC_HIEU')
    if 'Chưa ánh xạ được ICD' in rec['_GHI_CHU'] or \
       'Chưa gán được mã ICD' in rec['_GHI_CHU']:
        co.append('CON_CHAN_DOAN_CHUA_ANH_XA')
    if 'không ghi rõ bên mắt' in rec['_GHI_CHU']:
        co.append('THI_LUC_CHUA_RO_BEN_MAT')
    if not findings and pl_chung and pl_chung > 1:
        co.append('CO_PHAN_LOAI_NHUNG_KHONG_CO_CHAN_DOAN')
    if 'nhiều phân loại' in rec['_GHI_CHU']:
        co.append('NGUON_DANH_DAU_NHIEU_PHAN_LOAI')
    rec['_CO_QC'] = ';'.join(co)
    rec['_CAN_RA_SOAT'] = 'CÓ' if co else ''
    rec['_SO_LOI'] = len(co)
    return rec


def diem_day_du(row):
    """Điểm mức độ đầy đủ thông tin của 1 dòng nguồn — để chọn 100 ca mẫu."""
    d = 0
    cccd, note = clean_cccd(row[6])
    if cccd and len(cccd) == 12 and not note:
        d += 4
    if row[7]:  d += 2                       # ngày cấp CCCD
    if row[8]:  d += 1                       # nơi cấp
    if row[9]:  d += 1                       # đối tượng
    if row[10]: d += 1                       # thôn
    if row[2]:  d += 1                       # ngày khám
    pl, n = parse_pl(row)
    if pl and not n: d += 2
    dx = str(row[37] or '').strip()
    if dx:
        d += 2
        d += min(len(process_cell(dx)), 5)   # càng nhiều bệnh càng đủ thông tin
    if row[41]: d += 1                       # điện tim
    if row[42]: d += 1                       # siêu âm
    try:
        y = int(float(row[4]))
        if 1900 <= y <= 2010: d += 2
    except (TypeError, ValueError):
        pass
    return d


def write_xlsm(recs, path):
    """Ghi vào bản sao file mẫu .xlsm — giữ nguyên validation/định dạng/VBA."""
    shutil.copyfile(TPL, path)
    wb = openpyxl.load_workbook(path, keep_vba=True)
    ws = wb['Trên 18']

    code2col = {}
    for c in range(1, 104):
        k = ws.cell(2, c).value
        if k:
            code2col[str(k).strip()] = c

    # KHÔNG thêm bất kỳ cột nào ngoài 103 cột của mẫu — cổng BYT chỉ nhận đúng
    # cấu trúc mẫu. Toàn bộ trường bổ sung nằm ở file KSK_DuLieuQuanLy_*.xlsx.
    TEXT_COLS = {'SO_CCCD', 'NGAY_VAO', 'NGAY_SINH', 'NGAYCAP_CCCD'}
    # công thức cột phụ trợ M (TINH_ID). Cột O là lookup cấp huyện cũ, đã hỏng
    # (#REF!) sau khi bỏ cấp huyện — để trống, cột này không có mã trường ở
    # dòng 2 nên không thuộc dữ liệu xuất.
    for i, rec in enumerate(recs):
        r = FIRST_ROW + i
        ws.cell(r, 1, rec['TT'])
        for code, c in code2col.items():
            if code in rec:
                v = rec[code]
                cell = ws.cell(r, c, v if v != '' else None)
                if code in TEXT_COLS and v:
                    cell.number_format = '@'
        ws.cell(r, 13, f'=INDEX(dmtinh!$A:$B,MATCH(L{r},dmtinh!$A:$A,0),2)')
        ws.cell(r, 15, None)

    # xóa các dòng mẫu thừa còn lại của template (nếu dữ liệu ít hơn)
    last = FIRST_ROW + len(recs)
    for r in range(last, ws.max_row + 1):
        for c in range(1, 110):
            ws.cell(r, c).value = None

    # sheet cận lâm sàng: KHÔNG dùng — đợt khám dùng đường máu MAO MẠCH,
    # còn danh mục BYT chỉ có Glucose máu tĩnh mạch (máy sinh hóa) => không
    # đưa kết quả vào để tránh sai lệch bản chất xét nghiệm.
    w2 = wb['DM_CanLamSang']
    for r in range(3, w2.max_row + 1):
        for c in range(1, 10):
            w2.cell(r, c).value = None

    wb.save(path)
    # Giải phóng tường minh: openpyxl giữ tham chiếu toàn bộ ô của template
    # (kèm sheet dmicdme 35.735 dòng). Không dọn thì bộ nhớ tích lũy qua từng
    # xã và tiến trình bị hệ điều hành kết thúc giữa chừng.
    wb.close()
    del wb, ws, w2
    gc.collect()


TEN_CQ = {
    'TH': 'Tuần hoàn', 'HH': 'Hô hấp', 'TIEUHOA': 'Tiêu hóa',
    'THAN': 'Thận - Tiết niệu - Sinh dục', 'NOITIET': 'Nội tiết',
    'CXK': 'Cơ - Xương - Khớp', 'TK': 'Thần kinh', 'TT': 'Tâm thần',
    'NGOAI': 'Ngoại khoa', 'DALIEU': 'Da liễu', 'SAN': 'Sản phụ khoa',
    'MAT': 'Mắt', 'TMH': 'Tai - Mũi - Họng', 'RHM': 'Răng - Hàm - Mặt',
}
LOAI_SK = {1: 'I - Rất khỏe', 2: 'II - Khỏe', 3: 'III - Trung bình',
           4: 'IV - Yếu', 5: 'V - Rất yếu'}


# ---- định nghĩa cột dùng chung cho cả 2 chế độ ghi ----
COLS_BENH_NHAN = [
    ('MA_HO_SO', 22), ('TT', 6), ('HO_TEN', 24), ('GIOI_TINH', 10),
    ('NGAY_SINH', 12), ('NAM_SINH', 10), ('TUOI', 6), ('SO_CCCD', 16),
    ('NGAYCAP_CCCD', 13), ('NOICAP_CCCD', 24), ('NGAY_KHAM', 12),
    ('XA_PHUONG', 22), ('THON_TDP', 22), ('DOI_TUONG', 18),
    ('PHAN_LOAI_SK', 8), ('PHAN_LOAI_SK_CHU', 16),
    ('ICD_BENH_CHINH', 14), ('TEN_BENH_CHINH', 40), ('CO_QUAN_BENH_CHINH', 22),
    ('ICD_BENH_KEM', 30), ('TEN_BENH_KEM', 60),
    ('SO_LUONG_BENH', 8), ('TONG_SO_BENH_NGUON', 10),
    ('GLU_THOI_DIEM', 12), ('GLU_GIA_TRI', 10),
    ('THI_LUC_MAT_PHAI', 12), ('THI_LUC_MAT_TRAI', 12),
    ('KQ_DIEN_TIM', 34), ('KQ_SIEU_AM_O_BUNG', 34), ('CHAN_DOAN_GOC', 60),
    ('CAN_RA_SOAT', 10), ('SO_LOI', 8), ('CO_QC', 40), ('GHI_CHU_RA_SOAT', 45),
]
COLS_BENH_CHI_TIET = [
    ('MA_HO_SO', 22), ('TT', 6), ('HO_TEN', 24), ('SO_CCCD', 16),
    ('STT_BENH', 8), ('LA_BENH_CHINH', 12), ('MA_ICD', 12), ('TEN_ICD', 45),
    ('CO_QUAN', 24), ('MUC_DO_NANG', 10), ('CHUOI_GOC', 30),
    ('KHAI_NIEM_CHUAN_HOA', 30), ('NGUON_ANH_XA', 22),
    ('DIEN_GIAI_CUA_BS', 40), ('DO_TIN_CAY', 34), ('CAN_RA_SOAT', 10),
]
COLS_PHAN_LOAI = [
    ('MA_HO_SO', 22), ('TT', 6), ('HO_TEN', 24), ('MA_CO_QUAN', 12),
    ('CO_QUAN', 26), ('PHAN_LOAI', 10), ('KET_QUA_KHAM', 60),
]
COLS_TU_DIEN = [
    ('KHAI_NIEM_CHUAN_HOA', 34), ('MA_ICD', 12), ('TEN_ICD', 45),
    ('CO_QUAN', 24), ('MUC_DO_NANG', 10), ('SO_LAN_XUAT_HIEN', 12),
    ('NGUON_ANH_XA', 22), ('DIEN_GIAI_CUA_BS', 40), ('VI_DU_CHUOI_GOC', 34),
]


def row_benh_nhan(r):
    return [r['_MA_HO_SO'], r['TT'], r['HO_TEN'], r['GIOI_TINH'],
            r['NGAY_SINH'], r['_NAM_SINH_NGUON'], r['_TUOI'], r['SO_CCCD'],
            r['NGAYCAP_CCCD'], r['NOICAP_CCCD'], r['NGAY_VAO'],
            r['MAXA_CU_TRU'], r['_THON'], r['DOI_TUONG'],
            r['PHAN_LOAI_SK'], LOAI_SK.get(r['PHAN_LOAI_SK'], ''),
            r['_MA_BENH_CHINH'], r['KET_LUAN_BENH'],
            TEN_CQ.get(r['_CO_QUAN_BENH_CHINH'], ''),
            r['_MA_BENH_KEM'], r['_TEN_BENH_KEM'],
            len(r['_FINDINGS']), r['_TONG_SO_BENH'],
            r['_GLU_THOI_DIEM'], r['_GLU_GIA_TRI'],
            r['_THI_LUC_MP'], r['_THI_LUC_MT'],
            r['_DIEN_TIM'], r['_SIEU_AM'], r['CAC_BENH_TAT_NEU_CO'],
            r['_CAN_RA_SOAT'], r['_SO_LOI'], r['_CO_QC'], r['_GHI_CHU']]


def rows_benh_chi_tiet(r):
    bc = r['_ICD_CHINH_OBJ']
    out = []
    for j, f in enumerate(r['_FINDINGS'], 1):
        out.append([r['_MA_HO_SO'], r['TT'], r['HO_TEN'], r['SO_CCCD'], j,
                    'CHÍNH' if (bc and f is bc) else 'kèm theo',
                    f['icd'], ten_chinh_thuc(f['icd'], f['ten_icd']),
                    TEN_CQ.get(f['co_quan'], ''), f.get('_sev', ''),
                    f['atom'], f['concept'], f['nguon'],
                    f.get('nghia_khoi', ''), f.get('tin_cay', ''),
                    'CÓ' if f.get('can_ra_soat') else ''])
    return out


def rows_phan_loai(r):
    out = []
    for o in ORGANS:
        ct, cp = ORGAN_COLS[o]
        out.append([r['_MA_HO_SO'], r['TT'], r['HO_TEN'], o, TEN_CQ[o],
                    r[cp], r[ct]])
    return out


def write_quanly_lon(recs, path):
    """
    Bản ghi tuần tự (write_only) cho bộ dữ liệu LỚN.

    openpyxl ở chế độ thường giữ toàn bộ ô trong RAM và tạo một object style
    riêng cho MỖI ô — với 13.326 ca × 4 sheet thì tiến trình bị hệ điều hành
    kết thúc vì hết bộ nhớ. write_only ghi thẳng từng dòng ra đĩa.
    """
    from openpyxl.cell import WriteOnlyCell
    wb = openpyxl.Workbook(write_only=True)
    hdr_font = Font(bold=True, size=9)
    hdr_fill = PatternFill('solid', fgColor='DDEBF7')

    def add_sheet(ten, cols):
        ws = wb.create_sheet(ten)
        ws.freeze_panes = 'B2'
        for i, (h, w) in enumerate(cols, 1):
            ws.column_dimensions[L(i)].width = w
        hang = []
        for h, _ in cols:
            c = WriteOnlyCell(ws, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            hang.append(c)
        ws.append(hang)
        return ws

    ws = add_sheet('BENH_NHAN', COLS_BENH_NHAN)
    for r in recs:
        ws.append(row_benh_nhan(r))

    w2 = add_sheet('BENH_CHI_TIET', COLS_BENH_CHI_TIET)
    for r in recs:
        for v in rows_benh_chi_tiet(r):
            w2.append(v)

    w3 = add_sheet('PHAN_LOAI_CO_QUAN', COLS_PHAN_LOAI)
    for r in recs:
        for v in rows_phan_loai(r):
            w3.append(v)

    w4 = add_sheet('TU_DIEN_ICD', COLS_TU_DIEN)
    td = {}
    for r in recs:
        for f in r['_FINDINGS']:
            e = td.setdefault(f['concept'], {'n': 0, **f})
            e['n'] += 1
    for ck, e in sorted(td.items(), key=lambda kv: -kv[1]['n']):
        w4.append([ck, e['icd'], ten_chinh_thuc(e['icd'], e['ten_icd']),
                   TEN_CQ.get(e['co_quan'], ''), e.get('_sev', ''), e['n'],
                   e['nguon'], e.get('nghia_khoi', ''), e['atom']])

    wb.save(path)
    n_benh = sum(len(r['_FINDINGS']) for r in recs)
    return n_benh, len(td)


def write_quanly(recs, path):
    """
    File dữ liệu cho ỨNG DỤNG QUẢN LÝ (không dùng để nộp Bộ).
    Cấu trúc quan hệ, khóa nối = MA_HO_SO.

      Sheet BENH_NHAN      : 1 dòng / người
      Sheet BENH_CHI_TIET  : 1 dòng / bệnh  -> đếm, lọc, thống kê theo ICD
      Sheet PHAN_LOAI_CO_QUAN : 1 dòng / (người × cơ quan)
      Sheet TU_DIEN_ICD    : từ điển khái niệm -> ICD đã dùng (để app mở rộng)
    """
    hdr_fill = PatternFill('solid', fgColor='DDEBF7')

    def mk(ws, cols):
        for c, (h, w) in enumerate(cols, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True, size=9)
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.column_dimensions[L(c)].width = w
        ws.freeze_panes = 'B2'

    wb = openpyxl.Workbook()

    # ---------------- BENH_NHAN ----------------
    ws = wb.active
    ws.title = 'BENH_NHAN'
    cols = [('MA_HO_SO', 22), ('TT', 6), ('HO_TEN', 24), ('GIOI_TINH', 10),
            ('NGAY_SINH', 12), ('NAM_SINH', 10), ('TUOI', 6), ('SO_CCCD', 16),
            ('NGAYCAP_CCCD', 13), ('NOICAP_CCCD', 24), ('NGAY_KHAM', 12),
            ('XA_PHUONG', 22), ('THON_TDP', 22), ('DOI_TUONG', 18),
            ('PHAN_LOAI_SK', 8), ('PHAN_LOAI_SK_CHU', 16),
            ('ICD_BENH_CHINH', 14), ('TEN_BENH_CHINH', 40),
            ('CO_QUAN_BENH_CHINH', 22),
            ('ICD_BENH_KEM', 30), ('TEN_BENH_KEM', 60),
            ('SO_LUONG_BENH', 8), ('TONG_SO_BENH_NGUON', 10),
            ('GLU_THOI_DIEM', 12), ('GLU_GIA_TRI', 10),
            ('THI_LUC_MAT_PHAI', 12), ('THI_LUC_MAT_TRAI', 12),
            ('KQ_DIEN_TIM', 34), ('KQ_SIEU_AM_O_BUNG', 34),
            ('CHAN_DOAN_GOC', 60),
            ('CAN_RA_SOAT', 10), ('SO_LOI', 8), ('CO_QC', 40),
            ('GHI_CHU_RA_SOAT', 45)]
    mk(ws, cols)
    for i, r in enumerate(recs):
        row = 2 + i
        bc = r['_ICD_CHINH_OBJ']
        vals = [r['_MA_HO_SO'], r['TT'], r['HO_TEN'], r['GIOI_TINH'],
                r['NGAY_SINH'], r['_NAM_SINH_NGUON'], r['_TUOI'], r['SO_CCCD'],
                r['NGAYCAP_CCCD'], r['NOICAP_CCCD'], r['NGAY_VAO'],
                r['MAXA_CU_TRU'], r['_THON'], r['DOI_TUONG'],
                r['PHAN_LOAI_SK'], LOAI_SK.get(r['PHAN_LOAI_SK'], ''),
                r['_MA_BENH_CHINH'], r['KET_LUAN_BENH'],
                TEN_CQ.get(r['_CO_QUAN_BENH_CHINH'], ''),
                r['_MA_BENH_KEM'], r['_TEN_BENH_KEM'],
                len(r['_FINDINGS']), r['_TONG_SO_BENH'],
                r['_GLU_THOI_DIEM'], r['_GLU_GIA_TRI'],
                r['_THI_LUC_MP'], r['_THI_LUC_MT'],
                r['_DIEN_TIM'], r['_SIEU_AM'],
                r['CAC_BENH_TAT_NEU_CO'],
                r['_CAN_RA_SOAT'], r['_SO_LOI'], r['_CO_QC'], r['_GHI_CHU']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c in (5, 8, 9, 11):
                cell.number_format = '@'

    # ---------------- BENH_CHI_TIET ----------------
    w2 = wb.create_sheet('BENH_CHI_TIET')
    mk(w2, [('MA_HO_SO', 22), ('TT', 6), ('HO_TEN', 24), ('SO_CCCD', 16),
            ('STT_BENH', 8), ('LA_BENH_CHINH', 12),
            ('MA_ICD', 12), ('TEN_ICD', 45), ('CO_QUAN', 24),
            ('MUC_DO_NANG', 10), ('CHUOI_GOC', 30), ('KHAI_NIEM_CHUAN_HOA', 30),
            ('NGUON_ANH_XA', 22), ('DIEN_GIAI_CUA_BS', 40),
            ('DO_TIN_CAY', 34), ('CAN_RA_SOAT', 10)])
    r2 = 2
    for r in recs:
        bc = r['_ICD_CHINH_OBJ']
        for j, f in enumerate(r['_FINDINGS'], 1):
            vals = [r['_MA_HO_SO'], r['TT'], r['HO_TEN'], r['SO_CCCD'], j,
                    'CHÍNH' if (bc and f is bc) else 'kèm theo',
                    f['icd'], ten_chinh_thuc(f['icd'], f['ten_icd']),
                    TEN_CQ.get(f['co_quan'], ''), f.get('_sev', ''),
                    f['atom'], f['concept'], f['nguon'],
                    f.get('nghia_khoi', ''), f.get('tin_cay', ''),
                    'CÓ' if f.get('can_ra_soat') else '']
            for c, v in enumerate(vals, 1):
                cell = w2.cell(r2, c, v)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if c == 4:
                    cell.number_format = '@'
            r2 += 1

    # ---------------- PHAN_LOAI_CO_QUAN ----------------
    w3 = wb.create_sheet('PHAN_LOAI_CO_QUAN')
    mk(w3, [('MA_HO_SO', 22), ('TT', 6), ('HO_TEN', 24), ('MA_CO_QUAN', 12),
            ('CO_QUAN', 26), ('PHAN_LOAI', 10), ('KET_QUA_KHAM', 60)])
    r3 = 2
    for r in recs:
        for o in ORGANS:
            ct, cp = ORGAN_COLS[o]
            vals = [r['_MA_HO_SO'], r['TT'], r['HO_TEN'], o, TEN_CQ[o],
                    r[cp], r[ct]]
            for c, v in enumerate(vals, 1):
                cell = w3.cell(r3, c, v)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r3 += 1

    # ---------------- TU_DIEN_ICD ----------------
    w4 = wb.create_sheet('TU_DIEN_ICD')
    mk(w4, [('KHAI_NIEM_CHUAN_HOA', 34), ('MA_ICD', 12), ('TEN_ICD', 45),
            ('CO_QUAN', 24), ('MUC_DO_NANG', 10), ('SO_LAN_XUAT_HIEN', 12),
            ('NGUON_ANH_XA', 16), ('VI_DU_CHUOI_GOC', 34)])
    td = {}
    for r in recs:
        for f in r['_FINDINGS']:
            e = td.setdefault(f['concept'], {'n': 0, **f})
            e['n'] += 1
    for i, (ck, e) in enumerate(sorted(td.items(), key=lambda kv: -kv[1]['n'])):
        vals = [ck, e['icd'], ten_chinh_thuc(e['icd'], e['ten_icd']),
                TEN_CQ.get(e['co_quan'], ''), e.get('_sev', ''), e['n'],
                'rule' if e['nguon'] == 'rule' else 'theo cơ quan', e['atom']]
        for c, v in enumerate(vals, 1):
            cell = w4.cell(2 + i, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(path)
    return r2 - 2, len(td)


def _thong_ke(recs, bad):
    print(f'\nTổng {len(recs)} bản ghi | {len(bad)} ca cần rà soát CCCD')
    print('Phân loại SK:', dict(sorted(collections.Counter(
        r['PHAN_LOAI_SK'] for r in recs).items(), key=lambda x: str(x[0]))))
    print('Cơ quan bệnh chính:', collections.Counter(
        r['_CO_QUAN_BENH_CHINH'] for r in recs).most_common())
    print('Ca có cờ rà soát:', sum(1 for r in recs if r['_CAN_RA_SOAT']))
    co = collections.Counter()
    for r in recs:
        for c in r['_CO_QC'].split(';'):
            if c:
                co[c] += 1
    print('Chi tiết cờ:', co.most_common())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sample', type=int, default=0)
    ap.add_argument('--all', action='store_true')
    # Chạy RIÊNG 1 xã trong 1 tiến trình. Cần thiết vì openpyxl không trả lại
    # bộ nhớ sau khi ghi file .xlsm (template kèm dmicdme 35.735 dòng); chạy
    # 8 xã trong cùng tiến trình sẽ bị hệ điều hành kết thúc ở xã thứ 6.
    ap.add_argument('--xa', type=str, default='')
    ap.add_argument('--chi-xlsm', action='store_true',
                    help='chỉ xuất .xlsm, bỏ qua các file phụ')
    ap.add_argument('--chi-quanly', action='store_true',
                    help='chỉ xuất file dữ liệu cho ứng dụng quản lý')
    a = ap.parse_args()
    if not a.all and not a.sample:
        a.sample = 100
    os.makedirs(OUT, exist_ok=True)

    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    rows = [r for r in wb['TONG HOP FULL'].iter_rows(min_row=3, values_only=True)
            if r[3]]
    wb.close()
    print(f'Đọc {len(rows)} bản ghi')

    # nạp từ điển anh Khôi + dựng mỏ neo cho tầng khớp mờ
    import user_dict
    from icd_map import bat_fuzzy
    tk = user_dict.thong_ke()
    print(f"Từ điển bác sĩ: {tk['co_nghia']} khái niệm có diễn giải, "
          f"{tk['bo_qua']} khái niệm bỏ qua")
    freq = collections.Counter()
    for r in rows:
        for _atom in process_cell(r[37] or ''):
            k = concept_key(_atom)
            if k:
                freq[k] += 1
    print(f'Mỏ neo khớp mờ: {bat_fuzzy(freq)} khái niệm')

    if a.xa:
        sel = [(i, r) for i, r in enumerate(rows, 1)
               if XA_MAP.get(basic_clean(r[0]), basic_clean(r[0])) == a.xa]
        mode = re.sub(r'[^\w]+', '_', a.xa).strip('_')
        print(f'Xã/phường "{a.xa}": {len(sel)} ca')
    elif a.all:
        sel, mode = rows, 'TOANBO'
    else:
        scored = sorted(rows, key=lambda r: -diem_day_du(r))
        sel, mode = scored[:a.sample], f'MAU{a.sample}'
        print(f'Điểm đầy đủ: cao nhất {diem_day_du(sel[0])}, '
              f'thấp nhất trong mẫu {diem_day_du(sel[-1])}')

    recs, bad = [], []
    for i, item in enumerate(sel, 1):
        # chế độ --xa trả về cặp (chỉ số toàn cục, dòng); các chế độ khác trả
        # thẳng dòng dữ liệu — bản thân dòng cũng là tuple nên phải kiểm tra
        # theo ĐỘ DÀI, không chỉ theo kiểu.
        if isinstance(item, tuple) and len(item) == 2 and isinstance(item[0], int):
            gidx, r = item
        else:
            gidx, r = i, item
        rec = build_record(r, i, gidx)
        recs.append(rec)
        gc = rec['_GHI_CHU']
        if 'THIẾU CCCD' in gc or 'KHÔNG HỢP LỆ' in gc or 'KHÔNG XÁC ĐỊNH' in gc:
            bad.append([rec['TT'], rec['HO_TEN'], rec['SO_CCCD'], rec['NGAY_VAO'],
                        rec['MAXA_CU_TRU'], rec['DIA_CHI'],
                        rec['CAC_BENH_TAT_NEU_CO'], gc])

    dup = [k for k, v in collections.Counter(
        r['SO_CCCD'] for r in recs if r['SO_CCCD']).items() if v > 1]
    for rec in recs:
        if rec['SO_CCCD'] in dup:
            rec['_GHI_CHU'] = (rec['_GHI_CHU'] + ' | ' if rec['_GHI_CHU'] else '') \
                              + 'CCCD TRÙNG với bản ghi khác'
            bad.append([rec['TT'], rec['HO_TEN'], rec['SO_CCCD'], rec['NGAY_VAO'],
                        rec['MAXA_CU_TRU'], rec['DIA_CHI'],
                        rec['CAC_BENH_TAT_NEU_CO'], 'CCCD TRÙNG'])

    p1 = os.path.join(OUT, f'KSK_Import_{mode}.xlsm')
    p2 = os.path.join(OUT, f'KSK_DoiChieu_{mode}.xlsx')
    p3 = os.path.join(OUT, f'CAN_RA_SOAT_{mode}.xlsx')
    p4 = os.path.join(OUT, f'KSK_DuLieuQuanLy_{mode}.xlsx')

    if a.xa:
        os.makedirs(os.path.join(OUT, 'import_theo_xa'), exist_ok=True)
        fp = os.path.join(OUT, 'import_theo_xa', f'KSK_Import_{mode}.xlsm')
        for i, r in enumerate(recs, 1):
            r['TT'] = i
        write_xlsm(recs, fp)
        print(f'✔ {len(recs)} ca -> {fp}')
        return

    if a.chi_xlsm:
        write_xlsm(recs, p1)
        print(f'✔ {len(recs)} ca -> {p1}')
        return

    if a.chi_quanly:
        n_benh, n_td = write_quanly_lon(recs, p4)
        write_rasoat(bad, p3)
        print(f'✔ {len(recs)} ca | {n_benh} dòng bệnh | từ điển {n_td}')
        print(f'✔ {p4}')
        _thong_ke(recs, bad)
        return

    if len(recs) > 1500:
        # Bộ dữ liệu lớn: KHÔNG gộp vào 1 file .xlsm.
        # Lý do: openpyxl phải giữ toàn bộ template (kèm sheet dmicdme 35.735
        # dòng) cùng 13.326 × 103 ô trong RAM -> vượt bộ nhớ máy.
        # Tách theo XÃ cũng thuận tiện khi nộp và khi sửa lỗi theo địa bàn.
        # Gọi lại chính script này, MỖI XÃ MỘT TIẾN TRÌNH, để hệ điều hành thu
        # hồi bộ nhớ sau từng file (openpyxl không tự trả lại).
        import subprocess
        os.makedirs(os.path.join(OUT, 'import_theo_xa'), exist_ok=True)
        xas = sorted({r['MAXA_CU_TRU'] for r in recs})
        print(f'\nXuất file nộp theo {len(xas)} xã/phường '
              f'(mỗi xã một tiến trình riêng):')
        p1 = []
        me = os.path.abspath(__file__)
        for xa in xas:
            r = subprocess.run([sys.executable, '-u', me, '--xa', xa],
                               capture_output=True, text=True)
            dong = [x for x in r.stdout.split('\n') if x.startswith('✔')]
            print('  ', dong[0] if dong else f'LỖI ở {xa}: {r.stderr[-300:]}')
            ten = re.sub(r'[^\w]+', '_', xa).strip('_')
            p1.append(os.path.join(OUT, 'import_theo_xa',
                                   f'KSK_Import_{ten}.xlsm'))
    else:
        write_xlsm(recs, p1)
    write_rasoat(bad, p3)
    if len(recs) > 1000:
        # bộ dữ liệu lớn: ghi tuần tự để không hết bộ nhớ; bỏ file đối chiếu
        # vì nội dung đã nằm đầy đủ trong file quản lý
        n_benh, n_td = write_quanly_lon(recs, p4)
        p2 = '(bỏ qua — dùng file quản lý)'
    else:
        write_doichieu(recs, p2)
        n_benh, n_td = write_quanly(recs, p4)

    print(f'\n✔ {len(recs)} bản ghi | {len(bad)} ca cần rà soát')
    print(f'✔ File quản lý: {n_benh} dòng bệnh chi tiết | từ điển {n_td} khái niệm')
    print('Phân loại SK:', dict(sorted(collections.Counter(
        r['PHAN_LOAI_SK'] for r in recs).items(), key=lambda x: str(x[0]))))
    print('Cơ quan bệnh chính:', collections.Counter(
        r['_CO_QUAN_BENH_CHINH'] for r in recs).most_common())
    print('Tiền sử 5 năm = Có:', sum(1 for r in recs
                                     if r['TSBT_BENH_TRONG_5_NAM_QUA'] == 'Có'))
    print('\nFile:')
    if isinstance(p1, list):
        print(f'  {len(p1)} file .xlsm trong {os.path.join(OUT, "import_theo_xa")}')
    else:
        print(' ', p1)
    print(' ', p2, '\n ', p3, '\n ', p4)


if __name__ == '__main__':
    main()
